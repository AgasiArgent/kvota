"""
Financial Analytics & Reporting API

Admin-only analytics endpoints with query builder, aggregations, and exports.
Uses Redis caching (10-min TTL) and parameterized SQL for security.
"""

import os
import io
import time
import logging
from typing import Dict, Any, List, Optional
from uuid import UUID
from datetime import datetime, timedelta
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, status, Request, BackgroundTasks, Query
from fastapi.responses import FileResponse
from slowapi import Limiter
from slowapi.util import get_remote_address
from supabase import create_client, Client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv

from auth import get_current_user, User, check_admin_permissions
from models import (
    SavedReportCreate,
    SavedReportUpdate,
    SavedReport,
    AnalyticsQueryRequest,
    AnalyticsQueryResponse,
    AnalyticsAggregateResponse,
)
from analytics_security import (
    build_analytics_query,
    build_aggregation_query,
    QuerySecurityValidator,
)
from analytics_cache import (
    get_cache_key,
    get_cached_report,
    cache_report,
    invalidate_report_cache,
)
from db_pool import get_db_connection, release_db_connection
from async_supabase import async_supabase_call
from routes.quotes import set_rls_context

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/analytics", tags=["analytics"])
limiter = Limiter(key_func=get_remote_address)


# ============================================================================
# TASK 5: SAVED REPORTS CRUD ENDPOINTS
# ============================================================================

@router.get("/saved-reports", response_model=List[SavedReport])
async def list_saved_reports(user: User = Depends(get_current_user)):
    """
    List all saved reports for current organization.

    Returns both personal reports (created by user) and shared reports.
    RLS ensures organization isolation.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Query with RLS - returns personal + shared reports
        query = supabase.table("saved_reports") \
            .select("*") \
            .eq("organization_id", str(user.current_organization_id)) \
            .is_("deleted_at", "null") \
            .order("created_at", desc=True)

        result = await async_supabase_call(query)

        return [SavedReport(**report) for report in result.data]

    except Exception as e:
        logger.error(f"Error listing saved reports: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to list saved reports: {str(e)}"
        )


@router.get("/saved-reports/{report_id}", response_model=SavedReport)
async def get_saved_report(
    report_id: UUID,
    user: User = Depends(get_current_user)
):
    """
    Get single saved report with versions (future).

    RLS ensures user can only access reports in their organization.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        query = supabase.table("saved_reports") \
            .select("*") \
            .eq("id", str(report_id)) \
            .eq("organization_id", str(user.current_organization_id)) \
            .is_("deleted_at", "null")

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Saved report not found"
            )

        return SavedReport(**result.data[0])

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching saved report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch saved report: {str(e)}"
        )


@router.post("/saved-reports", response_model=SavedReport)
async def create_saved_report(
    report: SavedReportCreate,
    user: User = Depends(get_current_user)
):
    """
    Create new saved report template.

    Validates fields against whitelist and stores configuration.
    Invalidates cache after creation.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        # Validate selected fields
        validated_fields = QuerySecurityValidator.validate_fields(report.selected_fields)
        if not validated_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid fields provided"
            )

        # Validate filters
        safe_filters = QuerySecurityValidator.sanitize_filters(report.filters)

        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Insert report
        query = supabase.table("saved_reports").insert({
            "organization_id": str(user.current_organization_id),
            "created_by": str(user.id),
            "name": report.name,
            "description": report.description,
            "filters": safe_filters,
            "selected_fields": validated_fields,
            "aggregations": report.aggregations,
            "visibility": report.visibility
        })

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create saved report"
            )

        # Invalidate cache
        await invalidate_report_cache(str(user.current_organization_id))

        return SavedReport(**result.data[0])

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating saved report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create saved report: {str(e)}"
        )


@router.put("/saved-reports/{report_id}", response_model=SavedReport)
async def update_saved_report(
    report_id: UUID,
    report: SavedReportUpdate,
    user: User = Depends(get_current_user)
):
    """
    Update existing saved report.

    Auto-creates version record on change (via trigger).
    Invalidates cache after update.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Validate fields if provided
        update_data = {}
        if report.name is not None:
            update_data["name"] = report.name
        if report.description is not None:
            update_data["description"] = report.description
        if report.selected_fields is not None:
            validated_fields = QuerySecurityValidator.validate_fields(report.selected_fields)
            if not validated_fields:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="No valid fields provided"
                )
            update_data["selected_fields"] = validated_fields
        if report.filters is not None:
            safe_filters = QuerySecurityValidator.sanitize_filters(report.filters)
            update_data["filters"] = safe_filters
        if report.aggregations is not None:
            update_data["aggregations"] = report.aggregations
        if report.visibility is not None:
            update_data["visibility"] = report.visibility

        if not update_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No fields to update"
            )

        # Update report
        query = supabase.table("saved_reports") \
            .update(update_data) \
            .eq("id", str(report_id)) \
            .eq("organization_id", str(user.current_organization_id)) \
            .eq("created_by", str(user.id)) \
            .is_("deleted_at", "null")

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Saved report not found or you don't have permission to update it"
            )

        # Invalidate cache
        await invalidate_report_cache(str(user.current_organization_id))

        return SavedReport(**result.data[0])

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating saved report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update saved report: {str(e)}"
        )


@router.delete("/saved-reports/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_saved_report(
    report_id: UUID,
    user: User = Depends(get_current_user)
):
    """
    Soft delete saved report.

    Sets deleted_at timestamp. Report remains in database for audit.
    Invalidates cache after deletion.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Soft delete (set deleted_at)
        query = supabase.table("saved_reports") \
            .update({"deleted_at": datetime.utcnow().isoformat()}) \
            .eq("id", str(report_id)) \
            .eq("organization_id", str(user.current_organization_id)) \
            .eq("created_by", str(user.id)) \
            .is_("deleted_at", "null")

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Saved report not found or you don't have permission to delete it"
            )

        # Invalidate cache
        await invalidate_report_cache(str(user.current_organization_id))

        return None

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting saved report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete saved report: {str(e)}"
        )


# ============================================================================
# TASK 6: QUERY ENDPOINT (STANDARD MODE)
# ============================================================================

@router.post("/query", response_model=AnalyticsQueryResponse)
@limiter.limit("10/minute")
async def execute_analytics_query(
    request: Request,
    query_request: AnalyticsQueryRequest,
    user: User = Depends(get_current_user)
):
    """
    Execute analytics query (standard mode).

    Returns individual quote rows with selected fields.
    Uses Redis cache (10-min TTL) for repeated queries.
    If ≥2,000 quotes: Return task_id for background processing (stub).
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        # Generate cache key
        cache_key = get_cache_key(
            str(user.current_organization_id),
            query_request.filters,
            query_request.selected_fields
        )

        # Check cache first
        cached = await get_cached_report(cache_key)
        if cached:
            logger.info(f"Cache hit for query: {cache_key}")
            return AnalyticsQueryResponse(
                rows=cached.get("rows", []),
                count=cached.get("count", 0),
                total_count=cached.get("total_count"),
                has_more=cached.get("has_more", False),
                status="completed",
                message="Cached result"
            )

        # Count quotes to determine threshold
        conn = await get_db_connection()
        try:
            await set_rls_context(conn, user)

            # Count total matching quotes
            count_sql, count_params = build_analytics_query(
                user.current_organization_id,
                query_request.filters,
                ["id"],
                limit=1000000,  # High limit for counting
                offset=0
            )
            # Replace SELECT fields with COUNT(*)
            count_sql = count_sql.replace(
                "SELECT id",
                "SELECT COUNT(*) as total"
            ).replace("ORDER BY q.created_at DESC", "").replace("LIMIT $", "-- LIMIT $")

            # Remove LIMIT and OFFSET params (last 2)
            count_params = count_params[:-2]

            total_count = await conn.fetchval(count_sql, *count_params)

            # If ≥2,000 quotes: Stub for background processing
            if total_count >= 2000:
                # TODO: Implement background task queue
                import uuid
                task_id = str(uuid.uuid4())

                return AnalyticsQueryResponse(
                    rows=[],
                    count=0,
                    total_count=total_count,
                    has_more=False,
                    task_id=task_id,
                    status="processing",
                    message=f"Query processing in background. {total_count} quotes found. Check back with task_id."
                )

            # Execute query immediately
            start_time = time.time()

            sql, params = build_analytics_query(
                user.current_organization_id,
                query_request.filters,
                query_request.selected_fields,
                limit=query_request.limit,
                offset=query_request.offset
            )

            rows = await conn.fetch(sql, *params)

            execution_time_ms = int((time.time() - start_time) * 1000)

            # Convert rows to dicts
            result_rows = [dict(row) for row in rows]

            # Prepare response
            response_data = {
                "rows": result_rows,
                "count": len(result_rows),
                "total_count": total_count,
                "has_more": (query_request.offset + len(result_rows)) < total_count,
                "status": "completed",
                "message": f"Query executed in {execution_time_ms}ms"
            }

            # Cache result
            await cache_report(cache_key, response_data)

            return AnalyticsQueryResponse(**response_data)

        finally:
            await release_db_connection(conn)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error executing analytics query: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Query execution failed: {str(e)}"
        )


# ============================================================================
# TASK 7: AGGREGATE ENDPOINT (LIGHTWEIGHT MODE)
# ============================================================================

@router.post("/aggregate", response_model=AnalyticsAggregateResponse)
@limiter.limit("10/minute")
async def execute_analytics_aggregation(
    request: Request,
    query_request: AnalyticsQueryRequest,
    user: User = Depends(get_current_user)
):
    """
    Execute analytics aggregation (lightweight mode).

    Returns aggregations only (no individual rows).
    Faster than standard query for summary data.
    Uses Redis cache (10-min TTL).
    """
    # Admin check
    await check_admin_permissions(user)

    # Validate aggregation functions
    valid_functions = {'sum', 'avg', 'count', 'min', 'max'}

    if query_request.aggregations:
        for field, config in query_request.aggregations.items():
            func = config.get('function', '').lower()
            if func not in valid_functions:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid aggregation function: {func}. Allowed: {', '.join(sorted(valid_functions))}"
                )

    try:
        # Generate cache key
        cache_key = get_cache_key(
            str(user.current_organization_id),
            query_request.filters,
            [],
            aggregations=query_request.aggregations
        )

        # Check cache first
        cached = await get_cached_report(cache_key)
        if cached:
            logger.info(f"Cache hit for aggregation: {cache_key}")
            return AnalyticsAggregateResponse(
                aggregations=cached.get("aggregations", {}),
                execution_time_ms=cached.get("execution_time_ms", 0)
            )

        # Execute aggregation query
        conn = await get_db_connection()
        try:
            await set_rls_context(conn, user)

            start_time = time.time()

            sql, params = build_aggregation_query(
                user.current_organization_id,
                query_request.filters,
                query_request.aggregations or {}
            )

            row = await conn.fetchrow(sql, *params)

            execution_time_ms = int((time.time() - start_time) * 1000)

            # Convert row to dict
            result_aggregations = dict(row) if row else {}

            # Prepare response
            response_data = {
                "aggregations": result_aggregations,
                "execution_time_ms": execution_time_ms
            }

            # Cache result
            await cache_report(cache_key, response_data)

            return AnalyticsAggregateResponse(**response_data)

        finally:
            await release_db_connection(conn)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error executing aggregation: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Aggregation failed: {str(e)}"
        )


# ============================================================================
# TASK 8: EXPORT ENDPOINT (EXCEL/CSV)
# ============================================================================

@router.post("/export")
@limiter.limit("5/hour")
async def export_analytics_data(
    request: Request,
    background_tasks: BackgroundTasks,
    query_request: AnalyticsQueryRequest,
    user: User = Depends(get_current_user),
    format: str = Query(default="xlsx", description="Export format: xlsx or csv")
):
    """
    Export analytics data to Excel or CSV.

    Executes query, generates file, uploads to Supabase Storage,
    creates audit record in report_executions table.
    File expires in 7 days. Cleanup handled by background task.
    """
    # Admin check
    await check_admin_permissions(user)

    # Validate format
    export_format = format.lower()  # Normalize to lowercase
    if export_format not in ["xlsx", "csv"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid export format. Use 'xlsx' or 'csv'."
        )

    try:
        # Execute query
        conn = await get_db_connection()
        try:
            await set_rls_context(conn, user)

            start_time = time.time()

            sql, params = build_analytics_query(
                user.current_organization_id,
                query_request.filters,
                query_request.selected_fields,
                limit=query_request.limit,
                offset=query_request.offset
            )

            rows = await conn.fetch(sql, *params)

            execution_time_ms = int((time.time() - start_time) * 1000)

            if not rows:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="No data found for export"
                )

            # Generate file
            if export_format == "xlsx":
                file_path = await generate_excel_export(
                    rows,
                    query_request.selected_fields,
                    user.current_organization_id
                )
                media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:  # csv
                file_path = await generate_csv_export(
                    rows,
                    query_request.selected_fields,
                    user.current_organization_id
                )
                media_type = "text/csv"

            # Upload to Supabase Storage
            file_url = await upload_to_storage(
                file_path,
                str(user.current_organization_id)
            )

            # Get file size
            file_size_bytes = os.path.getsize(file_path)

            # Create execution record
            await create_execution_record(
                user=user,
                query_request=query_request,
                result_count=len(rows),
                execution_time_ms=execution_time_ms,
                export_format=export_format,
                export_file_url=file_url,
                file_size_bytes=file_size_bytes,
                request=request
            )

            # Schedule file cleanup
            background_tasks.add_task(cleanup_temp_file, file_path)

            # Return file
            return FileResponse(
                path=file_path,
                media_type=media_type,
                filename=os.path.basename(file_path)
            )

        finally:
            await release_db_connection(conn)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Export failed: {str(e)}"
        )


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

async def generate_excel_export(
    rows: List[Any],
    selected_fields: List[str],
    org_id: UUID
) -> str:
    """
    Generate Excel file with Russian formatting.

    Returns: Path to temporary file
    """
    import tempfile

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analytics Export"

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_idx, field in enumerate(selected_fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(selected_fields, start=1):
            value = row.get(field)

            # Convert Decimal to float for Excel
            if isinstance(value, Decimal):
                value = float(value)

            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to temp file
    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx",
        prefix=f"analytics_export_{org_id}_"
    )
    wb.save(temp_file.name)

    return temp_file.name


async def generate_csv_export(
    rows: List[Any],
    selected_fields: List[str],
    org_id: UUID
) -> str:
    """
    Generate CSV file.

    Returns: Path to temporary file
    """
    import tempfile

    # Create temp file
    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".csv",
        prefix=f"analytics_export_{org_id}_",
        mode="w",
        newline="",
        encoding="utf-8"
    )

    # Write CSV
    writer = csv.DictWriter(temp_file, fieldnames=selected_fields)
    writer.writeheader()

    for row in rows:
        # Convert Decimal to string
        row_data = {}
        for field in selected_fields:
            value = row.get(field)
            if isinstance(value, Decimal):
                row_data[field] = str(value)
            else:
                row_data[field] = value

        writer.writerow(row_data)

    temp_file.close()

    return temp_file.name


async def upload_to_storage(file_path: str, org_id: str) -> str:
    """
    Upload file to Supabase Storage and return public URL.

    Files stored in: analytics/{org_id}/{date}/{filename}
    """
    supabase: Client = create_client(
        os.getenv("SUPABASE_URL"),
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    )

    with open(file_path, 'rb') as f:
        file_data = f.read()

    filename = os.path.basename(file_path)
    storage_path = f"{org_id}/{datetime.now():%Y%m%d}/{filename}"

    # Upload to analytics bucket
    result = supabase.storage.from_("analytics").upload(
        storage_path,
        file_data,
        file_options={
            "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if filename.endswith(".xlsx") else "text/csv"
        }
    )

    # Get public URL
    url = supabase.storage.from_("analytics").get_public_url(storage_path)

    return url


async def create_execution_record(
    user: User,
    query_request: AnalyticsQueryRequest,
    result_count: int,
    execution_time_ms: int,
    export_format: str,
    export_file_url: str,
    file_size_bytes: int,
    request: Request
) -> None:
    """
    Create audit record in report_executions table.

    Immutable record for compliance and tracking.
    """
    supabase: Client = create_client(
        os.getenv("SUPABASE_URL"),
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    )

    # Get client IP and user agent
    ip_address = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent")

    # Calculate expiry (7 days)
    file_expires_at = datetime.utcnow() + timedelta(days=7)

    query = supabase.table("report_executions").insert({
        "organization_id": str(user.current_organization_id),
        "executed_by": str(user.id),
        "execution_type": "manual",
        "filters": query_request.filters,
        "selected_fields": query_request.selected_fields,
        "aggregations": query_request.aggregations,
        "result_summary": {
            "count": result_count,
            "limit": query_request.limit,
            "offset": query_request.offset
        },
        "quote_count": result_count,
        "export_file_url": export_file_url,
        "export_format": export_format,
        "file_size_bytes": file_size_bytes,
        "file_expires_at": file_expires_at.isoformat(),
        "ip_address": ip_address,
        "user_agent": user_agent,
        "execution_time_ms": execution_time_ms
    })

    await async_supabase_call(query)


async def cleanup_temp_file(file_path: str) -> None:
    """Background task to cleanup temporary file"""
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Cleaned up temp file: {file_path}")
    except Exception as e:
        logger.warning(f"Failed to cleanup temp file {file_path}: {e}")


# ============================================================================
# TASK 9: EXECUTION HISTORY - LIST ENDPOINT
# ============================================================================

@router.get("/executions")
async def list_executions(
    page: int = 1,
    page_size: int = 50,
    saved_report_id: Optional[UUID] = None,
    execution_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: User = Depends(get_current_user)
):
    """
    List execution history with pagination and filters.

    Returns paginated list of report executions with filters.
    Organization isolation via RLS.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Build query with filters
        query = supabase.table("report_executions") \
            .select("*", count="exact") \
            .eq("organization_id", str(user.current_organization_id))

        # Apply filters
        if saved_report_id:
            query = query.eq("saved_report_id", str(saved_report_id))

        if execution_type:
            query = query.eq("execution_type", execution_type)

        if date_from:
            query = query.gte("executed_at", date_from)

        if date_to:
            query = query.lte("executed_at", date_to)

        # Calculate pagination
        start = (page - 1) * page_size
        end = start + page_size - 1

        # Execute with pagination and ordering
        query = query.order("executed_at", desc=True).range(start, end)

        result = await async_supabase_call(query)

        total = result.count if result.count is not None else 0
        pages = max(1, (total + page_size - 1) // page_size) if page_size > 0 else 1

        return {
            "items": result.data,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing executions: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to list executions: {str(e)}"
        )


# ============================================================================
# TASK 10: EXECUTION HISTORY - GET SINGLE EXECUTION
# ============================================================================

@router.get("/executions/{execution_id}")
async def get_execution(
    execution_id: UUID,
    user: User = Depends(get_current_user)
):
    """
    Get detailed execution record by ID.

    Returns full execution record including query snapshot and results.
    Organization isolation via RLS.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        query = supabase.table("report_executions") \
            .select("*") \
            .eq("id", str(execution_id)) \
            .eq("organization_id", str(user.current_organization_id))

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Execution record not found"
            )

        return result.data[0]

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching execution: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch execution: {str(e)}"
        )


# ============================================================================
# TASK 11: EXECUTION HISTORY - DOWNLOAD FILE
# ============================================================================

@router.get("/executions/{execution_id}/download")
async def download_execution_file(
    execution_id: UUID,
    user: User = Depends(get_current_user)
):
    """
    Download exported file from execution record.

    Checks file expiration (7-day retention).
    Returns 410 Gone if file expired.
    Downloads from Supabase Storage and streams to client.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Get execution record
        query = supabase.table("report_executions") \
            .select("*") \
            .eq("id", str(execution_id)) \
            .eq("organization_id", str(user.current_organization_id))

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Execution record not found"
            )

        execution = result.data[0]

        # Check if file exists
        if not execution.get("export_file_url"):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No export file available for this execution"
            )

        # Check expiration
        file_expires_at = execution.get("file_expires_at")
        if file_expires_at:
            expiry_date = datetime.fromisoformat(file_expires_at.replace('Z', '+00:00'))
            if datetime.now(expiry_date.tzinfo) > expiry_date:
                raise HTTPException(
                    status_code=status.HTTP_410_GONE,
                    detail="Export file has expired (7-day retention)"
                )

        # Download file from storage
        file_url = execution["export_file_url"]
        temp_path = await download_from_storage(file_url)

        # Determine content type
        export_format = execution.get("export_format", "xlsx")
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if export_format == "xlsx"
            else "text/csv"
        )

        # Return file
        return FileResponse(
            path=temp_path,
            media_type=media_type,
            filename=os.path.basename(temp_path)
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading file: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to download file: {str(e)}"
        )


# ============================================================================
# TASK 12: SCHEDULED REPORTS - CRUD ENDPOINTS
# ============================================================================

@router.get("/scheduled")
async def list_scheduled_reports(user: User = Depends(get_current_user)):
    """
    List all scheduled reports for current organization.

    Returns scheduled reports with saved_report details (JOIN).
    Organization isolation via RLS.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Query with JOIN to get saved_report details
        query = supabase.table("scheduled_reports") \
            .select("*, saved_report:saved_reports(id, name, description, filters, selected_fields)") \
            .eq("organization_id", str(user.current_organization_id)) \
            .order("created_at", desc=True)

        result = await async_supabase_call(query)

        return result.data

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing scheduled reports: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to list scheduled reports: {str(e)}"
        )


@router.get("/scheduled/{schedule_id}")
async def get_scheduled_report(
    schedule_id: UUID,
    user: User = Depends(get_current_user)
):
    """
    Get single scheduled report by ID.

    Returns scheduled report with saved_report details.
    Organization isolation via RLS.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        query = supabase.table("scheduled_reports") \
            .select("*, saved_report:saved_reports(*)") \
            .eq("id", str(schedule_id)) \
            .eq("organization_id", str(user.current_organization_id))

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Scheduled report not found"
            )

        return result.data[0]

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching scheduled report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch scheduled report: {str(e)}"
        )


@router.post("/scheduled")
async def create_scheduled_report(
    schedule_data: Dict[str, Any],
    user: User = Depends(get_current_user)
):
    """
    Create new scheduled report.

    Validates cron expression and calculates next_run_at.
    Requires admin permissions.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        from croniter import croniter

        # Validate required fields
        if "saved_report_id" not in schedule_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="saved_report_id is required"
            )

        if "name" not in schedule_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="name is required"
            )

        if "schedule_cron" not in schedule_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="schedule_cron is required"
            )

        if "email_recipients" not in schedule_data or not schedule_data["email_recipients"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="At least one email recipient is required"
            )

        # Validate cron expression
        try:
            cron_expr = schedule_data["schedule_cron"]
            timezone = schedule_data.get("timezone", "Europe/Moscow")
            next_run = calculate_next_run(cron_expr, timezone)
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid cron expression: {str(e)}"
            )

        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Create schedule
        query = supabase.table("scheduled_reports").insert({
            "organization_id": str(user.current_organization_id),
            "saved_report_id": str(schedule_data["saved_report_id"]),
            "name": schedule_data["name"],
            "schedule_cron": cron_expr,
            "timezone": timezone,
            "email_recipients": schedule_data["email_recipients"],
            "include_file": schedule_data.get("include_file", True),
            "email_subject": schedule_data.get("email_subject"),
            "email_body": schedule_data.get("email_body"),
            "is_active": schedule_data.get("is_active", True),
            "next_run_at": next_run.isoformat(),
            "created_by": str(user.id)
        })

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create scheduled report"
            )

        return result.data[0]

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating scheduled report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create scheduled report: {str(e)}"
        )


@router.put("/scheduled/{schedule_id}")
async def update_scheduled_report(
    schedule_id: UUID,
    update_data: Dict[str, Any],
    user: User = Depends(get_current_user)
):
    """
    Update existing scheduled report.

    Recalculates next_run_at if cron expression changed.
    Requires admin permissions.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # If cron expression changed, recalculate next_run_at
        if "schedule_cron" in update_data:
            try:
                cron_expr = update_data["schedule_cron"]
                timezone = update_data.get("timezone", "Europe/Moscow")
                next_run = calculate_next_run(cron_expr, timezone)
                update_data["next_run_at"] = next_run.isoformat()
            except Exception as e:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid cron expression: {str(e)}"
                )

        # Update schedule
        query = supabase.table("scheduled_reports") \
            .update(update_data) \
            .eq("id", str(schedule_id)) \
            .eq("organization_id", str(user.current_organization_id))

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Scheduled report not found or you don't have permission to update it"
            )

        return result.data[0]

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating scheduled report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update scheduled report: {str(e)}"
        )


@router.delete("/scheduled/{schedule_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_scheduled_report(
    schedule_id: UUID,
    user: User = Depends(get_current_user)
):
    """
    Delete scheduled report.

    Hard delete (removes record from database).
    Requires admin permissions.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        query = supabase.table("scheduled_reports") \
            .delete() \
            .eq("id", str(schedule_id)) \
            .eq("organization_id", str(user.current_organization_id))

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Scheduled report not found or you don't have permission to delete it"
            )

        return None

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting scheduled report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete scheduled report: {str(e)}"
        )


# ============================================================================
# TASK 13: SCHEDULED REPORTS - MANUAL TRIGGER
# ============================================================================

@router.post("/scheduled/{schedule_id}/run")
async def run_scheduled_report(
    schedule_id: UUID,
    user: User = Depends(get_current_user)
):
    """
    Manually trigger scheduled report execution.

    Executes report, generates file, creates execution record,
    sends email (stub), and updates schedule status.
    """
    # Admin check
    await check_admin_permissions(user)

    try:
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        # Load schedule with saved_report
        query = supabase.table("scheduled_reports") \
            .select("*, saved_report:saved_reports(*)") \
            .eq("id", str(schedule_id)) \
            .eq("organization_id", str(user.current_organization_id))

        result = await async_supabase_call(query)

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Scheduled report not found"
            )

        schedule = result.data[0]
        saved_report = schedule.get("saved_report")

        if not saved_report:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Saved report not found for this schedule"
            )

        # Execute report
        execution = await execute_scheduled_report_internal(
            schedule=schedule,
            saved_report=saved_report,
            user=user,
            execution_type="manual"
        )

        return execution

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error running scheduled report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to run scheduled report: {str(e)}"
        )


# ============================================================================
# TASK 15: HELPER FUNCTIONS
# ============================================================================

async def download_from_storage(file_url: str) -> str:
    """
    Download file from Supabase Storage to temp directory.

    Returns: Path to temporary file
    """
    import tempfile

    supabase: Client = create_client(
        os.getenv("SUPABASE_URL"),
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    )

    # Extract bucket and path from URL
    # URL format: https://xxx.supabase.co/storage/v1/object/public/analytics/{path}
    bucket = "analytics"
    path = file_url.split(f"/{bucket}/")[1] if f"/{bucket}/" in file_url else file_url.split("/")[-1]

    # Download file
    file_data = supabase.storage.from_(bucket).download(path)

    # Save to temp
    filename = os.path.basename(path)
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1])
    temp_file.write(file_data)
    temp_file.close()

    return temp_file.name


def calculate_next_run(cron_expr: str, timezone: str = "Europe/Moscow") -> datetime:
    """
    Calculate next run time from cron expression.

    Returns: Datetime of next run (with timezone)
    """
    from croniter import croniter
    import pytz

    tz = pytz.timezone(timezone)
    now = datetime.now(tz)
    cron = croniter(cron_expr, now)
    next_run = cron.get_next(datetime)

    return next_run


def calculate_summary(rows: List[Any], selected_fields: List[str]) -> Dict[str, Any]:
    """
    Calculate aggregations from query results.

    Returns: Dictionary with count and field-specific aggregations
    """
    if not rows:
        return {"count": 0}

    summary = {
        "count": len(rows),
        "fields": {}
    }

    # Calculate aggregations for numeric fields
    for field in selected_fields:
        values = []
        for row in rows:
            value = row.get(field)
            if value is not None and isinstance(value, (int, float, Decimal)):
                values.append(float(value))

        if values:
            summary["fields"][field] = {
                "sum": sum(values),
                "avg": sum(values) / len(values),
                "min": min(values),
                "max": max(values)
            }

    return summary


async def execute_scheduled_report_internal(
    schedule: Dict,
    saved_report: Dict,
    user: User,
    execution_type: str = "manual"
) -> Dict:
    """
    Execute scheduled report and create audit record.

    Generates file, uploads to storage, creates execution record,
    sends email (stub), and updates schedule status.
    """
    start_time = time.time()

    try:
        # Execute query
        conn = await get_db_connection()
        try:
            await set_rls_context(conn, user)

            sql, params = build_analytics_query(
                user.current_organization_id,
                saved_report.get("filters", {}),
                saved_report.get("selected_fields", []),
                limit=100000,  # High limit for exports
                offset=0
            )

            rows = await conn.fetch(sql, *params)

        finally:
            await release_db_connection(conn)

        if not rows:
            raise ValueError("No data found for report")

        # Generate file
        file_path = await generate_excel_export(
            rows,
            saved_report["selected_fields"],
            user.current_organization_id
        )

        # Upload to storage
        file_url = await upload_to_storage(
            file_path,
            str(user.current_organization_id)
        )

        # Get file size
        file_size_bytes = os.path.getsize(file_path)

        execution_time_ms = int((time.time() - start_time) * 1000)

        # Calculate summary
        result_summary = calculate_summary(rows, saved_report["selected_fields"])

        # Create execution record
        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        file_expires_at = datetime.utcnow() + timedelta(days=7)

        query = supabase.table("report_executions").insert({
            "organization_id": str(user.current_organization_id),
            "executed_by": str(user.id),
            "saved_report_id": str(saved_report["id"]),
            "report_name": saved_report["name"],
            "execution_type": execution_type,
            "filters": saved_report.get("filters", {}),
            "selected_fields": saved_report["selected_fields"],
            "aggregations": saved_report.get("aggregations"),
            "result_summary": result_summary,
            "quote_count": len(rows),
            "export_file_url": file_url,
            "export_format": "xlsx",
            "file_size_bytes": file_size_bytes,
            "file_expires_at": file_expires_at.isoformat(),
            "ip_address": "system",
            "user_agent": "scheduled_task",
            "execution_time_ms": execution_time_ms
        })

        execution_result = await async_supabase_call(query)

        if not execution_result.data:
            raise ValueError("Failed to create execution record")

        execution = execution_result.data[0]

        # Update schedule status
        next_run = calculate_next_run(schedule["schedule_cron"], schedule.get("timezone", "Europe/Moscow"))

        update_query = supabase.table("scheduled_reports").update({
            "last_run_at": datetime.utcnow().isoformat(),
            "next_run_at": next_run.isoformat(),
            "last_run_status": "success",
            "consecutive_failures": 0,
            "last_error": None
        }).eq("id", schedule["id"])

        await async_supabase_call(update_query)

        # Stub: Send email
        logger.info(f"Email would be sent to: {schedule.get('email_recipients', [])}")
        logger.info(f"Subject: {schedule.get('email_subject', f'Scheduled Report: {saved_report['name']}')}")
        logger.info(f"Attachment: {file_url}")

        # Cleanup temp file
        if os.path.exists(file_path):
            os.remove(file_path)

        return execution

    except Exception as e:
        # Update schedule with failure status
        logger.error(f"Failed to execute scheduled report: {e}")

        supabase: Client = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        )

        consecutive_failures = schedule.get("consecutive_failures", 0) + 1

        update_query = supabase.table("scheduled_reports").update({
            "last_run_status": "failure",
            "last_error": str(e),
            "consecutive_failures": consecutive_failures
        }).eq("id", schedule["id"])

        await async_supabase_call(update_query)

        raise
