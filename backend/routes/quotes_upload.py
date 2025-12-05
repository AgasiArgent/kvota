"""
Quote Upload from Simplified Excel Template

Endpoint: POST /api/quotes/upload-excel
Accepts simplified Excel template, parses it, and runs calculation engine.
Optionally saves to database if customer_id is provided.

Created: 2025-11-28
Updated: 2025-12-01 - Added save-to-DB functionality
"""

import io
import os
import logging
from datetime import date, datetime, timedelta
from typing import Optional, List
from decimal import Decimal, ROUND_HALF_UP
from uuid import UUID
from urllib.parse import quote as url_quote

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from pydantic import BaseModel
from supabase import Client

from auth import get_current_user, User
from dependencies import get_supabase
from excel_parser.simplified_parser import (
    SimplifiedExcelParser,
    SimplifiedQuoteInput,
    ProductInput,
    parse_simplified_template,
)
from calculation_models import (
    QuoteCalculationInput,
    ProductInfo,
    FinancialParams,
    LogisticsParams,
    TaxesAndDuties,
    PaymentTerms,
    CustomsAndClearance,
    CompanySettings,
    SystemConfig,
    Currency as CalcCurrency,
    SupplierCountry,
    SellerCompany,
    OfferSaleType,
    Incoterms,
    DMFeeType,
)
from calculation_engine import calculate_multiproduct_quote
from services.exchange_rate_service import get_exchange_rate_service
from services.export_validation_service import generate_validation_export
from fastapi.responses import StreamingResponse

# Import helper functions from quotes_calc
from routes.quotes_calc import (
    generate_quote_number,
    aggregate_product_results_to_summary,
    get_rates_snapshot_to_usd,
    convert_decimals_to_float,
)

# Supabase client is now injected via dependency injection (see get_supabase)


logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/quotes", tags=["quotes-upload"])


# ============================================================================
# RESPONSE MODELS
# ============================================================================

class CalculationResultSummary(BaseModel):
    """Full calculation results with all intermediate values"""
    # Quote-level totals
    total_purchase_price: str  # S13
    total_cogs: str  # AB13
    total_revenue: str  # AK13 (without VAT)
    total_revenue_with_vat: str  # AL13
    total_profit: str  # AF13
    total_financing_cost: str  # BJ11
    total_credit_interest: str  # BL5
    margin_percent: str  # profit / COGS

    # Quote-level financing (BH/BI/BJ/BL series)
    evaluated_revenue: Optional[str] = None  # BH2
    client_advance: Optional[str] = None  # BH3
    supplier_payment: Optional[str] = None  # BH6
    supplier_financing_cost: Optional[str] = None  # BJ7
    operational_financing_cost: Optional[str] = None  # BJ10

    # All products with FULL intermediate values (30+ fields each)
    products: List[dict]  # Full ProductCalculationResult as dicts

    # Exchange rates used
    exchange_rates_used: dict


class UploadResponse(BaseModel):
    """Response from upload endpoint"""
    success: bool
    message: str
    quote_input: Optional[dict] = None
    calculation_results: Optional[CalculationResultSummary] = None
    errors: Optional[list] = None


class ParseOnlyResponse(BaseModel):
    """Response when only parsing (no calculation)"""
    success: bool
    parsed_data: dict
    errors: Optional[list] = None


# ============================================================================
# CURRENCY AND ENUM MAPPINGS
# ============================================================================

# Map simplified parser Currency to calculation_models Currency
CURRENCY_MAP = {
    "USD": CalcCurrency.USD,
    "EUR": CalcCurrency.EUR,
    "RUB": CalcCurrency.RUB,
    "TRY": CalcCurrency.TRY,
    "CNY": CalcCurrency.CNY,
}

# Map supplier country strings to SupplierCountry enum
SUPPLIER_COUNTRY_MAP = {
    "Турция": SupplierCountry.TURKEY,
    "Турция (транзитная зона)": SupplierCountry.TURKEY_TRANSIT,
    "Россия": SupplierCountry.RUSSIA,
    "Китай": SupplierCountry.CHINA,
    "Литва": SupplierCountry.LITHUANIA,
    "Латвия": SupplierCountry.LATVIA,
    "Болгария": SupplierCountry.BULGARIA,
    "Польша": SupplierCountry.POLAND,
    "ЕС (закупка между странами ЕС)": SupplierCountry.EU_CROSS_BORDER,
    "ОАЭ": SupplierCountry.UAE,
    "Прочие": SupplierCountry.OTHER,
}

# Map seller company strings to SellerCompany enum
SELLER_COMPANY_MAP = {
    "МАСТЕР БЭРИНГ ООО": SellerCompany.MASTER_BEARING_RU,
    "ЦМТО1 ООО": SellerCompany.CMTO1_RU,
    "РАД РЕСУРС ООО": SellerCompany.RAD_RESURS_RU,
    "TEXCEL OTOMOTİV TİCARET LİMİTED ŞİRKETİ": SellerCompany.TEXCEL_TR,
    "GESTUS DIŞ TİCARET LİMİTED ŞİRKETİ": SellerCompany.GESTUS_TR,
    "UPDOOR Limited": SellerCompany.UPDOOR_CN,
}

# Map sale type strings to OfferSaleType enum
SALE_TYPE_MAP = {
    "поставка": OfferSaleType.SUPPLY,
    "транзит": OfferSaleType.TRANSIT,
    "финтранзит": OfferSaleType.FIN_TRANSIT,
    "экспорт": OfferSaleType.EXPORT,
}

# Map incoterms strings to Incoterms enum
INCOTERMS_MAP = {
    "DDP": Incoterms.DDP,
    "DAP": Incoterms.DAP,
    "CIF": Incoterms.CIF,
    "FOB": Incoterms.FOB,
    "EXW": Incoterms.EXW,
}


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

async def get_exchange_rates(quote_currency: str, product_currencies: List[str]) -> dict:
    """
    Get exchange rates from CBR for all currencies used in the quote.

    Returns dict mapping currency pair to rate, e.g.:
    {"EUR/RUB": 95.5, "USD/RUB": 88.3, ...}
    """
    service = get_exchange_rate_service()
    rates = {}

    # Get unique currencies needed
    all_currencies = set(product_currencies)
    all_currencies.add(quote_currency)

    # Get rates to RUB (CBR base currency)
    for currency in all_currencies:
        if currency == "RUB":
            rates[f"{currency}/RUB"] = Decimal("1.0")
        else:
            rate = await service.get_rate(currency, "RUB")
            if rate:
                rates[f"{currency}/RUB"] = rate
            else:
                logger.warning(f"Could not get rate for {currency}/RUB")
                rates[f"{currency}/RUB"] = Decimal("1.0")  # Fallback

    return rates


def calculate_exchange_rate(
    from_currency: str,
    to_currency: str,
    rates: dict,
    for_division: bool = False
) -> Decimal:
    """
    Calculate exchange rate between two currencies using RUB as base.

    Args:
        from_currency: Source currency (e.g., "TRY")
        to_currency: Target currency (e.g., "EUR")
        rates: Dict of currency/RUB rates from CBR
        for_division: If True, returns rate for division (to/from).
                      If False, returns rate for multiplication (from/to).

    The calculation engine uses DIVISION: R16 = P16 / exchange_rate
    So for engine's exchange_rate_base_price_to_quote, use for_division=True.

    For converting costs (multiply), use for_division=False (default).
    """
    if from_currency == to_currency:
        return Decimal("1.0")

    from_rub = rates.get(f"{from_currency}/RUB", Decimal("1.0"))
    to_rub = rates.get(f"{to_currency}/RUB", Decimal("1.0"))

    if for_division:
        # For division: rate = to_rub / from_rub
        # Example: TRY→EUR, returns 4.92 (how many TRY per 1 EUR)
        # Engine does: 83333 TRY / 4.92 = 16941 EUR
        if from_rub == 0:
            return Decimal("1.0")
        result = to_rub / from_rub
    else:
        # For multiplication: rate = from_rub / to_rub
        # Example: EUR→RUB, returns 90.78 (how many RUB per 1 EUR)
        # Usage: 1500 EUR * 90.78 = 136170 RUB
        if to_rub == 0:
            return Decimal("1.0")
        result = from_rub / to_rub

    # Round to 4 decimal places for consistency with calculation engine
    return result.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


async def map_to_calculation_inputs(
    parsed: SimplifiedQuoteInput,
    rates: dict
) -> List[QuoteCalculationInput]:
    """
    Map SimplifiedQuoteInput to list of QuoteCalculationInput (one per product).
    """
    inputs = []
    quote_currency = CURRENCY_MAP.get(parsed.quote_currency.value, CalcCurrency.EUR)

    for product in parsed.products:
        # Calculate exchange rate for this product (for_division=True for engine's R16 = P16 / rate)
        product_currency = product.currency.value
        exchange_rate = calculate_exchange_rate(
            product_currency,
            parsed.quote_currency.value,
            rates,
            for_division=True  # Engine divides by this rate
        )

        # Map supplier country
        supplier_country = SUPPLIER_COUNTRY_MAP.get(
            product.supplier_country,
            SupplierCountry.TURKEY
        )

        # Build QuoteCalculationInput
        calc_input = QuoteCalculationInput(
            product=ProductInfo(
                base_price_VAT=product.base_price_vat,
                quantity=product.quantity,
                weight_in_kg=product.weight_kg or Decimal("0"),
                currency_of_base_price=CURRENCY_MAP.get(product_currency, CalcCurrency.EUR),
                customs_code=str(product.customs_code or "8708913509").zfill(10),
            ),
            financial=FinancialParams(
                currency_of_quote=quote_currency,
                exchange_rate_base_price_to_quote=exchange_rate,
                supplier_discount=product.supplier_discount * Decimal("100"),  # Convert 0.10 -> 10%
                markup=product.markup * Decimal("100"),  # Convert 0.15 -> 15%
                rate_forex_risk=Decimal("3"),  # Default
                dm_fee_type=DMFeeType.PERCENTAGE if parsed.dm_fee_type == "% от суммы" else DMFeeType.FIXED,
                dm_fee_value=parsed.dm_fee_value,
            ),
            logistics=LogisticsParams(
                supplier_country=supplier_country,
                offer_incoterms=INCOTERMS_MAP.get(parsed.incoterms, Incoterms.DDP),
                delivery_time=parsed.delivery_time,
                # Calculate delivery_date for VAT determination (22% for 2026+)
                delivery_date=date.today() + timedelta(days=parsed.delivery_time),
                # Convert logistics costs to quote currency
                logistics_supplier_hub=parsed.logistics_supplier_hub.value * calculate_exchange_rate(
                    parsed.logistics_supplier_hub.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
                logistics_hub_customs=parsed.logistics_hub_customs.value * calculate_exchange_rate(
                    parsed.logistics_hub_customs.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
                logistics_customs_client=parsed.logistics_customs_client.value * calculate_exchange_rate(
                    parsed.logistics_customs_client.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
            ),
            taxes=TaxesAndDuties(
                import_tariff=product.import_tariff * Decimal("100"),  # Convert 0.05 -> 5%
                excise_tax=Decimal("0"),
                util_fee=Decimal("0"),
            ),
            payment=PaymentTerms(
                # All percentages: convert from decimal (0.30) to percentage (30)
                advance_from_client=next(
                    (m.percentage * Decimal("100") for m in parsed.payment_milestones if m.name == "advance_from_client"),
                    Decimal("100")
                ),
                advance_to_supplier=parsed.advance_to_supplier * Decimal("100"),  # Convert 0.50 -> 50%
                time_to_advance=next(
                    (m.days or 0 for m in parsed.payment_milestones if m.name == "advance_from_client"),
                    0
                ),
                advance_on_loading=next(
                    (m.percentage * Decimal("100") for m in parsed.payment_milestones if m.name == "advance_on_loading"),
                    Decimal("0")
                ),
                time_to_advance_loading=next(
                    (m.days or 0 for m in parsed.payment_milestones if m.name == "advance_on_loading"),
                    0
                ),
                advance_on_going_to_country_destination=next(
                    (m.percentage * Decimal("100") for m in parsed.payment_milestones if m.name == "advance_on_shipping"),
                    Decimal("0")
                ),
                time_to_advance_going_to_country_destination=next(
                    (m.days or 0 for m in parsed.payment_milestones if m.name == "advance_on_shipping"),
                    0
                ),
                advance_on_customs_clearance=next(
                    (m.percentage * Decimal("100") for m in parsed.payment_milestones if m.name == "advance_on_customs"),
                    Decimal("0")
                ),
                time_to_advance_on_customs_clearance=next(
                    (m.days or 0 for m in parsed.payment_milestones if m.name == "advance_on_customs"),
                    0
                ),
                time_to_advance_on_receiving=next(
                    (m.days or 0 for m in parsed.payment_milestones if m.name == "on_receiving"),
                    0
                ),
            ),
            customs=CustomsAndClearance(
                brokerage_hub=parsed.brokerage_hub.value * calculate_exchange_rate(
                    parsed.brokerage_hub.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
                brokerage_customs=parsed.brokerage_customs.value * calculate_exchange_rate(
                    parsed.brokerage_customs.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
                warehousing_at_customs=parsed.warehousing.value * calculate_exchange_rate(
                    parsed.warehousing.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
                customs_documentation=parsed.documentation.value * calculate_exchange_rate(
                    parsed.documentation.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
                brokerage_extra=parsed.other_costs.value * calculate_exchange_rate(
                    parsed.other_costs.currency.value,
                    parsed.quote_currency.value,
                    rates
                ),
            ),
            company=CompanySettings(
                seller_company=SELLER_COMPANY_MAP.get(parsed.seller_company, SellerCompany.MASTER_BEARING_RU),
                offer_sale_type=SALE_TYPE_MAP.get(parsed.sale_type, OfferSaleType.SUPPLY),
            ),
            system=SystemConfig(),  # Use defaults
        )

        inputs.append(calc_input)

    return inputs


def format_decimal(value: Decimal, precision: int = 2) -> str:
    """Format decimal for display"""
    return f"{value:.{precision}f}"


def serialize_product_result(result, product_input) -> dict:
    """Serialize ProductCalculationResult to dict with all fields as strings"""
    # Get all fields from the result model
    data = result.model_dump()

    # Convert all Decimal values to formatted strings
    for key, value in data.items():
        if isinstance(value, Decimal):
            data[key] = format_decimal(value, 4)  # 4 decimal places for precision
        elif value is None:
            data[key] = None

    # Add input fields for reference
    data["_input_name"] = product_input.name
    data["_input_sku"] = product_input.sku
    data["_input_brand"] = product_input.brand
    data["_input_quantity"] = product_input.quantity
    data["_input_currency"] = product_input.currency.value
    data["_input_base_price_vat"] = format_decimal(product_input.base_price_vat)
    data["_input_markup"] = format_decimal(product_input.markup)

    return data


def build_calculation_summary(
    parsed: SimplifiedQuoteInput,
    results: List,  # List[ProductCalculationResult]
    rates: dict
) -> CalculationResultSummary:
    """Build summary from calculation results with ALL intermediate values"""

    # Serialize all products with full intermediate values
    products = []
    for product_input, product_result in zip(parsed.products, results):
        products.append(serialize_product_result(product_result, product_input))

    # Calculate totals by summing across all products
    total_purchase = sum(r.purchase_price_total_quote_currency for r in results)
    total_cogs = sum(r.cogs_per_product for r in results)
    total_profit = sum(r.profit for r in results)
    total_revenue_no_vat = sum(r.sales_price_total_no_vat for r in results)
    total_revenue_with_vat = sum(r.sales_price_total_with_vat for r in results)
    total_financing_cost = sum(r.financing_cost_initial + r.financing_cost_credit for r in results)
    total_credit_interest = Decimal("0")  # Will be populated from quote-level if available

    # Margin = Profit / COGS (as percentage)
    margin_percent = (total_profit / total_cogs * Decimal("100")) if total_cogs > 0 else Decimal("0")

    # Get quote-level financing from first product (if available)
    first_result = results[0] if results else None
    evaluated_revenue = first_result.quote_level_evaluated_revenue if first_result and first_result.quote_level_evaluated_revenue else None
    client_advance = first_result.quote_level_client_advance if first_result and first_result.quote_level_client_advance else None
    supplier_payment = first_result.quote_level_supplier_payment if first_result and first_result.quote_level_supplier_payment else None
    supplier_financing_cost = first_result.quote_level_supplier_financing_cost if first_result and first_result.quote_level_supplier_financing_cost else None
    operational_financing_cost = first_result.quote_level_operational_financing_cost if first_result and first_result.quote_level_operational_financing_cost else None

    if first_result and first_result.quote_level_credit_sales_interest:
        total_credit_interest = first_result.quote_level_credit_sales_interest

    return CalculationResultSummary(
        total_purchase_price=format_decimal(total_purchase),
        total_cogs=format_decimal(total_cogs),
        total_revenue=format_decimal(total_revenue_no_vat),
        total_revenue_with_vat=format_decimal(total_revenue_with_vat),
        total_profit=format_decimal(total_profit),
        total_financing_cost=format_decimal(total_financing_cost),
        total_credit_interest=format_decimal(total_credit_interest),
        margin_percent=format_decimal(margin_percent),
        evaluated_revenue=format_decimal(evaluated_revenue) if evaluated_revenue else None,
        client_advance=format_decimal(client_advance) if client_advance else None,
        supplier_payment=format_decimal(supplier_payment) if supplier_payment else None,
        supplier_financing_cost=format_decimal(supplier_financing_cost) if supplier_financing_cost else None,
        operational_financing_cost=format_decimal(operational_financing_cost) if operational_financing_cost else None,
        products=products,
        exchange_rates_used={k: format_decimal(v, 4) for k, v in rates.items()},
    )


# ============================================================================
# ENDPOINTS
# ============================================================================

@router.post("/upload-excel", response_model=UploadResponse)
async def upload_excel_quote(
    file: UploadFile = File(...),
    calculate: bool = True,
    user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase),
):
    """
    Upload simplified Excel template and optionally run calculation.

    Args:
        file: Excel file (.xlsx)
        calculate: If True, run calculation engine after parsing

    Returns:
        UploadResponse with parsed data and calculation results
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx)"
        )

    try:
        # Read file content
        content = await file.read()
        file_stream = io.BytesIO(content)

        # Parse Excel
        parser = SimplifiedExcelParser(file_stream)
        parsed_data = parser.parse()

        if not calculate:
            return UploadResponse(
                success=True,
                message="File parsed successfully",
                quote_input=parsed_data.dict(),
                calculation_results=None,
            )

        # Get exchange rates
        product_currencies = [p.currency.value for p in parsed_data.products]
        rates = await get_exchange_rates(parsed_data.quote_currency.value, product_currencies)

        # Map to calculation inputs
        calc_inputs = await map_to_calculation_inputs(parsed_data, rates)

        # Run calculation
        calc_result = calculate_multiproduct_quote(calc_inputs)

        # Build summary
        summary = build_calculation_summary(parsed_data, calc_result, rates)

        return UploadResponse(
            success=True,
            message="File parsed and calculated successfully",
            quote_input=parsed_data.dict(),
            calculation_results=summary,
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Error processing Excel upload")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )


@router.post("/parse-excel", response_model=ParseOnlyResponse)
async def parse_excel_only(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase),
):
    """
    Parse Excel template without running calculation.
    Useful for validation and preview.

    Args:
        file: Excel file (.xlsx)

    Returns:
        ParseOnlyResponse with parsed data
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx)"
        )

    try:
        content = await file.read()
        file_stream = io.BytesIO(content)

        parser = SimplifiedExcelParser(file_stream)
        parsed_data = parser.parse()

        return ParseOnlyResponse(
            success=True,
            parsed_data=parsed_data.dict(),
        )

    except ValueError as e:
        return ParseOnlyResponse(
            success=False,
            parsed_data={},
            errors=[str(e)],
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )


@router.post("/upload-excel-validation")
async def upload_excel_with_validation_export(
    file: UploadFile = File(...),
    customer_id: Optional[str] = Form(None),
    user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase),
):
    """
    Upload Excel, run calculation, and return validation Excel file.

    If customer_id is provided, also saves the quote to database (5 tables):
    - quotes, quote_items, quote_calculation_variables,
    - quote_calculation_results, quote_calculation_summaries

    The validation Excel contains:
    - API_Inputs tab: All uploaded input values
    - расчет tab: Modified to reference API_Inputs (formulas recalculate)
    - API_Results tab: API calculation outputs
    - Comparison tab: Detailed comparison with >0.01% differences highlighted

    Returns:
        Excel file (.xlsm) as downloadable attachment
        Headers include X-Quote-Id and X-Quote-Number if saved to DB
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx)"
        )

    try:
        # Read file content
        content = await file.read()
        file_stream = io.BytesIO(content)

        # Parse Excel
        parser = SimplifiedExcelParser(file_stream)
        parsed_data = parser.parse()

        # Get exchange rates
        product_currencies = [p.currency.value for p in parsed_data.products]
        rates = await get_exchange_rates(parsed_data.quote_currency.value, product_currencies)

        # Map to calculation inputs
        calc_inputs = await map_to_calculation_inputs(parsed_data, rates)

        # Run calculation
        calc_results = calculate_multiproduct_quote(calc_inputs)

        # Build quote-level inputs dict
        quote_inputs = {
            "seller_company": parsed_data.seller_company,
            "offer_sale_type": parsed_data.sale_type,
            "incoterms": parsed_data.incoterms,
            "quote_currency": parsed_data.quote_currency.value,
            "delivery_time": parsed_data.delivery_time,
            "advance_to_supplier": float(parsed_data.advance_to_supplier),
            # Payment milestones
            "advance_from_client": float(parsed_data.payment_milestones[0].percentage) if parsed_data.payment_milestones else 0,
            "time_to_advance": parsed_data.payment_milestones[0].days if parsed_data.payment_milestones else 0,
            # Days to payment after receipt (F7 in расчет) - from on_receiving milestone
            "time_to_payment": next(
                (m.days or 0 for m in parsed_data.payment_milestones if m.name == "on_receiving"),
                0
            ),
            # Logistics costs
            "logistics_supplier_hub": float(parsed_data.logistics_supplier_hub.value),
            "logistics_hub_customs": float(parsed_data.logistics_hub_customs.value),
            "logistics_customs_client": float(parsed_data.logistics_customs_client.value),
            # Brokerage costs
            "brokerage_hub": float(parsed_data.brokerage_hub.value),
            "brokerage_customs": float(parsed_data.brokerage_customs.value),
            "warehousing": float(parsed_data.warehousing.value),
            "documentation": float(parsed_data.documentation.value),
            "other_costs": float(parsed_data.other_costs.value),
            # DM Fee
            "dm_fee_type": parsed_data.dm_fee_type,
            "dm_fee_value": float(parsed_data.dm_fee_value),
            # Admin settings (defaults)
            "rate_forex_risk": 0.03,  # 3%
        }

        # Build product inputs list
        product_inputs = []
        for p in parsed_data.products:
            product_inputs.append({
                "brand": p.brand or "",
                "sku": p.sku or "",
                "name": p.name,
                "quantity": p.quantity,
                "weight_in_kg": float(p.weight_kg) if p.weight_kg else 0,
                "currency_of_base_price": p.currency.value,
                "base_price_vat": float(p.base_price_vat),
                "supplier_country": p.supplier_country if p.supplier_country else "Турция",
                "supplier_discount": float(p.supplier_discount) if p.supplier_discount else 0,
                "exchange_rate": float(calculate_exchange_rate(p.currency.value, parsed_data.quote_currency.value, rates, for_division=True)),
                "customs_code": p.customs_code or "",
                "import_tariff": float(p.import_tariff) if p.import_tariff else 0,
                "markup": float(p.markup),
            })

        # Build API results dict (quote-level)
        total_purchase = sum(r.purchase_price_total_quote_currency for r in calc_results)
        total_cogs = sum(r.cogs_per_product for r in calc_results)
        total_profit = sum(r.profit for r in calc_results)
        total_revenue_no_vat = sum(r.sales_price_total_no_vat for r in calc_results)
        total_revenue_with_vat = sum(r.sales_price_total_with_vat for r in calc_results)
        # Logistics totals
        total_logistics_first = sum(r.logistics_first_leg for r in calc_results)
        total_logistics_last = sum(r.logistics_last_leg for r in calc_results)
        total_logistics = sum(r.logistics_total for r in calc_results)

        first_result = calc_results[0] if calc_results else None

        api_results = {
            # Quote totals
            "total_purchase_price": float(total_purchase),
            "total_logistics_first": float(total_logistics_first),
            "total_logistics_last": float(total_logistics_last),
            "total_logistics": float(total_logistics),
            "total_cogs": float(total_cogs),
            "total_revenue": float(total_revenue_no_vat),
            "total_revenue_with_vat": float(total_revenue_with_vat),
            "total_profit": float(total_profit),
            # Financing
            "evaluated_revenue": float(first_result.quote_level_evaluated_revenue) if first_result and first_result.quote_level_evaluated_revenue else 0,
            "client_advance": float(first_result.quote_level_client_advance) if first_result and first_result.quote_level_client_advance else 0,
            "total_before_forwarding": float(first_result.quote_level_total_before_forwarding) if first_result and hasattr(first_result, 'quote_level_total_before_forwarding') and first_result.quote_level_total_before_forwarding else 0,
            "supplier_payment": float(first_result.quote_level_supplier_payment) if first_result and first_result.quote_level_supplier_payment else 0,
            "supplier_financing_cost": float(first_result.quote_level_supplier_financing_cost) if first_result and first_result.quote_level_supplier_financing_cost else 0,
            "operational_financing_cost": float(first_result.quote_level_operational_financing_cost) if first_result and first_result.quote_level_operational_financing_cost else 0,
            "total_financing_cost": float(first_result.quote_level_total_financing_cost) if first_result and first_result.quote_level_total_financing_cost else 0,
            "credit_sales_amount": float(first_result.quote_level_credit_sales_amount) if first_result and hasattr(first_result, 'quote_level_credit_sales_amount') and first_result.quote_level_credit_sales_amount else 0,
            "credit_sales_fv": float(first_result.quote_level_credit_sales_fv) if first_result and hasattr(first_result, 'quote_level_credit_sales_fv') and first_result.quote_level_credit_sales_fv else 0,
            "credit_sales_interest": float(first_result.quote_level_credit_sales_interest) if first_result and first_result.quote_level_credit_sales_interest else 0,
        }

        # Build product results list
        product_results = []
        for r in calc_results:
            product_results.append({
                "purchase_price_no_vat": float(r.purchase_price_no_vat),
                "purchase_price_after_discount": float(r.purchase_price_after_discount),
                "purchase_price_per_unit_quote_currency": float(r.purchase_price_per_unit_quote_currency),
                "purchase_price_total_quote_currency": float(r.purchase_price_total_quote_currency),
                "logistics_first_leg": float(r.logistics_first_leg),
                "logistics_last_leg": float(r.logistics_last_leg),
                "logistics_total": float(r.logistics_total),
                "customs_fee": float(r.customs_fee),
                "excise_tax_amount": float(r.excise_tax_amount),
                "cogs_per_unit": float(r.cogs_per_unit),
                "cogs_per_product": float(r.cogs_per_product),
                "sale_price_per_unit_excl_financial": float(r.sale_price_per_unit_excl_financial),
                "sale_price_total_excl_financial": float(r.sale_price_total_excl_financial),
                "profit": float(r.profit),
                "dm_fee": float(r.dm_fee),
                "forex_reserve": float(r.forex_reserve),
                "financial_agent_fee": float(r.financial_agent_fee),
                "sales_price_per_unit_no_vat": float(r.sales_price_per_unit_no_vat),
                "sales_price_total_no_vat": float(r.sales_price_total_no_vat),
                "sales_price_total_with_vat": float(r.sales_price_total_with_vat),
                "sales_price_per_unit_with_vat": float(r.sales_price_per_unit_with_vat),
                "vat_from_sales": float(r.vat_from_sales),
                "vat_on_import": float(r.vat_on_import),
                "vat_net_payable": float(r.vat_net_payable),
                "transit_commission": float(r.transit_commission),
                "internal_sale_price_per_unit": float(r.internal_sale_price_per_unit),
                "internal_sale_price_total": float(r.internal_sale_price_total),
                "financing_cost_initial": float(r.financing_cost_initial),
                "financing_cost_credit": float(r.financing_cost_credit),
            })

        # ================================================================
        # SAVE TO DATABASE (if customer_id provided)
        # ================================================================
        quote_id = None
        quote_number = None

        if customer_id and user.current_organization_id:
            try:
                # Generate quote number
                quote_number = generate_quote_number(supabase, str(user.current_organization_id))

                # Get exchange rates snapshot for audit
                rates_snapshot = get_rates_snapshot_to_usd(date.today(), supabase)

                # Calculate VAT totals
                total_vat_on_import = sum(r.vat_on_import for r in calc_results)
                total_vat_payable = sum(r.vat_net_payable for r in calc_results)

                # Get quote currency to USD conversion rate
                # Calculation engine outputs in QUOTE CURRENCY, we need to convert to USD
                quote_currency = parsed_data.quote_currency.value
                if quote_currency == "USD":
                    quote_to_usd_rate = Decimal("1.0")
                else:
                    # rates contains {currency}/RUB rates
                    # quote_to_usd = quote_currency/RUB ÷ USD/RUB
                    usd_rub = rates.get("USD/RUB", Decimal("1.0"))
                    quote_rub = rates.get(f"{quote_currency}/RUB", Decimal("1.0"))
                    quote_to_usd_rate = quote_rub / usd_rub if usd_rub else Decimal("1.0")

                # Convert from quote currency to USD
                # total_revenue_* values are already in quote currency from calc engine
                total_revenue_no_vat_usd = total_revenue_no_vat * quote_to_usd_rate
                total_revenue_with_vat_usd = total_revenue_with_vat * quote_to_usd_rate
                total_profit_usd = total_profit * quote_to_usd_rate

                # 1. Create quote record
                quote_data = {
                    "organization_id": str(user.current_organization_id),
                    "customer_id": customer_id,
                    "quote_number": quote_number,
                    "title": f"КП от {date.today().strftime('%d.%m.%Y')}",
                    "status": "draft",
                    "created_by": str(user.id),
                    "quote_date": date.today().isoformat(),
                    "valid_until": (date.today() + timedelta(days=30)).isoformat(),
                    "currency": quote_currency,
                    "subtotal": float(total_purchase),
                    "total_amount": float(total_revenue_no_vat),  # In quote currency
                    # USD totals for display (converted from quote currency)
                    "total_usd": float(total_revenue_no_vat_usd),  # AK16 sum converted to USD
                    "total_with_vat_usd": float(total_revenue_with_vat_usd),  # AL16 sum converted to USD
                    "total_profit_usd": float(total_profit_usd),  # AF16 sum converted to USD
                    "total_vat_on_import_usd": float(total_vat_on_import * quote_to_usd_rate),
                    "total_vat_payable_usd": float(total_vat_payable * quote_to_usd_rate),
                    # Quote currency totals (already in quote currency from calc engine)
                    "total_amount_quote": float(total_revenue_no_vat),
                    "total_with_vat_quote": float(total_revenue_with_vat),
                }

                quote_response = supabase.table("quotes").insert(quote_data).execute()
                if not quote_response.data:
                    raise Exception("Failed to create quote")
                quote_id = quote_response.data[0]['id']

                # 2. Create quote_items records
                items_data = []
                for idx, p in enumerate(parsed_data.products):
                    custom_fields = {
                        "currency_of_base_price": p.currency.value,
                        "exchange_rate_base_price_to_quote": float(calculate_exchange_rate(
                            p.currency.value, parsed_data.quote_currency.value, rates, for_division=True
                        )),
                        "supplier_discount": float(p.supplier_discount) if p.supplier_discount else 0,
                        "markup": float(p.markup),
                        "import_tariff": float(p.import_tariff) if p.import_tariff else 0,
                    }
                    items_data.append({
                        "quote_id": quote_id,
                        "position": idx,
                        "product_name": p.name,
                        "product_code": p.sku or "",
                        "base_price_vat": float(p.base_price_vat),
                        "quantity": p.quantity,
                        "weight_in_kg": float(p.weight_kg) if p.weight_kg else 0,
                        "customs_code": p.customs_code or "",
                        "supplier_country": p.supplier_country or "Турция",
                        "custom_fields": custom_fields,
                    })

                items_response = supabase.table("quote_items").insert(items_data).execute()
                if not items_response.data:
                    raise Exception("Failed to create quote items")

                # 3. Save calculation variables
                variables_data = {
                    "quote_id": quote_id,
                    "variables": {
                        **quote_inputs,
                        "exchange_rates": {k: float(v) for k, v in rates.items()},
                        "rates_snapshot": rates_snapshot,
                    }
                }
                supabase.table("quote_calculation_variables").insert(variables_data).execute()

                # 4. Save calculation results for each product
                for idx, (r, item_record) in enumerate(zip(calc_results, items_response.data)):
                    result_dict = convert_decimals_to_float(r.dict())
                    result_dict["rates_snapshot"] = rates_snapshot

                    results_data = {
                        "quote_id": quote_id,
                        "quote_item_id": item_record['id'],
                        "phase_results": result_dict,
                    }
                    supabase.table("quote_calculation_results").insert(results_data).execute()

                # 5. Save quote calculation summary
                quote_summary = aggregate_product_results_to_summary(calc_results, quote_inputs)
                quote_summary["quote_id"] = quote_id
                supabase.table("quote_calculation_summaries").upsert(quote_summary).execute()

                logger.info(f"Quote {quote_number} (ID: {quote_id}) saved to database for customer {customer_id}")

            except Exception as e:
                logger.exception(f"Error saving quote to database: {e}")
                # Don't fail the whole request - still return validation Excel
                # but log the error

        # Generate validation Excel
        excel_bytes = generate_validation_export(
            quote_inputs=quote_inputs,
            product_inputs=product_inputs,
            api_results=api_results,
            product_results=product_results,
        )

        # Return as downloadable file
        base_name = file.filename.replace('.xlsx', '').replace('.xls', '').replace('.xlsm', '')
        filename = f"validation_{base_name}.xlsm"

        # RFC 5987 encoding for non-ASCII filenames
        # Use ASCII-safe filename as fallback, and UTF-8 encoded filename* for modern browsers
        safe_filename = f"validation_{url_quote(base_name, safe='')}.xlsm"
        filename_encoded = url_quote(filename, safe='')

        # Build response headers
        response_headers = {
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{filename_encoded}"
        }

        # Add quote info headers if saved to DB
        if quote_id:
            response_headers["X-Quote-Id"] = str(quote_id)
            # URL-encode quote_number for HTTP header (contains Cyrillic "КП")
            response_headers["X-Quote-Number"] = url_quote(quote_number, safe='')
            # Expose custom headers to frontend
            response_headers["Access-Control-Expose-Headers"] = "X-Quote-Id, X-Quote-Number"

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            headers=response_headers
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Error generating validation export")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )


# ============================================================================
# TEMPLATE DOWNLOAD & EXPORT ENDPOINTS
# ============================================================================

# Path to blank template (relative to backend directory for Railway deployment)
# Railway deploys only the backend/ directory, so templates must be inside backend/
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "templates",
    "validation",
    "template_quote_input_v6_test.xlsx"
)


def export_quote_as_template(
    quote: dict,
    quote_items: list,
    variables: dict,
) -> bytes:
    """
    Export a quote to the simplified input template format.

    Args:
        quote: Quote data from database
        quote_items: List of quote items from database
        variables: Quote calculation variables from database

    Returns:
        Excel file bytes
    """
    import openpyxl
    from pathlib import Path

    # Load the blank template
    template_path = Path(TEMPLATE_PATH)
    if not template_path.exists():
        raise ValueError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Котировка"] if "Котировка" in wb.sheetnames else wb.active

    # ========== Section 1: Quote Settings (B2-B7) ==========
    ws["B2"] = variables.get("seller_company", "МАСТЕР БЭРИНГ ООО")
    ws["B3"] = variables.get("offer_sale_type", "поставка")
    ws["B4"] = variables.get("offer_incoterms", "DDP")
    ws["B5"] = quote.get("currency") or variables.get("currency_of_quote", "EUR")
    ws["B6"] = variables.get("delivery_time", 30)
    ws["B7"] = float(variables.get("advance_to_supplier", 100)) / 100  # Convert to decimal for Excel %

    # ========== Section 2: Payment Terms (E3-F7) ==========
    advance_from_client = float(variables.get("advance_from_client", 100))
    # Convert percentage (0-100) to decimal (0-1) for Excel
    ws["E3"] = advance_from_client / 100
    ws["F3"] = int(variables.get("time_to_advance", 0))

    # Other payment milestones - typically 0 for most quotes
    ws["E4"] = 0  # On loading
    ws["E5"] = 0  # On shipping
    ws["E6"] = 0  # On customs
    # E7 is formula =1-SUM(E3:E6), don't overwrite
    ws["F7"] = int(variables.get("time_to_payment", 0))

    # ========== Section 3: Logistics (I3-J5) ==========
    ws["I3"] = float(variables.get("logistics_supplier_hub", 0))
    ws["J3"] = variables.get("logistics_supplier_hub_currency", "EUR")

    ws["I4"] = float(variables.get("logistics_hub_customs", 0))
    ws["J4"] = variables.get("logistics_hub_customs_currency", "EUR")

    ws["I5"] = float(variables.get("logistics_customs_client", 0))
    ws["J5"] = variables.get("logistics_customs_client_currency", "RUB")

    # ========== Section 4: Brokerage (I8-J12) ==========
    ws["I8"] = float(variables.get("brokerage_hub", 0))
    ws["J8"] = variables.get("brokerage_hub_currency", "EUR")

    ws["I9"] = float(variables.get("brokerage_customs", 0))
    ws["J9"] = variables.get("brokerage_customs_currency", "RUB")

    ws["I10"] = float(variables.get("warehousing", 0))
    ws["J10"] = variables.get("warehousing_currency", "RUB")

    ws["I11"] = float(variables.get("documentation", 0))
    ws["J11"] = variables.get("documentation_currency", "RUB")

    ws["I12"] = float(variables.get("other_costs", 0))
    ws["J12"] = variables.get("other_costs_currency", "RUB")

    # ========== Section 5: DM Fee / LPR (E11-F12) ==========
    dm_fee_type = variables.get("dm_fee_type", "% от суммы")
    ws["E11"] = dm_fee_type

    dm_fee_value = float(variables.get("dm_fee_value", 0))
    if dm_fee_type == "% от суммы":
        ws["E12"] = dm_fee_value / 100  # Convert to decimal for Excel %
    else:
        ws["E12"] = dm_fee_value
        ws["F12"] = variables.get("dm_fee_currency", "EUR")

    # ========== Section 6: Products (Row 16+) ==========
    # Product columns: A=brand, B=sku, C=name, D=quantity, E=weight,
    # F=currency, G=base_price_vat, H=supplier_country, I=discount,
    # J=customs_code, K=import_tariff, L=markup

    row = 16
    for item in quote_items:
        ws[f"A{row}"] = item.get("brand", "")
        ws[f"B{row}"] = item.get("sku", "")
        ws[f"C{row}"] = item.get("product_name", "")
        ws[f"D{row}"] = int(item.get("quantity", 1))
        ws[f"E{row}"] = float(item.get("weight_in_kg", 0)) if item.get("weight_in_kg") else None
        ws[f"F{row}"] = item.get("currency_of_base_price", "EUR")
        ws[f"G{row}"] = float(item.get("base_price_vat", 0))
        ws[f"H{row}"] = item.get("supplier_country", "Турция")

        # Discount as decimal for Excel %
        supplier_discount = float(item.get("supplier_discount", 0))
        ws[f"I{row}"] = supplier_discount / 100 if supplier_discount > 1 else supplier_discount

        ws[f"J{row}"] = item.get("customs_code") or ""

        # Import tariff as decimal for Excel %
        import_tariff = float(item.get("import_tariff", 0))
        ws[f"K{row}"] = import_tariff / 100 if import_tariff > 1 else import_tariff

        # Markup as decimal for Excel %
        markup = float(item.get("markup", 15))
        ws[f"L{row}"] = markup / 100 if markup > 1 else markup

        row += 1

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/upload/download-template")
async def download_template(
    user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase),
):
    """
    Download blank quote input template.

    Returns the template Excel file that users can fill in and upload.
    """
    from pathlib import Path

    template_path = Path(TEMPLATE_PATH)
    if not template_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Template file not found"
        )

    with open(template_path, "rb") as f:
        content = f.read()

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=\"quote_input_template.xlsx\""
        }
    )


@router.get("/upload/export-as-template/{quote_id}")
async def export_quote_as_template_endpoint(
    quote_id: str,
    user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase),
):
    """
    Export an existing quote as the big validation Excel file (.xlsm).

    This regenerates the same 1.5MB+ Excel file that was returned during upload,
    using the stored quote data and calculation results from the database.
    """
    if not user.current_organization_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User is not associated with any organization"
        )

    try:
        # Get quote and verify organization ownership
        quote_result = supabase.table("quotes").select("*").eq("id", quote_id).execute()
        if not quote_result.data:
            raise HTTPException(status_code=404, detail="Quote not found")

        quote = quote_result.data[0]
        if quote["organization_id"] != str(user.current_organization_id):
            raise HTTPException(status_code=403, detail="Not authorized")

        # Get quote items (these become product_inputs)
        items_result = supabase.table("quote_items").select("*").eq("quote_id", quote_id).order("position").execute()
        quote_items = items_result.data or []

        if not quote_items:
            raise HTTPException(status_code=404, detail="No items found for quote")

        # Get calculation variables (these become quote_inputs)
        vars_result = supabase.table("quote_calculation_variables").select("*").eq("quote_id", quote_id).execute()
        if not vars_result.data:
            raise HTTPException(status_code=404, detail="No calculation variables found for quote")

        stored_variables = vars_result.data[0].get("variables", {})

        # Build quote_inputs from stored variables
        # Note: export_validation_service expects specific field names
        quote_currency = quote.get("currency") or stored_variables.get("currency_of_quote", "EUR")
        quote_inputs = {
            "seller_company": stored_variables.get("seller_company", "МАСТЕР БЭРИНГ ООО"),
            "offer_sale_type": stored_variables.get("offer_sale_type", "поставка"),
            "incoterms": stored_variables.get("offer_incoterms", "DDP"),  # export service expects "incoterms"
            "quote_currency": quote_currency,  # export service expects "quote_currency"
            "delivery_time": stored_variables.get("delivery_time", 30),
            "advance_to_supplier": stored_variables.get("advance_to_supplier", 100),
            "advance_from_client": stored_variables.get("advance_from_client", 100),
            "time_to_advance": stored_variables.get("time_to_advance", 0),
            "time_to_payment": stored_variables.get("time_to_payment", 0),
            "logistics_supplier_hub": stored_variables.get("logistics_supplier_hub", 0),
            "logistics_supplier_hub_currency": stored_variables.get("logistics_supplier_hub_currency", "EUR"),
            "logistics_hub_customs": stored_variables.get("logistics_hub_customs", 0),
            "logistics_hub_customs_currency": stored_variables.get("logistics_hub_customs_currency", "EUR"),
            "logistics_customs_client": stored_variables.get("logistics_customs_client", 0),
            "logistics_customs_client_currency": stored_variables.get("logistics_customs_client_currency", "RUB"),
            "brokerage_hub": stored_variables.get("brokerage_hub", 0),
            "brokerage_hub_currency": stored_variables.get("brokerage_hub_currency", "EUR"),
            "brokerage_customs": stored_variables.get("brokerage_customs", 0),
            "brokerage_customs_currency": stored_variables.get("brokerage_customs_currency", "RUB"),
            "warehousing": stored_variables.get("warehousing", 0),
            "warehousing_currency": stored_variables.get("warehousing_currency", "RUB"),
            "documentation": stored_variables.get("documentation", 0),
            "documentation_currency": stored_variables.get("documentation_currency", "RUB"),
            "other_costs": stored_variables.get("other_costs", 0),
            "other_costs_currency": stored_variables.get("other_costs_currency", "RUB"),
            "dm_fee_type": stored_variables.get("dm_fee_type", "% от суммы"),
            "dm_fee_value": stored_variables.get("dm_fee_value", 0),
            "dm_fee_currency": stored_variables.get("dm_fee_currency", "EUR"),
            "rate_forex_risk": stored_variables.get("rate_forex_risk", 0),
            "rate_fin_comm": stored_variables.get("rate_fin_comm", 0),
            "rate_loan_interest_daily": stored_variables.get("rate_loan_interest_daily", 0),
            "exchange_rates": stored_variables.get("exchange_rates", {}),
        }

        # Build product_inputs from quote_items
        # Note: Some fields are stored in custom_fields JSONB column
        product_inputs = []
        for item in quote_items:
            custom_fields = item.get("custom_fields", {}) or {}

            # Get values from custom_fields (product-level overrides) or use defaults
            currency_of_base_price = custom_fields.get("currency_of_base_price", "EUR")
            exchange_rate = custom_fields.get("exchange_rate_base_price_to_quote", 1.0)
            supplier_discount = custom_fields.get("supplier_discount", 0)
            import_tariff = custom_fields.get("import_tariff", 0)
            markup = custom_fields.get("markup", 15)

            product_inputs.append({
                "brand": item.get("brand", ""),
                "sku": item.get("sku", ""),
                "name": item.get("product_name", ""),  # export service expects "name" not "product_name"
                "quantity": item.get("quantity", 1),
                "weight_in_kg": item.get("weight_in_kg") or 0,
                "currency_of_base_price": currency_of_base_price,
                "base_price_vat": item.get("base_price_vat", 0),
                "supplier_country": item.get("supplier_country", "Турция"),
                "supplier_discount": supplier_discount,
                "exchange_rate": exchange_rate,  # required for validation export
                "customs_code": item.get("customs_code", ""),
                "import_tariff": import_tariff,  # from custom_fields
                "markup": markup,  # from custom_fields
            })

        # Get calculation results from DB (these become product_results)
        calc_results = supabase.table("quote_calculation_results").select("*").eq("quote_id", quote_id).execute()
        calc_by_item = {r["quote_item_id"]: r.get("phase_results", {}) for r in calc_results.data or []}

        product_results = []
        for item in quote_items:
            item_id = item.get("id")
            if item_id in calc_by_item:
                product_results.append(calc_by_item[item_id])
            else:
                product_results.append({})  # Empty if not found

        # Get quote summary (for api_results)
        # The DB uses column names like "calc_s16_total_purchase_price"
        # but export service expects "total_purchase_price"
        summary_result = supabase.table("quote_calculation_summaries").select("*").eq("quote_id", quote_id).execute()
        raw_summary = summary_result.data[0] if summary_result.data else {}

        # Map DB column names to export service expected field names
        api_results = {
            # Quote totals (QUOTE_TOTAL_CELLS in export_validation_service.py)
            "total_purchase_price": raw_summary.get("calc_s16_total_purchase_price", 0),
            "total_logistics_first": raw_summary.get("calc_t16_first_leg_logistics", 0),
            "total_logistics_last": raw_summary.get("calc_u16_last_leg_logistics", 0),
            "total_logistics": raw_summary.get("calc_v16_total_logistics", 0),
            "total_cogs": raw_summary.get("calc_ab16_cogs_total", 0),
            "total_revenue": raw_summary.get("calc_ak16_final_price_total", 0),
            "total_revenue_with_vat": raw_summary.get("calc_al16_total_with_vat", 0),
            # total_profit = revenue - cogs
            "total_profit": (
                float(raw_summary.get("calc_ak16_final_price_total", 0) or 0)
                - float(raw_summary.get("calc_ab16_cogs_total", 0) or 0)
            ),
            # Financing (FINANCING_CELLS in export_validation_service.py)
            "evaluated_revenue": raw_summary.get("calc_bh2_revenue_estimated", 0),
            "client_advance": raw_summary.get("calc_bh3_client_advance", 0),
            "total_before_forwarding": raw_summary.get("calc_bh4_before_forwarding", 0),
            "supplier_payment": raw_summary.get("calc_bh6_supplier_payment", 0),
            "supplier_financing_cost": raw_summary.get("calc_bj7_supplier_financing_cost", 0),
            "operational_financing_cost": raw_summary.get("calc_bj10_operational_cost", 0),
            "total_financing_cost": raw_summary.get("calc_bj11_total_financing_cost", 0),
            "credit_sales_amount": raw_summary.get("calc_bl3_credit_sales_amount", 0),
            "credit_sales_fv": raw_summary.get("calc_bl4_credit_sales_with_interest", 0),
            "credit_sales_interest": raw_summary.get("calc_bl5_credit_sales_interest", 0),
        }

        # Generate the big validation Excel file
        excel_bytes = generate_validation_export(
            quote_inputs=quote_inputs,
            product_inputs=product_inputs,
            api_results=api_results,
            product_results=product_results,
        )

        # Build filename
        quote_number = quote.get("quote_number", quote_id)
        safe_quote_number = ''.join(c if ord(c) < 128 else '_' for c in str(quote_number))
        filename = f"validation_{safe_quote_number}.xlsm"
        filename_encoded = url_quote(f"validation_{quote_number}.xlsm", safe='')

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{filename_encoded}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error exporting quote as validation Excel")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exporting quote: {str(e)}"
        )
