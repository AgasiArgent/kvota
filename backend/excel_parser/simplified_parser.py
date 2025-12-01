"""
Simplified Excel Template Parser

Parses the simplified quote input template (template_quote_input_v5.xlsx).
This template has a horizontal layout different from the reference Excel.

Template Structure:
- Row 1: Section headers (НАСТРОЙКИ, УСЛОВИЯ ОПЛАТЫ, ЛОГИСТИКА)
- Rows 2-7: Quote settings (A-B), Payment terms (D-F), Logistics (H-J)
- Row 6: БРОКЕРСКИЕ УСЛУГИ header
- Rows 7-12: Brokerage content (H-J)
- Rows 10-12: LPR section (D-F)
- Row 14: ТОВАРЫ header
- Row 15: Product column headers
- Row 16+: Product data

Created: 2025-11-28
"""

import openpyxl
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Any
from decimal import Decimal, InvalidOperation
from pydantic import BaseModel, Field, validator
from enum import Enum


# ============================================================================
# ENUMS (must match calculation_models.py)
# ============================================================================

class Currency(str, Enum):
    USD = "USD"
    EUR = "EUR"
    RUB = "RUB"
    TRY = "TRY"
    CNY = "CNY"


class SupplierCountry(str, Enum):
    TURKEY = "Турция"
    TURKEY_TRANSIT = "Турция (транзитная зона)"
    RUSSIA = "Россия"
    CHINA = "Китай"
    LITHUANIA = "Литва"
    LATVIA = "Латвия"
    BULGARIA = "Болгария"
    POLAND = "Польша"
    EU_CROSS_BORDER = "ЕС (закупка между странами ЕС)"
    UAE = "ОАЭ"
    OTHER = "Прочие"


class SellerCompany(str, Enum):
    MASTER_BEARING_RU = "МАСТЕР БЭРИНГ ООО"
    CMTO1_RU = "ЦМТО1 ООО"
    RAD_RESURS_RU = "РАД РЕСУРС ООО"
    TEXCEL_TR = "TEXCEL OTOMOTİV TİCARET LİMİTED ŞİRKETİ"
    GESTUS_TR = "GESTUS DIŞ TİCARET LİMİTED ŞİRKETİ"
    UPDOOR_CN = "UPDOOR Limited"


class OfferSaleType(str, Enum):
    SUPPLY = "поставка"
    TRANSIT = "транзит"
    FIN_TRANSIT = "финтранзит"
    EXPORT = "экспорт"


class Incoterms(str, Enum):
    DDP = "DDP"
    DAP = "DAP"
    CIF = "CIF"
    FOB = "FOB"
    EXW = "EXW"


class DMFeeType(str, Enum):
    PERCENTAGE = "% от суммы"
    FIXED = "фикс. сумма"


# ============================================================================
# PYDANTIC MODELS
# ============================================================================

class MonetaryValue(BaseModel):
    """Amount with currency"""
    value: Decimal
    currency: Currency


class PaymentMilestone(BaseModel):
    """Single payment milestone"""
    name: str
    percentage: Decimal = Field(ge=0, le=100)
    days: Optional[int] = None


class ProductInput(BaseModel):
    """Single product from template"""
    brand: Optional[str] = None
    sku: Optional[str] = None
    name: str
    quantity: int = Field(gt=0)
    weight_kg: Optional[Decimal] = Field(default=None, ge=0)
    currency: Currency
    base_price_vat: Decimal = Field(gt=0)
    supplier_country: str
    supplier_discount: Decimal = Field(default=Decimal("0"), ge=0, le=100)
    customs_code: Optional[int] = None
    import_tariff: Decimal = Field(default=Decimal("0"), ge=0, le=100)
    markup: Decimal = Field(default=Decimal("15"), ge=0)


class SimplifiedQuoteInput(BaseModel):
    """Complete parsed quote from simplified template"""
    # Quote settings
    seller_company: str
    sale_type: str
    incoterms: str
    quote_currency: Currency
    delivery_time: int = Field(gt=0)
    advance_to_supplier: Decimal = Field(ge=0, le=100)

    # Payment terms
    payment_milestones: List[PaymentMilestone]

    # Logistics costs
    logistics_supplier_hub: MonetaryValue
    logistics_hub_customs: MonetaryValue
    logistics_customs_client: MonetaryValue

    # Brokerage costs
    brokerage_hub: MonetaryValue
    brokerage_customs: MonetaryValue
    warehousing: MonetaryValue
    documentation: MonetaryValue
    other_costs: MonetaryValue

    # DM Fee (LPR)
    dm_fee_type: str
    dm_fee_value: Decimal = Field(ge=0)
    dm_fee_currency: Optional[Currency] = None  # Only for fixed type

    # Products
    products: List[ProductInput]

    @validator('payment_milestones')
    def validate_payment_totals(cls, milestones):
        total = sum(m.percentage for m in milestones)
        if abs(total - Decimal("100")) > Decimal("0.01"):
            raise ValueError(f"Payment milestones must sum to 100%, got {total}%")
        return milestones


# ============================================================================
# PARSER
# ============================================================================

class SimplifiedExcelParser:
    """
    Parser for simplified quote input template.

    Usage:
        parser = SimplifiedExcelParser(file_path_or_bytes)
        quote_input = parser.parse()
    """

    def __init__(self, source):
        """
        Initialize parser.

        Args:
            source: File path (str) or file-like object (BytesIO)
        """
        if isinstance(source, str):
            self.workbook = openpyxl.load_workbook(source, data_only=True)
        else:
            self.workbook = openpyxl.load_workbook(source, data_only=True)

        self.sheet = self._find_sheet()
        self.errors: List[str] = []

    def _find_sheet(self):
        """Find the quote sheet"""
        # Try exact name first
        if "Котировка" in self.workbook.sheetnames:
            return self.workbook["Котировка"]

        # Try first sheet
        return self.workbook.active

    def _get_value(self, cell: str, default=None):
        """Get cell value with default"""
        value = self.sheet[cell].value
        return value if value is not None else default

    def _get_decimal(self, cell: str, default: Decimal = Decimal("0")) -> Decimal:
        """Get cell value as Decimal"""
        value = self.sheet[cell].value
        if value is None:
            return default
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            self.errors.append(f"Invalid number in cell {cell}: {value}")
            return default

    def _get_int(self, cell: str, default: int = 0) -> int:
        """Get cell value as int"""
        value = self.sheet[cell].value
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            self.errors.append(f"Invalid integer in cell {cell}: {value}")
            return default

    def _get_currency(self, cell: str, default: Currency = Currency.EUR) -> Currency:
        """Get cell value as Currency enum"""
        value = self.sheet[cell].value
        if value is None:
            return default
        try:
            return Currency(value.upper() if isinstance(value, str) else value)
        except ValueError:
            self.errors.append(f"Invalid currency in cell {cell}: {value}")
            return default

    def parse(self) -> SimplifiedQuoteInput:
        """
        Parse the Excel template and return structured data.

        Returns:
            SimplifiedQuoteInput: Parsed and validated quote data

        Raises:
            ValueError: If validation fails
        """
        self.errors = []

        # Parse each section
        settings = self._parse_settings()
        payment = self._parse_payment_terms()
        logistics = self._parse_logistics()
        brokerage = self._parse_brokerage()
        lpr = self._parse_lpr()
        products = self._parse_products()

        if self.errors:
            raise ValueError(f"Parse errors:\n" + "\n".join(self.errors))

        # Build the complete input
        return SimplifiedQuoteInput(
            # Settings
            seller_company=settings["seller_company"],
            sale_type=settings["sale_type"],
            incoterms=settings["incoterms"],
            quote_currency=settings["quote_currency"],
            delivery_time=settings["delivery_time"],
            advance_to_supplier=settings["advance_to_supplier"],

            # Payment
            payment_milestones=payment,

            # Logistics
            logistics_supplier_hub=logistics["supplier_hub"],
            logistics_hub_customs=logistics["hub_customs"],
            logistics_customs_client=logistics["customs_client"],

            # Brokerage
            brokerage_hub=brokerage["hub"],
            brokerage_customs=brokerage["customs"],
            warehousing=brokerage["warehousing"],
            documentation=brokerage["documentation"],
            other_costs=brokerage["other"],

            # LPR
            dm_fee_type=lpr["type"],
            dm_fee_value=lpr["value"],
            dm_fee_currency=lpr.get("currency"),

            # Products
            products=products
        )

    def _parse_settings(self) -> Dict[str, Any]:
        """Parse НАСТРОЙКИ КОТИРОВКИ section (A2-B7)"""
        return {
            "seller_company": self._get_value("B2", ""),
            "sale_type": self._get_value("B3", "поставка"),
            "incoterms": self._get_value("B4", "DDP"),
            "quote_currency": self._get_currency("B5", Currency.EUR),
            "delivery_time": self._get_int("B6", 30),
            "advance_to_supplier": self._get_decimal("B7", Decimal("100")),
        }

    def _parse_payment_terms(self) -> List[PaymentMilestone]:
        """Parse УСЛОВИЯ ОПЛАТЫ section (D2-F7)"""
        milestones = []

        # Row 3: Аванс от клиента
        milestones.append(PaymentMilestone(
            name="advance_from_client",
            percentage=self._get_decimal("E3", Decimal("0")),
            days=self._get_int("F3", 0)
        ))

        # Row 4: При заборе груза
        milestones.append(PaymentMilestone(
            name="advance_on_loading",
            percentage=self._get_decimal("E4", Decimal("0")),
            days=self._get_int("F4") if self._get_value("F4") else None
        ))

        # Row 5: При отправке в РФ
        milestones.append(PaymentMilestone(
            name="advance_on_shipping",
            percentage=self._get_decimal("E5", Decimal("0")),
            days=self._get_int("F5") if self._get_value("F5") else None
        ))

        # Row 6: При таможне
        milestones.append(PaymentMilestone(
            name="advance_on_customs",
            percentage=self._get_decimal("E6", Decimal("0")),
            days=self._get_int("F6") if self._get_value("F6") else None
        ))

        # Row 7: При получении (auto-calculated)
        # Note: Formula cells may not be evaluated in files created by openpyxl
        # So we calculate: 100% - sum of all other milestones
        other_percentages = sum(m.percentage for m in milestones)
        on_receiving_pct = Decimal("100") - other_percentages

        milestones.append(PaymentMilestone(
            name="on_receiving",
            percentage=on_receiving_pct,
            days=self._get_int("F7", 0)
        ))

        return milestones

    def _parse_logistics(self) -> Dict[str, MonetaryValue]:
        """Parse ЛОГИСТИКА section (H2-J5)"""
        return {
            "supplier_hub": MonetaryValue(
                value=self._get_decimal("I3", Decimal("0")),
                currency=self._get_currency("J3", Currency.EUR)
            ),
            "hub_customs": MonetaryValue(
                value=self._get_decimal("I4", Decimal("0")),
                currency=self._get_currency("J4", Currency.EUR)
            ),
            "customs_client": MonetaryValue(
                value=self._get_decimal("I5", Decimal("0")),
                currency=self._get_currency("J5", Currency.RUB)
            ),
        }

    def _parse_brokerage(self) -> Dict[str, MonetaryValue]:
        """Parse БРОКЕРСКИЕ УСЛУГИ section (H6-J12)"""
        return {
            "hub": MonetaryValue(
                value=self._get_decimal("I8", Decimal("0")),
                currency=self._get_currency("J8", Currency.EUR)
            ),
            "customs": MonetaryValue(
                value=self._get_decimal("I9", Decimal("0")),
                currency=self._get_currency("J9", Currency.RUB)
            ),
            "warehousing": MonetaryValue(
                value=self._get_decimal("I10", Decimal("0")),
                currency=self._get_currency("J10", Currency.RUB)
            ),
            "documentation": MonetaryValue(
                value=self._get_decimal("I11", Decimal("0")),
                currency=self._get_currency("J11", Currency.RUB)
            ),
            "other": MonetaryValue(
                value=self._get_decimal("I12", Decimal("0")),
                currency=self._get_currency("J12", Currency.RUB)
            ),
        }

    def _parse_lpr(self) -> Dict[str, Any]:
        """Parse ВОЗНАГРАЖДЕНИЕ ЛПР section (D10-F12)"""
        fee_type = self._get_value("E11", "% от суммы")
        fee_value = self._get_decimal("E12", Decimal("0"))

        result = {
            "type": fee_type,
            "value": fee_value,
        }

        # Currency only matters for fixed amount
        if fee_type == "фикс. сумма":
            result["currency"] = self._get_currency("F12", Currency.EUR)

        return result

    def _parse_products(self) -> List[ProductInput]:
        """Parse ТОВАРЫ section (A14+)"""
        products = []
        row = 16  # First product row

        # Continue while there's data in column A (brand) or D (name)
        while True:
            brand = self._get_value(f"A{row}")
            name = self._get_value(f"C{row}")
            quantity = self._get_value(f"D{row}")

            # Stop if row is empty
            if not name and not quantity:
                break

            # Skip if no name
            if not name:
                row += 1
                continue

            product = ProductInput(
                brand=brand,
                sku=self._get_value(f"B{row}"),
                name=str(name),
                quantity=self._get_int(f"D{row}", 1),
                weight_kg=self._get_decimal(f"E{row}") if self._get_value(f"E{row}") else None,
                currency=self._get_currency(f"F{row}", Currency.EUR),
                base_price_vat=self._get_decimal(f"G{row}", Decimal("0")),
                supplier_country=self._get_value(f"H{row}", "Турция"),
                supplier_discount=self._get_decimal(f"I{row}", Decimal("0")),
                customs_code=self._get_int(f"J{row}") if self._get_value(f"J{row}") else None,
                import_tariff=self._get_decimal(f"K{row}", Decimal("0")),
                markup=self._get_decimal(f"L{row}", Decimal("15")),
            )

            products.append(product)
            row += 1

            # Safety limit
            if row > 1000:
                break

        return products


def parse_simplified_template(source) -> SimplifiedQuoteInput:
    """
    Convenience function to parse simplified template.

    Args:
        source: File path (str) or file-like object (BytesIO)

    Returns:
        SimplifiedQuoteInput: Parsed and validated data
    """
    parser = SimplifiedExcelParser(source)
    return parser.parse()
