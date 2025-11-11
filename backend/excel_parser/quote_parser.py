import openpyxl
import os
from dataclasses import dataclass
from typing import Dict, List, Optional, Any
from openpyxl.worksheet.worksheet import Worksheet

@dataclass
class QuoteData:
    """Parsed quote data from Excel"""
    filename: str
    sheet_name: str
    inputs: Dict[str, Any]
    expected_results: Dict[str, Any]

class ExcelQuoteParser:
    """Parse quotes from Excel files with formula-based layout"""

    def __init__(self, filepath: str) -> None:
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath, data_only=True)
        self.sheet = self._find_calculation_sheet()

    def _find_calculation_sheet(self):
        """Find calculation sheet with 3-level fallback"""
        # Strategy 1: Try exact name "Расчет"
        if "Расчет" in self.workbook.sheetnames:
            sheet = self.workbook["Расчет"]
            if self._validate_sheet_structure(sheet):
                return sheet

        # Strategy 2: Try similar names
        similar_names = ["расчет", "Расчёт", "расчёт", "Calculation", "calc"]
        for name in self.workbook.sheetnames:
            if any(similar in name.lower() for similar in similar_names):
                sheet = self.workbook[name]
                if self._validate_sheet_structure(sheet):
                    return sheet

        # Strategy 3: Search all sheets for markers
        for sheet in self.workbook.worksheets:
            if self._validate_sheet_structure(sheet):
                return sheet

        # Fail with clear error
        available = ", ".join(self.workbook.sheetnames)
        raise ValueError(
            f"Cannot find calculation sheet in {self.filepath}.\n"
            f"Available sheets: {available}\n"
            f"Expected sheet with cells D5, E16, K16"
        )

    def _validate_sheet_structure(self, sheet: Worksheet) -> bool:
        """Check if sheet has expected structure"""
        try:
            # Check E16 is number (quantity)
            e16 = sheet["E16"].value
            if e16 and isinstance(e16, (int, float)) and e16 > 0:
                # Check K16 is number (price)
                k16 = sheet["K16"].value
                if k16 and isinstance(k16, (int, float)) and k16 > 0:
                    return True
            return False
        except Exception:
            return False

    def parse(self) -> QuoteData:
        """Extract complete quote data from Excel"""
        return QuoteData(
            filename=os.path.basename(self.filepath),
            sheet_name=self.sheet.title,
            inputs=self._extract_inputs(),
            expected_results=self._extract_results()
        )

    def _extract_inputs(self) -> Dict[str, Any]:
        """Extract all input variables"""
        # Quote-level (fixed cells)
        quote_vars = {
            "seller_company": self.sheet["D5"].value,
            "offer_sale_type": self.sheet["D6"].value,
            "offer_incoterms": self.sheet["D7"].value,
            "currency_of_quote": self.sheet["D8"].value,
            "delivery_time": self.sheet["D9"].value,  # Days
            "advance_to_supplier": self.sheet["D11"].value,  # %
            "advance_from_client": self.sheet["J5"].value,
            "time_to_advance": self.sheet["K5"].value,

            # Financial params (critical for accuracy)
            "rate_forex_risk": self.sheet["AH11"].value,  # Forex risk %
            # Note: rate_fin_comm and rate_loan_interest_daily are on separate sheet (helpsheet)
            # BL5 is intermediate calculation value, not input
            "dm_fee_type": self.sheet["AG3"].value,  # DM fee type
            "dm_fee_value": self.sheet["AG7"].value,  # DM fee value

            # Logistics (quote-level totals)
            "logistics_supplier_hub": self.sheet["W2"].value,
            "logistics_hub_customs": self.sheet["W3"].value,
            "logistics_customs_client": self.sheet["W4"].value,
            "brokerage_hub": self.sheet["W5"].value,
            "brokerage_customs": self.sheet["W6"].value,
            "warehousing_at_customs": self.sheet["W7"].value,
            "customs_documentation": self.sheet["W8"].value,
            "brokerage_extra": self.sheet["W9"].value,
            "insurance": self.sheet["W10"].value,  # Insurance (quote-level)
        }

        # Products (dynamic rows starting from 16)
        products = []
        row = 16  # First product

        while self.sheet[f"E{row}"].value:  # While quantity exists
            base_price = self.sheet[f"K{row}"].value

            # Skip products with zero or null price (empty rows, bonuses, etc.)
            if base_price is None or base_price == 0:
                row += 1
                continue

            product = {
                "quantity": self.sheet[f"E{row}"].value,
                "weight_in_kg": self.sheet[f"G{row}"].value,
                "currency_of_base_price": self.sheet[f"J{row}"].value,
                "base_price_VAT": base_price,
                "supplier_country": self.sheet[f"L{row}"].value,
                "supplier_discount": self.sheet[f"O{row}"].value,
                "exchange_rate_base_price_to_quote": self.sheet[f"Q{row}"].value,
                "customs_code": self.sheet[f"W{row}"].value,
                "import_tariff": self.sheet[f"X{row}"].value,
                "excise_tax": self.sheet[f"Z{row}"].value,

                # Product-level overrides (critical for accuracy)
                "markup": self.sheet[f"AC{row}"].value,  # Markup % per product
            }
            products.append(product)
            row += 1

        return {"quote": quote_vars, "products": products}

    def _extract_results(self) -> Dict[str, Any]:
        """Extract calculated results for comparison"""
        results = []
        row = 16

        while self.sheet[f"E{row}"].value:
            base_price = self.sheet[f"K{row}"].value

            # Skip products with zero or null price (must match _extract_inputs logic)
            if base_price is None or base_price == 0:
                row += 1
                continue

            result = {
                # Summary fields
                "AK16": self.sheet[f"AK{row}"].value,  # Final price
                "AM16": self.sheet[f"AM{row}"].value,  # With VAT
                "AQ16": self.sheet[f"AQ{row}"].value,  # Profit

                # Detailed fields (for detailed mode)
                "M16": self.sheet[f"M{row}"].value,
                "S16": self.sheet[f"S{row}"].value,
                "T16": self.sheet[f"T{row}"].value,
                "V16": self.sheet[f"V{row}"].value,
                "Y16": self.sheet[f"Y{row}"].value,
                "AB16": self.sheet[f"AB{row}"].value,
            }
            results.append(result)
            row += 1

        return {"products": results}
