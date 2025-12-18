"""
Secure File Processing Service for Russian B2B Quotation System
Handles Excel/CSV import with security validation and Russian business context
"""
import io
import re
import hashlib
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook


class SecureFileProcessor:
    """
    Secure file processing with Russian business context
    Handles Excel and CSV imports with proper validation
    """

    # Security settings
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}
    MAX_ROWS = 1000  # Maximum rows to process

    # Russian column mappings
    RUSSIAN_COLUMN_MAP = {
        # Russian terms
        'наименование': 'description',
        'название': 'description',
        'товар': 'description',
        'продукт': 'description',
        'услуга': 'description',
        'описание': 'description',

        'количество': 'quantity',
        'кол-во': 'quantity',
        'штук': 'quantity',
        'шт': 'quantity',
        'кол': 'quantity',

        'цена': 'unit_price',
        'стоимость': 'unit_price',
        'цена за ед': 'unit_price',
        'цена за единицу': 'unit_price',
        'прайс': 'unit_price',

        'сумма': 'line_total',
        'общая стоимость': 'line_total',
        'итого': 'line_total',
        'всего': 'line_total',

        'единица измерения': 'unit',
        'ед изм': 'unit',
        'ед.изм': 'unit',
        'ед.изм.': 'unit',
        'единица': 'unit',

        'категория': 'category',
        'группа': 'category',
        'тип': 'category',
        'класс': 'category',

        'бренд': 'brand',
        'марка': 'brand',
        'производитель': 'brand',
        'изготовитель': 'brand',

        'примечание': 'notes',
        'примечания': 'notes',
        'комментарий': 'notes',
        'комментарии': 'notes',
        'заметки': 'notes',

        'артикул': 'product_code',
        'код': 'product_code',
        'код товара': 'product_code',
        'номер': 'product_code',

        'страна': 'country_of_origin',
        'страна происхождения': 'country_of_origin',
        'происхождение': 'country_of_origin',

        # English terms
        'name': 'description',
        'title': 'description',
        'description': 'description',
        'product': 'description',
        'item': 'description',

        'quantity': 'quantity',
        'qty': 'quantity',
        'amount': 'quantity',
        'count': 'quantity',

        'price': 'unit_price',
        'unit_price': 'unit_price',
        'cost': 'unit_price',
        'rate': 'unit_price',

        'total': 'line_total',
        'line_total': 'line_total',
        'sum': 'line_total',
        'total_price': 'line_total',

        'unit': 'unit',
        'measure': 'unit',
        'uom': 'unit',

        'category': 'category',
        'group': 'category',
        'type': 'category',

        'brand': 'brand',
        'manufacturer': 'brand',
        'vendor': 'brand',

        'notes': 'notes',
        'comment': 'notes',
        'remarks': 'notes',

        'sku': 'product_code',
        'code': 'product_code',
        'product_code': 'product_code',
        'item_code': 'product_code',

        'country': 'country_of_origin',
        'origin': 'country_of_origin',
        'country_of_origin': 'country_of_origin',
    }

    # CSV injection patterns
    CSV_INJECTION_PATTERNS = [
        r'^[=+\-@]',  # Formula injection
        r'cmd\|',     # Command execution
        r'powershell', # PowerShell
        r'<script',   # XSS attempts
        r'javascript:', # JavaScript execution
        r'data:',     # Data URLs
    ]

    def __init__(self):
        """Initialize file processor"""
        self.temp_dir = Path("/tmp/claude")
        self.temp_dir.mkdir(parents=True, exist_ok=True)

    def validate_file_security(self, file: UploadFile) -> bool:
        """
        Comprehensive file security validation

        Args:
            file: FastAPI UploadFile object

        Returns:
            bool: True if file passes security checks

        Raises:
            HTTPException: If file fails security validation
        """
        try:
            # File size check
            file.file.seek(0, 2)  # Seek to end
            file_size = file.file.tell()
            file.file.seek(0)  # Reset to beginning

            if file_size > self.MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File too large. Maximum size: {self.MAX_FILE_SIZE // 1024 // 1024}MB"
                )

            # File extension check
            file_ext = Path(file.filename or '').suffix.lower()
            if file_ext not in self.ALLOWED_EXTENSIONS:
                raise HTTPException(
                    status_code=415,
                    detail=f"Unsupported file type. Allowed: {', '.join(self.ALLOWED_EXTENSIONS)}"
                )

            # Content type validation
            allowed_content_types = {
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                'application/vnd.ms-excel',  # .xls
                'text/csv',  # .csv
                'application/csv',
                'text/plain',
            }

            if file.content_type and file.content_type not in allowed_content_types:
                print(f"Warning: Unexpected content type: {file.content_type}")

            # Read first few bytes for magic number validation
            file_content = file.file.read(100)
            file.file.seek(0)

            if file_ext in ['.xlsx', '.xls']:
                # Excel files should start with ZIP signature (xlsx) or OLE signature (xls)
                if not (file_content.startswith(b'PK') or file_content.startswith(b'\xd0\xcf\x11\xe0')):
                    raise HTTPException(
                        status_code=400,
                        detail="Invalid Excel file format"
                    )

            return True

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"File validation error: {str(e)}"
            )

    def sanitize_cell_content(self, value: Any) -> str:
        """
        Sanitize cell content to prevent CSV injection and XSS

        Args:
            value: Cell value to sanitize

        Returns:
            str: Sanitized value
        """
        if pd.isna(value) or value == '':
            return ''

        value_str = str(value).strip()

        # Check for CSV injection patterns
        for pattern in self.CSV_INJECTION_PATTERNS:
            if re.search(pattern, value_str, re.IGNORECASE):
                # Neutralize by prefixing with single quote
                return f"'{value_str}"

        # Basic XSS prevention
        value_str = value_str.replace('<', '&lt;').replace('>', '&gt;')

        return value_str

    def normalize_column_names(self, columns: List[str]) -> Dict[str, str]:
        """
        Normalize column names using Russian/English mappings

        Args:
            columns: List of column names from file

        Returns:
            dict: Mapping of original column names to standard field names
        """
        column_mapping = {}

        for col in columns:
            col_clean = col.lower().strip()

            # Remove extra spaces and punctuation
            col_clean = re.sub(r'[^\w\s]', ' ', col_clean)
            col_clean = re.sub(r'\s+', ' ', col_clean).strip()

            # Check direct mapping
            if col_clean in self.RUSSIAN_COLUMN_MAP:
                column_mapping[col] = self.RUSSIAN_COLUMN_MAP[col_clean]
                continue

            # Check partial matching for compound names
            for russian_term, english_term in self.RUSSIAN_COLUMN_MAP.items():
                if russian_term in col_clean or col_clean in russian_term:
                    column_mapping[col] = english_term
                    break

        return column_mapping

    def parse_numeric_value(self, value: Any, field_name: str) -> Decimal:
        """
        Parse numeric value with Russian number formatting support

        Args:
            value: Value to parse
            field_name: Field name for error messages

        Returns:
            Decimal: Parsed decimal value
        """
        if pd.isna(value) or value == '':
            return Decimal('0')

        try:
            # Convert to string and clean
            value_str = str(value).strip()

            # Remove currency symbols
            value_str = re.sub(r'[₽$€¥£]', '', value_str)

            # Handle Russian number formatting (space as thousands separator)
            value_str = value_str.replace(' ', '')

            # Handle both comma and dot as decimal separators
            if ',' in value_str and '.' in value_str:
                # Both present - assume comma is thousands separator
                value_str = value_str.replace(',', '')
            elif ',' in value_str:
                # Only comma - could be decimal separator in Russian format
                if value_str.count(',') == 1 and value_str.rfind(',') > len(value_str) - 4:
                    # Likely decimal separator
                    value_str = value_str.replace(',', '.')
                else:
                    # Thousands separator
                    value_str = value_str.replace(',', '')

            # Remove any remaining non-numeric characters except dot
            value_str = re.sub(r'[^\d.-]', '', value_str)

            if not value_str:
                return Decimal('0')

            return Decimal(value_str)

        except (InvalidOperation, ValueError) as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid numeric value in field '{field_name}': {value} - {str(e)}"
            )

    def process_excel_file(self, file: UploadFile) -> List[Dict[str, Any]]:
        """
        Process Excel file and extract quote items

        Args:
            file: FastAPI UploadFile object

        Returns:
            List[Dict]: List of processed quote items
        """
        try:
            # Read Excel file
            file_content = file.file.read()
            file.file.seek(0)

            # Try pandas first (more robust)
            try:
                df = pd.read_excel(
                    io.BytesIO(file_content),
                    engine='openpyxl',
                    dtype=str,
                    na_filter=False,
                    nrows=self.MAX_ROWS
                )
            except Exception:
                # Fallback to openpyxl
                workbook = load_workbook(io.BytesIO(file_content), read_only=True)
                sheet = workbook.active

                # Convert to DataFrame
                data = []
                headers = []

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_idx == 0:
                        headers = [str(cell) if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
                    else:
                        if row_idx > self.MAX_ROWS:
                            break
                        data.append([str(cell) if cell is not None else '' for cell in row])

                df = pd.DataFrame(data, columns=headers)

            return self._process_dataframe(df)

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Excel processing error: {str(e)}"
            )

    def process_csv_file(self, file: UploadFile) -> List[Dict[str, Any]]:
        """
        Process CSV file and extract quote items

        Args:
            file: FastAPI UploadFile object

        Returns:
            List[Dict]: List of processed quote items
        """
        try:
            # Read CSV content
            file_content = file.file.read()
            file.file.seek(0)

            # Try different encodings
            encodings = ['utf-8', 'cp1251', 'utf-8-sig', 'latin1']
            df = None

            for encoding in encodings:
                try:
                    # Try different separators
                    separators = [',', ';', '\t']

                    for sep in separators:
                        try:
                            df = pd.read_csv(
                                io.BytesIO(file_content),
                                encoding=encoding,
                                sep=sep,
                                dtype=str,
                                na_filter=False,
                                nrows=self.MAX_ROWS
                            )

                            # Check if we got meaningful data (more than 1 column)
                            if len(df.columns) > 1:
                                break

                        except Exception:
                            continue

                    if df is not None and len(df.columns) > 1:
                        break

                except Exception:
                    continue

            if df is None or df.empty:
                raise HTTPException(
                    status_code=400,
                    detail="Could not parse CSV file. Please check file format and encoding."
                )

            return self._process_dataframe(df)

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"CSV processing error: {str(e)}"
            )

    def _process_dataframe(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Process pandas DataFrame into quote items

        Args:
            df: Pandas DataFrame with file data

        Returns:
            List[Dict]: Processed quote items
        """
        try:
            # Remove empty rows
            df = df.dropna(how='all')

            if df.empty:
                raise HTTPException(
                    status_code=400,
                    detail="File contains no data"
                )

            # Normalize column names
            original_columns = df.columns.tolist()
            column_mapping = self.normalize_column_names(original_columns)

            if not column_mapping:
                raise HTTPException(
                    status_code=400,
                    detail=f"Could not identify required columns. Found columns: {', '.join(original_columns)}"
                )

            # Apply column mapping
            df_mapped = df.rename(columns=column_mapping)

            # Ensure required fields exist
            required_fields = ['description']
            missing_fields = [field for field in required_fields if field not in df_mapped.columns]

            if missing_fields:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns: {', '.join(missing_fields)}. "
                          f"Please ensure your file has columns for: {', '.join(required_fields)}"
                )

            # Process each row
            items = []
            errors = []

            for idx, row in df_mapped.iterrows():
                try:
                    item = self._process_row(row, idx + 1)
                    if item:
                        items.append(item)
                except Exception as e:
                    errors.append(f"Row {idx + 1}: {str(e)}")

            if not items and errors:
                raise HTTPException(
                    status_code=400,
                    detail=f"No valid items found. Errors: {'; '.join(errors[:5])}"
                )

            if errors:
                print(f"Processing warnings: {'; '.join(errors[:10])}")

            return items

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Data processing error: {str(e)}"
            )

    def _process_row(self, row: pd.Series, row_number: int) -> Optional[Dict[str, Any]]:
        """
        Process a single row into quote item

        Args:
            row: Pandas Series representing a row
            row_number: Row number for error reporting

        Returns:
            Dict or None: Processed quote item or None if row should be skipped
        """
        try:
            # Sanitize all string fields
            description = self.sanitize_cell_content(row.get('description', ''))

            if not description:
                return None  # Skip rows without description

            # Parse numeric fields
            quantity = self.parse_numeric_value(row.get('quantity', '1'), 'quantity')
            unit_price = self.parse_numeric_value(row.get('unit_price', '0'), 'unit_price')

            # Calculate line total if not provided
            line_total = row.get('line_total')
            if line_total and pd.notna(line_total):
                line_total = self.parse_numeric_value(line_total, 'line_total')
            else:
                line_total = quantity * unit_price

            # Process optional fields
            item = {
                'description': description,
                'quantity': float(quantity),
                'unit_price': float(unit_price),
                'line_total': float(line_total),
                'unit': self.sanitize_cell_content(row.get('unit', 'шт.')),
                'category': self.sanitize_cell_content(row.get('category', '')),
                'brand': self.sanitize_cell_content(row.get('brand', '')),
                'notes': self.sanitize_cell_content(row.get('notes', '')),
                'product_code': self.sanitize_cell_content(row.get('product_code', '')),
                'country_of_origin': self.sanitize_cell_content(row.get('country_of_origin', ''))
            }

            return item

        except Exception as e:
            raise ValueError(f"Error processing row data: {str(e)}")

    async def process_file(self, file: UploadFile) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        """
        Main file processing method

        Args:
            file: FastAPI UploadFile object

        Returns:
            Tuple: (List of quote items, Processing summary)
        """
        try:
            # Security validation
            self.validate_file_security(file)

            # Determine file type and process
            file_ext = Path(file.filename or '').suffix.lower()

            if file_ext in ['.xlsx', '.xls']:
                items = self.process_excel_file(file)
            elif file_ext == '.csv':
                items = self.process_csv_file(file)
            else:
                raise HTTPException(
                    status_code=415,
                    detail="Unsupported file type"
                )

            # Generate processing summary
            summary = {
                'total_rows_processed': len(items),
                'file_name': file.filename,
                'file_size': file.size or 0,
                'file_type': file_ext,
                'processing_timestamp': pd.Timestamp.now().isoformat(),
                'items_with_prices': len([item for item in items if item['unit_price'] > 0]),
                'total_value': sum(item['line_total'] for item in items),
                'unique_categories': len(set(item['category'] for item in items if item['category'])),
                'unique_brands': len(set(item['brand'] for item in items if item['brand']))
            }

            return items, summary

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"File processing failed: {str(e)}"
            )


# Global file processor instance
file_processor = SecureFileProcessor()