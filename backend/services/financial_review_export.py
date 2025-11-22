"""
Financial Review Excel Export Service

Generates Excel files for financial manager review with:
- Quote data and calculation results
- Auto-validation highlighting (red cells for issues)
- Comments explaining problems
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from decimal import Decimal
from typing import Dict, List, Any, Optional


# Color fills for validation
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")


def create_financial_review_excel(quote_data: Dict[str, Any]) -> Workbook:
    """
    Create Excel workbook for financial manager review

    Args:
        quote_data: Quote data with calculation results including:
            - quote_number: str
            - customer_name: str
            - seller_company: str (D5)
            - offer_sale_type: str (D6)
            - offer_incoterms: str (D7)
            - currency_of_quote: str (D8)
            - delivery_time: int (D9, days)
            - advance_to_supplier: Decimal (D11, %)
            - advance_from_client: Decimal (J5, %)
            - time_to_advance: int (K5, days)
            - logistics fields: W2-W10
            - totals: V13, AB13, AC13, AK13, AL13, AM13
            - dm_fee_type: str (AG3)
            - dm_fee_value: Decimal (AG7)
            - products: List[Dict] - product-level data
            - vat_removed: bool - Was VAT removed from purchase price?

    Returns:
        openpyxl Workbook ready for download
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    # ========== HEADER ==========
    ws['A1'] = 'ФИНАНСОВЫЙ АНАЛИЗ КП'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:M1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # ========== QUOTE INFO (Row 3) ==========
    ws['A3'] = 'КП №:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = quote_data.get('quote_number', '')
    ws['B3'].font = Font(bold=True, size=12)

    ws['D3'] = 'Клиент:'
    ws['D3'].font = Font(bold=True)
    ws['E3'] = quote_data.get('customer_name', '')
    ws.merge_cells('E3:H3')

    ws['J3'] = 'Сумма с НДС:'
    ws['J3'].font = Font(bold=True)
    ws['K3'] = float(quote_data.get('total_amount', 0))
    ws['K3'].number_format = '#,##0.00 ₽'
    ws['K3'].font = Font(bold=True)

    # ========== SECTION HEADERS (Row 5) ==========
    ws['A5'] = 'ОСНОВНЫЕ ДАННЫЕ'
    ws['A5'].font = Font(bold=True, size=11)

    ws['F5'] = 'УСЛОВИЯ ОПЛАТЫ'
    ws['F5'].font = Font(bold=True, size=11)

    ws['I5'] = 'ЛОГИСТИКА'
    ws['I5'].font = Font(bold=True, size=11)

    # ========== BASIC INFO (A6:D11) + PAYMENT TERMS (F6:G8) + LOGISTICS (I6:J14) - HORIZONTAL ==========
    # Basic Info (left side)
    basic_info = [
        ('Продавец', 'seller_company'),
        ('Тип продажи', 'offer_sale_type'),
        ('Инкотермс', 'offer_incoterms'),
        ('Валюта', 'currency_of_quote'),
        ('Срок поставки (дни)', 'delivery_time'),
        ('Аванс поставщику (%)', 'advance_to_supplier'),
    ]

    row = 6
    for label, key in basic_info:
        ws[f'A{row}'] = label
        value = quote_data.get(key, '')
        ws[f'B{row}'] = float(value) if isinstance(value, Decimal) else value
        if 'аванс' in label.lower():
            ws[f'B{row}'].number_format = '0.0"%"'
        row += 1

    # Payment Terms (middle)
    payment_terms = [
        ('Аванс клиента (%)', 'advance_from_client'),
        ('Дни до аванса', 'time_to_advance'),
    ]

    row = 6
    for label, key in payment_terms:
        ws[f'F{row}'] = label
        value = quote_data.get(key, 0)
        ws[f'G{row}'] = float(value)
        if '%' in label:
            ws[f'G{row}'].number_format = '0.0"%"'
        row += 1

    # Logistics (right side)
    logistics_fields = [
        ('Поставщик → Хаб', 'logistics_supplier_hub'),
        ('Хаб → Таможня', 'logistics_hub_customs'),
        ('Таможня → Клиент', 'logistics_customs_client'),
        ('Брокераж хаб', 'brokerage_hub'),
        ('Брокераж таможня', 'brokerage_customs'),
        ('Хранение', 'warehousing_at_customs'),
        ('Оформление', 'customs_documentation'),
        ('Доп. брокераж', 'brokerage_extra'),
        ('Страховка', 'insurance'),
    ]

    row = 6
    for label, key in logistics_fields:
        ws[f'I{row}'] = label
        value = quote_data.get(key, 0)
        ws[f'J{row}'] = float(value) if value else 0
        ws[f'J{row}'].number_format = '#,##0.00 ₽'
        row += 1

    # ========== QUOTE TOTALS (Row 17 - Horizontal like test_raschet Row 13) ==========
    row += 2
    ws[f'A{row}'] = 'ИТОГО ПО КП'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:M{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    row += 1
    # Headers
    totals_headers = ['Логистика', 'Себест-ть', 'Наценка %', 'Цена без НДС', 'Цена с НДС', 'Маржа', 'DM Гонорар']
    col_letters = ['A', 'C', 'E', 'G', 'I', 'K', 'M']

    for idx, header in enumerate(totals_headers):
        cell = ws[f'{col_letters[idx]}{row}']
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')

    row += 1
    # Values
    advance = Decimal(str(quote_data.get('advance_from_client', 0)))
    delivery = quote_data.get('delivery_time', 0)
    total_cogs = Decimal(str(quote_data.get('total_cogs', 0)))
    total_margin = Decimal(str(quote_data.get('total_margin', 0)))
    dm_fee = Decimal(str(quote_data.get('dm_fee_value', 0)))

    # Calculate ACHIEVED markup (actual profit margin considering product-level overrides)
    # Formula: achieved_markup % = (total_margin / total_cogs) × 100
    if total_cogs > 0:
        achieved_markup = (total_margin / total_cogs) * 100
    else:
        achieved_markup = Decimal("0")

    totals_values = [
        ('A', quote_data.get('total_logistics', 0), '#,##0.00 ₽', None),
        ('C', total_cogs, '#,##0.00 ₽', None),
        ('E', achieved_markup, '0.00"%"', validate_markup(achieved_markup, advance, delivery, level='quote')),
        ('G', quote_data.get('total_revenue_no_vat', 0), '#,##0.00 ₽', None),
        ('I', quote_data.get('total_revenue_with_vat', 0), '#,##0.00 ₽', None),
        ('K', total_margin, '#,##0.00 ₽', None),
        ('M', dm_fee, '#,##0.00 ₽', validate_dm_fee(dm_fee, total_margin)),
    ]

    for col, value, fmt, validation in totals_values:
        cell = ws[f'{col}{row}']
        cell.value = float(value) if value else 0
        cell.number_format = fmt
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='right')

        # Apply validation if exists
        if validation:
            is_valid, error_msg = validation
            apply_validation_to_cell(cell, is_valid, error_msg)

    # ========== VAT & DM FEE INFO (Row after totals) ==========
    row += 2
    ws[f'A{row}'] = 'НДС очищен:'
    vat_removed = quote_data.get('vat_removed', False)
    ws[f'B{row}'] = 'ДА' if vat_removed else 'НЕТ'
    ws[f'B{row}'].font = Font(bold=True)

    # Highlight if VAT not removed
    if not vat_removed:
        ws[f'B{row}'].fill = YELLOW_FILL
        ws[f'B{row}'].comment = Comment(
            "⚠️ VAT не был очищен от закупочной цены. Проверьте.",
            "System"
        )

    ws[f'D{row}'] = 'DM Гонорар тип:'
    ws[f'E{row}'] = quote_data.get('dm_fee_type', '')

    # ========== PRODUCTS TABLE ==========
    row += 2
    ws[f'A{row}'] = 'ТОВАРЫ'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:M{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    row += 1
    # Headers - Expanded to show cost buildup for financial manager transparency
    headers = [
        '№',             # A
        'Наименование',  # B
        'Кол-во',        # D
        'Цена закупки',  # E - purchase_price_supplier (from supplier)
        'После скидки',  # F - purchase_price_after_discount
        'Логистика',     # G - logistics (distributed)
        'Таможня+Акциз', # H - customs_fee + excise_tax
        'Финансирование',# I - financing costs
        'С/с за ед.',    # J - cogs_per_unit
        'С/с всего',     # K - cogs total
        'Наценка %',     # L - markup %
        'Цена б/НДС',    # M - price_no_vat
        'Цена с НДС',    # N - price_with_vat
        'Маржа'          # O - profit
    ]
    col_letters_products = ['A', 'B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']

    for idx, header in enumerate(headers):
        cell = ws[f'{col_letters_products[idx]}{row}']
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row += 1
    # Products data
    products = quote_data.get('products', [])
    advance = Decimal(str(quote_data.get('advance_from_client', 0)))
    delivery = quote_data.get('delivery_time', 0)

    for idx, product in enumerate(products, start=1):
        # Column A: №
        ws[f'A{row}'] = idx
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Column B-C: Product name (merged)
        ws[f'B{row}'] = product.get('name', f'Товар {idx}')
        ws.merge_cells(f'B{row}:C{row}')

        # Column D: Quantity
        ws[f'D{row}'] = product.get('quantity', 0)
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        # Column E: Purchase price (original from supplier)
        ws[f'E{row}'] = float(product.get('purchase_price_supplier', 0))
        ws[f'E{row}'].number_format = '#,##0.00 ₽'
        ws[f'E{row}'].alignment = Alignment(horizontal='right')

        # Column F: After discount
        ws[f'F{row}'] = float(product.get('purchase_price_after_discount', 0))
        ws[f'F{row}'].number_format = '#,##0.00 ₽'
        ws[f'F{row}'].alignment = Alignment(horizontal='right')

        # Column G: Logistics (distributed to this product)
        ws[f'G{row}'] = float(product.get('logistics', 0))
        ws[f'G{row}'].number_format = '#,##0.00 ₽'
        ws[f'G{row}'].alignment = Alignment(horizontal='right')

        # Column H: Customs + Excise
        customs_excise_total = (
            Decimal(str(product.get('customs_fee', 0))) +
            Decimal(str(product.get('excise_tax', 0)))
        )
        ws[f'H{row}'] = float(customs_excise_total)
        ws[f'H{row}'].number_format = '#,##0.00 ₽'
        ws[f'H{row}'].alignment = Alignment(horizontal='right')

        # Column I: Financing costs
        ws[f'I{row}'] = float(product.get('financing', 0))
        ws[f'I{row}'].number_format = '#,##0.00 ₽'
        ws[f'I{row}'].alignment = Alignment(horizontal='right')

        # Column J: COGS per unit
        ws[f'J{row}'] = float(product.get('cogs_per_unit', 0))
        ws[f'J{row}'].number_format = '#,##0.00 ₽'
        ws[f'J{row}'].alignment = Alignment(horizontal='right')

        # Column K: COGS total
        ws[f'K{row}'] = float(product.get('cogs', 0))
        ws[f'K{row}'].number_format = '#,##0.00 ₽'
        ws[f'K{row}'].alignment = Alignment(horizontal='right')
        ws[f'K{row}'].font = Font(bold=True)  # Highlight COGS

        # Column L: Markup % with validation
        product_markup = Decimal(str(product.get('markup', 0)))
        markup_cell = ws[f'L{row}']
        markup_cell.value = float(product_markup)
        markup_cell.number_format = '0.00"%"'
        markup_cell.alignment = Alignment(horizontal='right')
        markup_cell.font = Font(bold=True)  # Highlight markup

        # Validate product-level markup
        is_valid, error_msg = validate_markup(product_markup, advance, delivery, level='product')
        apply_validation_to_cell(markup_cell, is_valid, error_msg)

        # Column M: Price no VAT
        ws[f'M{row}'] = float(product.get('price_no_vat', 0))
        ws[f'M{row}'].number_format = '#,##0.00 ₽'
        ws[f'M{row}'].alignment = Alignment(horizontal='right')

        # Column N: Price with VAT
        ws[f'N{row}'] = float(product.get('price_with_vat', 0))
        ws[f'N{row}'].number_format = '#,##0.00 ₽'
        ws[f'N{row}'].alignment = Alignment(horizontal='right')
        ws[f'N{row}'].font = Font(bold=True)  # Highlight final price

        # Column O: Profit
        profit_val = float(product.get('profit', 0))
        ws[f'O{row}'] = profit_val
        ws[f'O{row}'].number_format = '#,##0.00 ₽'
        ws[f'O{row}'].alignment = Alignment(horizontal='right')
        # Color code profit: green if positive, red if negative
        if profit_val > 0:
            ws[f'O{row}'].font = Font(bold=True, color="006600")  # Green
        elif profit_val < 0:
            ws[f'O{row}'].font = Font(bold=True, color="CC0000")  # Red
            ws[f'O{row}'].fill = YELLOW_FILL

        row += 1

    # Set column widths for the expanded product table
    column_widths = {
        'A': 6,   # № (row number)
        'B': 25,  # Наименование (product name)
        'C': 5,   # Merged with B
        'D': 8,   # Кол-во (quantity)
        'E': 12,  # Цена закупки (purchase price)
        'F': 12,  # После скидки (after discount)
        'G': 12,  # Логистика (logistics)
        'H': 13,  # Таможня+Акциз (customs+excise)
        'I': 13,  # Финансирование (financing)
        'J': 11,  # С/с за ед. (COGS per unit)
        'K': 12,  # С/с всего (COGS total)
        'L': 10,  # Наценка % (markup)
        'M': 13,  # Цена б/НДС (price no VAT)
        'N': 13,  # Цена с НДС (price with VAT)
        'O': 12,  # Маржа (profit)
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    return wb


def validate_vat_removal(product: Dict[str, Any]) -> tuple[bool, str]:
    """
    Check if VAT was properly removed from purchase price

    Args:
        product: Product data with purchase_price_with_vat, purchase_price_no_vat, vat_rate

    Returns:
        (is_valid, error_message)
    """
    price_with_vat = Decimal(str(product.get('purchase_price_with_vat', 0)))
    price_no_vat = Decimal(str(product.get('purchase_price_no_vat', 0)))
    vat_rate = Decimal(str(product.get('vat_seller_country', 0)))

    if vat_rate == 0:
        return (True, '')  # No VAT expected

    # Calculate expected price without VAT
    expected_no_vat = price_with_vat / (Decimal('1') + vat_rate / Decimal('100'))

    # Allow 1% tolerance for rounding
    diff = abs(price_no_vat - expected_no_vat)
    tolerance = expected_no_vat * Decimal('0.01')

    if diff > tolerance:
        return (False, f"⚠️ VAT not removed properly. Expected: {expected_no_vat:.2f}, Got: {price_no_vat:.2f}")

    return (True, '')


def calculate_required_markup(advance_percent: Decimal, delivery_days: int) -> Decimal:
    """
    Calculate required markup based on advance and delivery time

    Rules:
    - If advance > 50%: require 5% minimum
    - If advance ≤ 50%: add 5% penalty per 30 days of delivery

    Examples:
    - advance=50%, delivery=60 days → required=10% (5% base + 10% penalty)
    - advance=30%, delivery=90 days → required=20% (5% base + 15% penalty)
    - advance=70%, delivery=60 days → required=5% (no penalty)

    Args:
        advance_percent: Client advance payment percentage
        delivery_days: Delivery time in days

    Returns:
        Required markup percentage
    """
    base_markup = Decimal('5.0')

    if advance_percent <= 50:
        penalty_periods = Decimal(str(delivery_days)) / Decimal('30')
        penalty = penalty_periods * Decimal('5.0')
        return base_markup + penalty

    return base_markup


def validate_markup(
    markup: Decimal,
    advance_percent: Decimal,
    delivery_days: int,
    level: str = 'quote'
) -> tuple[bool, str]:
    """
    Validate markup meets threshold requirements

    Args:
        markup: Actual markup percentage
        advance_percent: Client advance percentage
        delivery_days: Delivery time in days
        level: 'quote' or 'product' (for error message)

    Returns:
        (is_valid, error_message)
    """
    required_markup = calculate_required_markup(advance_percent, delivery_days)

    if markup < required_markup:
        return (
            False,
            f"⚠️ {level.capitalize()}-level markup {markup:.2f}% below required threshold {required_markup:.2f}%.\n"
            f"Advance: {advance_percent}%, Delivery: {delivery_days} days."
        )

    return (True, '')


def validate_dm_fee(dm_fee_value: Decimal, margin_value: Decimal) -> tuple[bool, str]:
    """
    Check if DM fee exceeds deal margin

    Args:
        dm_fee_value: DM fee amount
        margin_value: Deal margin amount

    Returns:
        (is_valid, error_message)
    """
    if dm_fee_value > margin_value:
        return (
            False,
            f"⚠️ DM fee ({dm_fee_value:,.2f} ₽) exceeds deal margin ({margin_value:,.2f} ₽)"
        )

    return (True, '')


def apply_validation_to_cell(cell, is_valid: bool, error_message: str):
    """
    Apply color and comment to cell based on validation result

    Args:
        cell: openpyxl cell object
        is_valid: True if validation passed
        error_message: Error message to show in comment (if failed)
    """
    if not is_valid:
        cell.fill = RED_FILL
        if error_message:
            cell.comment = Comment(error_message, "System")
    else:
        # Optionally apply green for passed validations
        # cell.fill = GREEN_FILL
        pass
