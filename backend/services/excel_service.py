"""
Excel Export Service

Generates Excel exports in 2 formats:
1. Validation Export - Input/Output comparison for checking against old Excel
2. Grid Export - Professional 2-sheet export (КП поставка + КП open book)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from decimal import Decimal
from typing import Dict, Any, List
import sys
import os

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from services.export_data_mapper import ExportData


class QuoteExcelService:
    """Excel export service for quotes"""

    @staticmethod
    def format_russian_number(value: Any) -> str:
        """
        Format number with Russian style: 1 234,56

        Args:
            value: Number to format (Decimal, float, int, or None)

        Returns:
            Formatted string with space as thousand separator and comma as decimal
        """
        if value is None:
            return "0,00"

        # Convert to float
        try:
            num_value = float(value)
        except (ValueError, TypeError):
            return "0,00"

        # Format with 2 decimals
        str_val = f"{num_value:,.2f}"

        # Replace comma (thousand separator) with space, dot (decimal) with comma
        # Python's format uses comma for thousands and dot for decimal
        return str_val.replace(',', ' ').replace('.', ',')

    @staticmethod
    def generate_validation_export(export_data: ExportData) -> bytes:
        """
        Generate Input/Output validation Excel.

        Structure:
        - Section 1: All 42 quote-level input variables (shown once)
        - Section 2+: For each product:
          - Left side (A-C): Product inputs + overrides
          - Right side (E-G): All calculated outputs

        Column Layout:
        A: Input Cell | B: Input Name | C: Input Value | D: (empty) | E: Output Cell | F: Output Name | G: Output Value

        Args:
            export_data: Export data with quote, items, calculations

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Проверка расчетов"

        # ========== HEADERS ==========
        ws['A1'] = 'ВХОДНЫЕ ДАННЫЕ (INPUTS)'
        ws['E1'] = 'ВЫХОДНЫЕ ДАННЫЕ (OUTPUTS)'
        ws['A2'] = 'Excel Cell'
        ws['B2'] = 'Название'
        ws['C2'] = 'Значение'
        ws['E2'] = 'Excel Cell'
        ws['F2'] = 'Название'
        ws['G2'] = 'Значение'

        # Style headers
        header_fill = PatternFill(start_color="FF2C5AA0", end_color="FF2C5AA0", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFFFF")
        section_font = Font(bold=True, size=13, color="FF2C5AA0")

        for cell in ['A1', 'E1']:
            ws[cell].font = section_font
        for cell in ['A2', 'B2', 'C2', 'E2', 'F2', 'G2']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')

        row = 4

        # ========== SECTION 1: QUOTE-LEVEL INPUTS (ALL 42 VARIABLES) ==========
        ws[f'A{row}'] = '═══ УРОВЕНЬ КОТИРОВКИ (QUOTE-LEVEL) ═══'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:C{row}')
        row += 2

        # All quote-level variables with Excel cell references
        quote_vars = [
            # Company & Type
            ('D5', 'Компания-продавец', 'seller_company'),
            ('D6', 'Вид КП', 'offer_sale_type'),
            ('D7', 'Базис поставки Incoterms', 'offer_incoterms'),
            ('D8', 'Валюта КП', 'currency_of_quote'),
            ('D9', 'Срок поставки (дней) [default]', 'delivery_time'),
            ('', '', ''),  # Blank row
            # Payment Terms
            ('J5', 'Аванс от клиента (%)', 'advance_from_client'),
            ('K5', 'Дней от подписания до аванса', 'time_to_advance'),
            ('D11', 'Аванс поставщику (%) [default]', 'advance_to_supplier'),
            ('J6', 'Аванс при заборе груза (%)', 'advance_on_loading'),
            ('K6', 'Дней от забора до аванса', 'time_to_advance_loading'),
            ('J7', 'Аванс при отправке в РФ (%)', 'advance_on_going_to_country_destination'),
            ('K7', 'Дней от отправки до аванса', 'time_to_advance_going_to_country_destination'),
            ('J8', 'Аванс при таможне (%)', 'advance_on_customs_clearance'),
            ('K8', 'Дней от таможни до аванса', 'time_to_advance_on_customs_clearance'),
            ('K9', 'Дней от получения до оплаты', 'time_to_advance_on_receiving'),
            ('', '', ''),  # Blank row
            # Logistics & Customs Costs
            ('W2', 'Логистика Поставщик-Турция', 'logistics_supplier_hub'),
            ('W3', 'Логистика Турция-РФ', 'logistics_hub_customs'),
            ('W4', 'Логистика Таможня-Клиент', 'logistics_customs_client'),
            ('W5', 'Брокерские Турция', 'brokerage_hub'),
            ('W6', 'Брокерские РФ', 'brokerage_customs'),
            ('W7', 'Расходы на СВХ', 'warehousing_at_customs'),
            ('W8', 'Разрешительные документы', 'customs_documentation'),
            ('W9', 'Прочее', 'brokerage_extra'),
            ('', '', ''),  # Blank row
            # Financial & Fees
            ('AG3', 'Вознаграждение ЛПР (тип)', 'dm_fee_type'),
            ('AG7', 'Вознаграждение ЛПР (значение)', 'dm_fee_value'),
            ('AH11', 'Резерв на курсовой разнице (%)', 'rate_forex_risk'),
            ('(admin)', 'Комиссия ФинАгента (%)', 'rate_fin_comm'),
            ('(admin)', 'Дневная стоимость денег (%)', 'rate_loan_interest_daily'),
            ('', 'Утилизационный сбор', 'util_fee'),
            ('', '', ''),  # Blank row
            # Defaults (can be overridden per product)
            ('J16', 'Валюта цены закупки [default]', 'currency_of_base_price'),
            ('L16', 'Страна закупки [default]', 'supplier_country'),
            ('O16', 'Скидка поставщика (%) [default]', 'supplier_discount'),
            ('Q16', 'Курс к валюте КП [default]', 'exchange_rate_base_price_to_quote'),
            ('W16', 'Код ТН ВЭД [default]', 'customs_code'),
            ('X16', 'Пошлина (%) [default]', 'import_tariff'),
            ('Z16', 'Акциз [default]', 'excise_tax'),
            ('AC16', 'Наценка (%) [default]', 'markup'),
        ]

        for cell_ref, name, var_key in quote_vars:
            if not cell_ref and not name:  # Blank row
                row += 1
                continue

            value = export_data.variables.get(var_key, '') if var_key else ''
            ws[f'A{row}'] = cell_ref
            ws[f'B{row}'] = name
            ws[f'C{row}'] = str(value) if value != '' else ''
            row += 1

        row += 2

        # ========== TOTALS SECTION (Row 13 in original Excel) - OUTPUT COLUMN ==========
        ws[f'E{row}'] = '═══ ИТОГИ (TOTALS - Quote Level) ═══'
        ws[f'E{row}'].font = Font(bold=True, size=12, color="FF2C5AA0")
        ws.merge_cells(f'E{row}:G{row}')
        row += 2

        # Calculate comprehensive totals from all products
        totals = {
            # Input totals
            'quantity': 0,
            'weight': Decimal('0'),
            # Purchase totals
            'purchase_no_vat': Decimal('0'),
            'purchase_after_discount': Decimal('0'),
            'purchase_quote_currency': Decimal('0'),
            # Logistics totals
            'logistics_dist_1': Decimal('0'),
            'logistics_dist_2': Decimal('0'),
            'logistics_total': Decimal('0'),
            # Customs & fees totals
            'customs_fee': Decimal('0'),
            'excise_tax': Decimal('0'),
            # Cost totals
            'cogs': Decimal('0'),
            'sale_price_excl_financial': Decimal('0'),
            'profit': Decimal('0'),
            # Financial totals
            'dm_fee': Decimal('0'),
            'forex_risk': Decimal('0'),
            'fin_comm': Decimal('0'),
            # Sales totals
            'sales_price_per_unit': Decimal('0'),
            'sales_total_no_vat': Decimal('0'),
            'sales_total_with_vat': Decimal('0'),
            'sales_price_with_vat_per_unit': Decimal('0'),
            # VAT totals
            'vat_from_sales': Decimal('0'),
            'vat_at_import': Decimal('0'),
            'vat_payable': Decimal('0'),
            # Other
            'transit_commission': Decimal('0'),
        }

        for item in export_data.items:
            calc_results = item.get('calculation_results', {})
            # Input totals
            totals['quantity'] += item.get('quantity', 0)
            totals['weight'] += Decimal(str(item.get('weight_in_kg', 0)))
            # Purchase totals
            totals['purchase_no_vat'] += Decimal(str(calc_results.get('purchase_price_no_vat', 0)))
            totals['purchase_after_discount'] += Decimal(str(calc_results.get('purchase_price_after_discount', 0)))
            totals['purchase_quote_currency'] += Decimal(str(calc_results.get('purchase_price_total_quote_currency', 0)))
            # Logistics totals
            totals['logistics_dist_1'] += Decimal(str(calc_results.get('logistics_distributed_1', 0)))
            totals['logistics_dist_2'] += Decimal(str(calc_results.get('logistics_distributed_2', 0)))
            totals['logistics_total'] += Decimal(str(calc_results.get('logistics_total', 0)))
            # Customs & fees
            totals['customs_fee'] += Decimal(str(calc_results.get('customs_fee', 0)))
            totals['excise_tax'] += Decimal(str(calc_results.get('excise_tax_amount', 0)))
            # Cost totals
            totals['cogs'] += Decimal(str(calc_results.get('cogs', 0)))
            totals['sale_price_excl_financial'] += Decimal(str(calc_results.get('sale_price_excl_financial', 0)))
            totals['profit'] += Decimal(str(calc_results.get('profit', 0)))
            # Financial
            totals['dm_fee'] += Decimal(str(calc_results.get('dm_fee_amount', 0)))
            totals['forex_risk'] += Decimal(str(calc_results.get('forex_risk_amount', 0)))
            totals['fin_comm'] += Decimal(str(calc_results.get('fin_comm_amount', 0)))
            # Sales totals
            totals['sales_price_per_unit'] += Decimal(str(calc_results.get('sales_price_per_unit', 0)))
            totals['sales_total_no_vat'] += Decimal(str(calc_results.get('sales_price_total_no_vat', 0)))
            totals['sales_total_with_vat'] += Decimal(str(calc_results.get('sales_price_total_with_vat', 0)))
            totals['sales_price_with_vat_per_unit'] += Decimal(str(calc_results.get('sales_price_per_unit_with_vat', 0)))
            # VAT
            totals['vat_from_sales'] += Decimal(str(calc_results.get('vat_amount', 0)))
            totals['vat_at_import'] += Decimal(str(calc_results.get('vat_at_import', 0)))
            totals['vat_payable'] += Decimal(str(calc_results.get('vat_payable', 0)))
            # Other
            totals['transit_commission'] += Decimal(str(calc_results.get('transit_commission', 0)))

        # Write comprehensive totals to OUTPUT column (E-G) with cell references (row 13 in original Excel)
        total_cells = [
            # Input totals
            ('E13', 'Итого кол-во', totals['quantity']),
            ('G13', 'Итого вес (кг)', totals['weight']),
            ('', '', ''),  # Blank row
            # Purchase totals
            ('N13', 'Итого цена без НДС', totals['purchase_no_vat']),
            ('P13', 'Итого после скидки', totals['purchase_after_discount']),
            ('R13', 'Итого в валюте КП', totals['purchase_quote_currency']),
            ('S13', 'Итого покупка', totals['purchase_quote_currency']),
            ('', '', ''),  # Blank row
            # Logistics
            ('T13', 'Итого логистика (распр. 1)', totals['logistics_dist_1']),
            ('U13', 'Итого логистика (распр. 2)', totals['logistics_dist_2']),
            ('V13', 'Итого логистика', totals['logistics_total']),
            ('', '', ''),  # Blank row
            # Customs & fees
            ('Y13', 'Итого таможенный сбор', totals['customs_fee']),
            ('Z13', 'Итого акциз', totals['excise_tax']),
            ('', '', ''),  # Blank row
            # Cost structure
            ('AB13', 'Итого COGS', totals['cogs']),
            ('AD13', 'Итого цена без фин.', totals['sale_price_excl_financial']),
            ('AF13', 'Итого прибыль', totals['profit']),
            ('', '', ''),  # Blank row
            # Financial
            ('AG13', 'Итого вознаграждение ЛПР', totals['dm_fee']),
            ('AH13', 'Итого резерв курсовой', totals['forex_risk']),
            ('AI13', 'Итого комиссия ФинАгента', totals['fin_comm']),
            ('', '', ''),  # Blank row
            # Sales
            ('AJ13', 'Итого цена продажи за ед.', totals['sales_price_per_unit']),
            ('AK13', 'Итого продажи (без НДС)', totals['sales_total_no_vat']),
            ('AL13', 'Итого с НДС', totals['sales_total_with_vat']),
            ('AM13', 'Итого с НДС за ед.', totals['sales_price_with_vat_per_unit']),
            ('', '', ''),  # Blank row
            # VAT
            ('AN13', 'Итого НДС от продаж', totals['vat_from_sales']),
            ('AO13', 'Итого НДС при импорте', totals['vat_at_import']),
            ('AP13', 'Итого НДС к уплате', totals['vat_payable']),
            ('', '', ''),  # Blank row
            # Other
            ('AQ13', 'Итого комиссия транзит', totals['transit_commission']),
        ]

        for cell_ref, name, value in total_cells:
            if not cell_ref and not name:  # Blank row
                row += 1
                continue
            ws[f'E{row}'] = cell_ref
            ws[f'F{row}'] = name
            if isinstance(value, (int, float, Decimal)):
                ws[f'G{row}'] = QuoteExcelService.format_russian_number(value)
            else:
                ws[f'G{row}'] = str(value)
            ws[f'E{row}'].font = Font(bold=True, color="FF2C5AA0")
            ws[f'F{row}'].font = Font(bold=True)
            ws[f'G{row}'].font = Font(bold=True, color="FF2C5AA0")
            row += 1

        row += 2

        # ========== SECTION 2+: PER-PRODUCT INPUTS & OUTPUTS ==========
        for idx, item in enumerate(export_data.items, 1):
            calc_results = item.get('calculation_results', {})

            # Product header
            product_name = item.get('product_name', '')
            brand = item.get('brand', '')
            product_code = item.get('product_code', '')
            item_title = f'═══ ТОВАР {idx}: {brand} {product_code} - {product_name} ═══'

            ws[f'A{row}'] = item_title
            ws[f'A{row}'].font = Font(bold=True, size=11, color="FF1F4788")
            ws.merge_cells(f'A{row}:G{row}')
            row += 1

            # Calculate Excel row number for this product (Product 1 = row 16, Product 2 = row 17, etc.)
            excel_row = 15 + idx

            # Product input data (Left side: columns A-C)
            product_inputs = [
                (f'B{excel_row}', 'Бренд', brand),
                (f'C{excel_row}', 'Артикул', product_code),
                (f'D{excel_row}', 'Наименование', product_name[:30] + '...' if len(product_name) > 30 else product_name),
                (f'E{excel_row}', 'Кол-во', item.get('quantity', 0)),
                (f'G{excel_row}', 'Вес (кг)', QuoteExcelService.format_russian_number(item.get('weight_in_kg', 0))),
                (f'K{excel_row}', 'Цена закупки (с НДС)', QuoteExcelService.format_russian_number(item.get('base_price_vat', 0))),
                ('', '', ''),
                ('', '--- Переопределения (если есть) ---', ''),
                (f'J{excel_row}', 'Валюта', item.get('currency_of_base_price', '') or '(default)'),
                (f'L{excel_row}', 'Страна', item.get('supplier_country', '') or '(default)'),
                (f'O{excel_row}', 'Скидка (%)', item.get('supplier_discount', '') or '(default)'),
                (f'Q{excel_row}', 'Курс', item.get('exchange_rate_base_price_to_quote', '') or '(default)'),
                (f'W{excel_row}', 'ТН ВЭД', item.get('customs_code', '') or '(default)'),
                (f'X{excel_row}', 'Пошлина (%)', item.get('import_tariff', '') or '(default)'),
                (f'Z{excel_row}', 'Акциз', item.get('excise_tax', '') or '(default)'),
                (f'AC{excel_row}', 'Наценка (%)', item.get('markup', '') or '(default)'),
            ]

            # Calculation outputs (Right side: columns E-G)
            output_start_row = row
            if calc_results:
                outputs = [
                    (f'N{excel_row}', 'Цена без НДС', calc_results.get('purchase_price_no_vat', 0)),
                    (f'P{excel_row}', 'Цена после скидки', calc_results.get('purchase_price_after_discount', 0)),
                    (f'R{excel_row}', 'Цена за ед. в вал. КП', calc_results.get('purchase_price_per_unit_quote_currency', 0)),
                    (f'S{excel_row}', 'Цена покупки итого', calc_results.get('purchase_price_total_quote_currency', 0)),
                    (f'T{excel_row}', 'Логистика (распред. 1)', calc_results.get('logistics_distributed_1', 0)),
                    (f'U{excel_row}', 'Логистика (распред. 2)', calc_results.get('logistics_distributed_2', 0)),
                    (f'V{excel_row}', 'Логистика итого', calc_results.get('logistics_total', 0)),
                    (f'Y{excel_row}', 'Таможенный сбор', calc_results.get('customs_fee', 0)),
                    (f'Z{excel_row}', 'Акциз', calc_results.get('excise_tax_amount', 0)),
                    (f'AB{excel_row}', 'COGS (себестоимость)', calc_results.get('cogs', 0)),
                    (f'AD{excel_row}', 'Цена продажи без фин.', calc_results.get('sale_price_excl_financial', 0)),
                    (f'AF{excel_row}', 'Прибыль (без НДС)', calc_results.get('profit', 0)),
                    (f'AG{excel_row}', 'Вознаграждение ЛПР', calc_results.get('dm_fee_amount', 0)),
                    (f'AH{excel_row}', 'Резерв курсовой', calc_results.get('forex_risk_amount', 0)),
                    (f'AI{excel_row}', 'Комиссия ФинАгента', calc_results.get('fin_comm_amount', 0)),
                    (f'AJ{excel_row}', 'Цена продажи за ед.', calc_results.get('sales_price_per_unit', 0)),
                    (f'AK{excel_row}', 'Сумма продаж (без НДС)', calc_results.get('sales_price_total_no_vat', 0)),
                    (f'AL{excel_row}', 'Сумма с НДС итого', calc_results.get('sales_price_total_with_vat', 0)),
                    (f'AM{excel_row}', 'Цена с НДС за ед.', calc_results.get('sales_price_per_unit_with_vat', 0)),
                    (f'AN{excel_row}', 'НДС от продаж', calc_results.get('vat_amount', 0)),
                    (f'AO{excel_row}', 'НДС при импорте', calc_results.get('vat_at_import', 0)),
                    (f'AP{excel_row}', 'НДС к уплате (разница)', calc_results.get('vat_payable', 0)),
                    (f'AQ{excel_row}', 'Комиссия транзит', calc_results.get('transit_commission', 0)),
                ]

                for output_cell, output_name, output_value in outputs:
                    ws[f'E{output_start_row}'] = output_cell
                    ws[f'F{output_start_row}'] = output_name
                    ws[f'G{output_start_row}'] = QuoteExcelService.format_russian_number(output_value)
                    output_start_row += 1

            # Write product inputs
            for input_cell, input_name, input_value in product_inputs:
                if not input_cell and not input_name:  # Blank row
                    row += 1
                    continue
                ws[f'A{row}'] = input_cell
                ws[f'B{row}'] = input_name
                ws[f'C{row}'] = str(input_value) if input_value not in ['', '(default)'] else input_value
                row += 1

            row = max(row, output_start_row) + 2  # Space between products

        # ========== COLUMN WIDTHS ==========
        ws.column_dimensions['A'].width = 12  # Cell reference
        ws.column_dimensions['B'].width = 35  # Name
        ws.column_dimensions['C'].width = 18  # Value
        ws.column_dimensions['D'].width = 2   # Separator
        ws.column_dimensions['E'].width = 12  # Output cell
        ws.column_dimensions['F'].width = 30  # Output name
        ws.column_dimensions['G'].width = 18  # Output value

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_grid_export(export_data: ExportData) -> bytes:
        """
        Generate professional 2-sheet Excel.
        Sheet 1: "КП поставка" (9 columns)
        Sheet 2: "КП open book" (21 columns)

        Args:
            export_data: Export data with quote, items, calculations

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # ========== Sheet 1: КП поставка (9 columns) ==========
        ws1 = wb.active
        ws1.title = "КП поставка"

        # Headers (9 columns)
        headers_supply = [
            'Бренд', 'Артикул', 'Наименование', 'Кол-во',
            'Цена за ед. (₽)', 'Сумма (₽)', 'НДС (₽)',
            'Цена с НДС за ед. (₽)', 'Сумма с НДС (₽)'
        ]

        # Style headers
        header_fill = PatternFill(start_color="FF2C5AA0", end_color="FF2C5AA0", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, header in enumerate(headers_supply, 1):
            cell = ws1.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        row_idx = 2
        totals = {'quantity': 0, 'col5': Decimal('0'), 'col6': Decimal('0'), 'col7': Decimal('0'), 'col8': Decimal('0'), 'col9': Decimal('0')}

        for item in export_data.items:
            calc_results = item.get('calculation_results', {})

            # Basic info
            ws1.cell(row=row_idx, column=1).value = item.get('brand', '')
            ws1.cell(row=row_idx, column=2).value = item.get('product_code', '')
            ws1.cell(row=row_idx, column=3).value = item.get('product_name', '')
            ws1.cell(row=row_idx, column=4).value = item.get('quantity', 0)

            # Selling prices from calculations
            selling_price_per_unit = Decimal(str(calc_results.get('sales_price_per_unit', 0)))
            selling_price_total = Decimal(str(calc_results.get('sales_price_total_no_vat', 0)))
            vat_from_sales = Decimal(str(calc_results.get('vat_amount', 0)))
            selling_price_with_vat_per_unit = Decimal(str(calc_results.get('sales_price_per_unit_with_vat', 0)))
            selling_price_with_vat_total = Decimal(str(calc_results.get('sales_price_total_with_vat', 0)))

            ws1.cell(row=row_idx, column=5).value = float(selling_price_per_unit)
            ws1.cell(row=row_idx, column=6).value = float(selling_price_total)
            ws1.cell(row=row_idx, column=7).value = float(vat_from_sales)
            ws1.cell(row=row_idx, column=8).value = float(selling_price_with_vat_per_unit)
            ws1.cell(row=row_idx, column=9).value = float(selling_price_with_vat_total)

            # Accumulate totals
            totals['quantity'] += item.get('quantity', 0)
            totals['col5'] += selling_price_per_unit
            totals['col6'] += selling_price_total
            totals['col7'] += vat_from_sales
            totals['col8'] += selling_price_with_vat_per_unit
            totals['col9'] += selling_price_with_vat_total

            # Apply currency format to monetary columns (5-9)
            for col in range(5, 10):
                ws1.cell(row=row_idx, column=col).number_format = '#,##0.00 ₽'

            row_idx += 1

        # Totals row
        ws1.cell(row=row_idx, column=1).value = 'ИТОГО'
        ws1.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        ws1.cell(row=row_idx, column=4).value = totals['quantity']
        ws1.cell(row=row_idx, column=4).font = Font(bold=True)

        for col_idx, col_key in enumerate(['col5', 'col6', 'col7', 'col8', 'col9'], 5):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.value = float(totals[col_key])
            cell.number_format = '#,##0.00 ₽'
            cell.font = Font(bold=True)

        # Set column widths
        ws1.column_dimensions['A'].width = 12  # Бренд
        ws1.column_dimensions['B'].width = 15  # Артикул
        ws1.column_dimensions['C'].width = 35  # Наименование
        ws1.column_dimensions['D'].width = 10  # Кол-во
        for col_letter in ['E', 'F', 'G', 'H', 'I']:
            ws1.column_dimensions[col_letter].width = 15

        # Freeze header row
        ws1.freeze_panes = 'A2'

        # ========== Sheet 2: КП open book (20 columns) ==========
        ws2 = wb.create_sheet("КП open book")

        headers_openbook = [
            'Бренд', 'Артикул', 'Наименование', 'Кол-во', 'Валюта',
            'Цена без НДС', 'Сумма инвойса', 'Цена в валюте КП (₽)',
            'Логистика (₽)', 'Код ТН ВЭД', 'Пошлина (%)', 'Таможенный сбор (₽)',
            'Акциз (₽)', 'Утиль. сбор (₽)', 'Комиссия транзит (₽)',
            'Цена за ед. (₽)', 'Сумма (₽)', 'НДС (₽)',
            'Цена с НДС за ед. (₽)', 'Сумма с НДС (₽)'
        ]

        for col_idx, header in enumerate(headers_openbook, 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10, color="FFFFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data rows for Sheet 2
        row_idx = 2
        for item in export_data.items:
            calc_results = item.get('calculation_results', {})

            # Columns 1-4: Same as Sheet 1
            ws2.cell(row=row_idx, column=1).value = item.get('brand', '')
            ws2.cell(row=row_idx, column=2).value = item.get('product_code', '')
            ws2.cell(row=row_idx, column=3).value = item.get('product_name', '')
            ws2.cell(row=row_idx, column=4).value = item.get('quantity', 0)

            # Column 5: Currency
            ws2.cell(row=row_idx, column=5).value = export_data.variables.get('currency_of_base_price', 'USD')

            # Column 6: Price without VAT
            purchase_price_no_vat = Decimal(str(calc_results.get('purchase_price_no_vat', 0)))
            ws2.cell(row=row_idx, column=6).value = float(purchase_price_no_vat)

            # Column 7: Invoice amount (FORMULA: Price × Quantity)
            quantity = item.get('quantity', 0)
            invoice_amount = purchase_price_no_vat * Decimal(str(quantity))
            ws2.cell(row=row_idx, column=7).value = float(invoice_amount)

            # Column 8: Price in quote currency
            ws2.cell(row=row_idx, column=8).value = float(Decimal(str(calc_results.get('purchase_price_total_quote_currency', 0))))

            # Column 9: Logistics
            ws2.cell(row=row_idx, column=9).value = float(Decimal(str(calc_results.get('logistics_total', 0))))

            # Column 10: Customs code
            ws2.cell(row=row_idx, column=10).value = item.get('customs_code', '')

            # Column 11: Import tariff (%)
            ws2.cell(row=row_idx, column=11).value = float(item.get('import_tariff', 0))

            # Column 12: Customs fee
            ws2.cell(row=row_idx, column=12).value = float(Decimal(str(calc_results.get('customs_fee', 0))))

            # Column 13: Excise tax
            ws2.cell(row=row_idx, column=13).value = float(item.get('excise_tax', 0))

            # Column 14: Util fee
            ws2.cell(row=row_idx, column=14).value = float(Decimal(str(calc_results.get('util_fee', 0))))

            # Column 15: Transit commission
            ws2.cell(row=row_idx, column=15).value = float(Decimal(str(calc_results.get('transit_commission', 0))))

            # Columns 16-20: Same selling prices as Sheet 1
            ws2.cell(row=row_idx, column=16).value = float(Decimal(str(calc_results.get('sales_price_per_unit', 0))))
            ws2.cell(row=row_idx, column=17).value = float(Decimal(str(calc_results.get('sales_price_total_no_vat', 0))))
            ws2.cell(row=row_idx, column=18).value = float(Decimal(str(calc_results.get('vat_amount', 0))))
            ws2.cell(row=row_idx, column=19).value = float(Decimal(str(calc_results.get('sales_price_per_unit_with_vat', 0))))
            ws2.cell(row=row_idx, column=20).value = float(Decimal(str(calc_results.get('sales_price_total_with_vat', 0))))

            # Apply number formats
            # Monetary columns: 6, 7, 8, 9, 12, 14, 15, 16, 17, 18, 19, 20
            for col in [6, 7, 8, 9, 12, 14, 15, 16, 17, 18, 19, 20]:
                ws2.cell(row=row_idx, column=col).number_format = '#,##0.00 ₽'

            # Percentage columns: 11
            ws2.cell(row=row_idx, column=11).number_format = '0.00%'

            # Excise tax (column 13) - monetary
            ws2.cell(row=row_idx, column=13).number_format = '#,##0.00 ₽'

            row_idx += 1

        # Set column widths for Sheet 2
        ws2.column_dimensions['A'].width = 10  # Бренд
        ws2.column_dimensions['B'].width = 12  # Артикул
        ws2.column_dimensions['C'].width = 30  # Наименование
        ws2.column_dimensions['D'].width = 8   # Кол-во
        ws2.column_dimensions['E'].width = 8   # Валюта
        ws2.column_dimensions['J'].width = 13  # Код ТН ВЭД

        # Other columns default to 12
        for col_letter in ['F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
            ws2.column_dimensions[col_letter].width = 12

        # Freeze header row
        ws2.freeze_panes = 'A2'

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_supply_grid_export(export_data: ExportData) -> bytes:
        """
        Generate КП поставка Excel export (9 columns) with header info
        
        Structure:
        - Rows 1-11: General info (Seller, Buyer, Quote details)
        - Row 12: Blank
        - Row 13: Column headers
        - Row 14+: Data
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "КП поставка"
        
        # Styling
        header_fill = PatternFill(start_color="FF2C5AA0", end_color="FF2C5AA0", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFFFF")
        info_font = Font(bold=True, size=11)
        
        # ==== 3-COLUMN HEADER CARDS (Like PDF) ====
        card_fill = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
        card_border = Border(
            left=Side(style='thin', color='FF2C5AA0'),
            right=Side(style='thin', color='FF2C5AA0'),
            top=Side(style='thin', color='FF2C5AA0'),
            bottom=Side(style='thin', color='FF2C5AA0')
        )
        label_font = Font(bold=True, size=10)
        value_font = Font(size=10)
        title_font = Font(bold=True, size=11, color="FF2C5AA0")

        # Card titles row
        ws['A1'] = "Продавец"
        ws['A1'].font = title_font
        ws['D1'] = "Покупатель"
        ws['D1'].font = title_font
        ws['G1'] = "Информация о поставке"
        ws['G1'].font = title_font

        # Card 1: Seller (A2:C8)
        seller_data = [
            ("Компания:", export_data.variables.get('seller_company', '')),
            ("Менеджер:", export_data.manager.get('full_name', '') if export_data.manager else ''),
            ("Тел:", export_data.manager.get('phone', '') if export_data.manager else ''),
            ("Email:", export_data.manager.get('email', '') if export_data.manager else ''),
        ]

        row = 2
        for label, value in seller_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = value_font
            ws.merge_cells(f'B{row}:C{row}')
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Card 2: Buyer (D2:F8)
        customer_data = [
            ("Компания:", export_data.customer.get('name', '') if export_data.customer else ''),
            ("Контакт:", export_data.contact.get('full_name', '') if export_data.contact else ''),
            ("Тел:", export_data.contact.get('phone', '') if export_data.contact else ''),
            ("Email:", export_data.contact.get('email', '') if export_data.contact else ''),
        ]

        row = 2
        for label, value in customer_data:
            ws[f'D{row}'] = label
            ws[f'D{row}'].font = label_font
            ws[f'E{row}'] = value
            ws[f'E{row}'].font = value_font
            ws.merge_cells(f'E{row}:F{row}')
            for col in ['D', 'E', 'F']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Card 3: Quote Info (G2:I11)
        # Calculate total
        total = Decimal('0')
        for item in export_data.items:
            calc_results = item.get('calculation_results', {})
            total += Decimal(str(calc_results.get('sales_price_total_with_vat', 0)))

        quote_data = [
            ("Дата КП:", export_data.quote.get('quote_date', '')),
            ("Срок поставки:", f"{export_data.variables.get('delivery_time_days', '')} дней"),
            ("Базис:", export_data.variables.get('offer_incoterms', '')),
            ("Оплата:", f"{export_data.variables.get('advance_from_client', '')}%"),
            ("Сумма с НДС:", f"{float(total):,.2f} ₽"),
        ]

        # Add delivery description if present
        delivery_desc = export_data.quote.get('delivery_terms', '')
        if delivery_desc:
            quote_data.append(("", delivery_desc))  # No label, just description

        row = 2
        for label, value in quote_data:
            ws[f'G{row}'] = label
            ws[f'G{row}'].font = label_font if label else Font(size=9, italic=True)
            ws[f'H{row}'] = value
            ws[f'H{row}'].font = Font(bold=True, size=11, color="FF2C5AA0") if "Сумма" in label else value_font
            ws.merge_cells(f'H{row}:I{row}')
            for col in ['G', 'H', 'I']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Blank row before table
        row = max(6, row) + 1
        
        # ==== DATA TABLE ====
        table_start_row = row
        
        # Column headers
        headers = [
            'Бренд', 'Артикул', 'Наименование', 'Кол-во',
            'Цена за ед. (₽)', 'Сумма (₽)', 'НДС (₽)',
            'Цена с НДС за ед. (₽)', 'Сумма с НДС (₽)'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row += 1
        
        # Data rows
        totals = {'quantity': 0, 'col5': Decimal('0'), 'col6': Decimal('0'), 'col7': Decimal('0'), 'col8': Decimal('0'), 'col9': Decimal('0')}
        
        for item in export_data.items:
            calc_results = item.get('calculation_results', {})
            
            ws.cell(row=row, column=1).value = item.get('brand', '')
            ws.cell(row=row, column=2).value = item.get('product_code', '')
            ws.cell(row=row, column=3).value = item.get('product_name', '')
            ws.cell(row=row, column=4).value = item.get('quantity', 0)

            selling_price_per_unit = Decimal(str(calc_results.get('sales_price_per_unit_no_vat', 0)))
            selling_price_total = Decimal(str(calc_results.get('sales_price_total_no_vat', 0)))
            vat_amount = Decimal(str(calc_results.get('vat_net_payable', 0)))
            selling_price_with_vat_per_unit = Decimal(str(calc_results.get('sales_price_per_unit_with_vat', 0)))
            selling_price_total_with_vat = Decimal(str(calc_results.get('sales_price_total_with_vat', 0)))
            
            ws.cell(row=row, column=5).value = float(selling_price_per_unit)
            ws.cell(row=row, column=5).number_format = '#,##0.00 ₽'
            
            ws.cell(row=row, column=6).value = float(selling_price_total)
            ws.cell(row=row, column=6).number_format = '#,##0.00 ₽'
            
            ws.cell(row=row, column=7).value = float(vat_amount)
            ws.cell(row=row, column=7).number_format = '#,##0.00 ₽'
            
            ws.cell(row=row, column=8).value = float(selling_price_with_vat_per_unit)
            ws.cell(row=row, column=8).number_format = '#,##0.00 ₽'
            
            ws.cell(row=row, column=9).value = float(selling_price_total_with_vat)
            ws.cell(row=row, column=9).number_format = '#,##0.00 ₽'
            
            totals['quantity'] += item.get('quantity', 0)
            totals['col5'] += selling_price_per_unit
            totals['col6'] += selling_price_total
            totals['col7'] += vat_amount
            totals['col8'] += selling_price_with_vat_per_unit
            totals['col9'] += selling_price_total_with_vat
            
            row += 1
        
        # Totals row
        ws.cell(row=row, column=3).value = "ИТОГО:"
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).value = totals['quantity']
        ws.cell(row=row, column=4).font = Font(bold=True)
        
        for col_idx, col_key in enumerate(['col5', 'col6', 'col7', 'col8', 'col9'], 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = float(totals[col_key])
            cell.number_format = '#,##0.00 ₽'
            cell.font = Font(bold=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 10
        for col_letter in ['E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col_letter].width = 15
        
        # Freeze panes at table start
        ws.freeze_panes = f'A{table_start_row + 1}'
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_openbook_grid_export(export_data: ExportData) -> bytes:
        """
        Generate КП open book Excel export (21 columns) with header info
        
        Structure:
        - Rows 1-11: General info (Seller, Buyer, Quote details)
        - Row 12: Blank
        - Row 13: Column headers
        - Row 14+: Data
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "КП open book"
        
        # Styling
        header_fill = PatternFill(start_color="FF2C5AA0", end_color="FF2C5AA0", fill_type="solid")
        header_font = Font(bold=True, size=10, color="FFFFFFFF")
        info_font = Font(bold=True, size=11)
        
        # ==== 3-COLUMN HEADER CARDS (Like PDF) ====
        card_fill = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
        card_border = Border(
            left=Side(style='thin', color='FF2C5AA0'),
            right=Side(style='thin', color='FF2C5AA0'),
            top=Side(style='thin', color='FF2C5AA0'),
            bottom=Side(style='thin', color='FF2C5AA0')
        )
        label_font = Font(bold=True, size=10)
        value_font = Font(size=10)
        title_font = Font(bold=True, size=11, color="FF2C5AA0")

        # Card titles row
        ws['A1'] = "Продавец"
        ws['A1'].font = title_font
        ws['H1'] = "Покупатель"
        ws['H1'].font = title_font
        ws['O1'] = "Информация о поставке"
        ws['O1'].font = title_font

        # Card 1: Seller (A2:G8) - spans 7 columns for wider card
        seller_data = [
            ("Компания:", export_data.variables.get('seller_company', '')),
            ("Менеджер:", export_data.manager.get('full_name', '') if export_data.manager else ''),
            ("Тел:", export_data.manager.get('phone', '') if export_data.manager else ''),
            ("Email:", export_data.manager.get('email', '') if export_data.manager else ''),
        ]

        row = 2
        for label, value in seller_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = value_font
            ws.merge_cells(f'B{row}:G{row}')
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Empty rows to match card height
        for _ in range(2):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Card 2: Buyer (H2:N8) - spans 7 columns
        customer_data = [
            ("Компания:", export_data.customer.get('name', '') if export_data.customer else ''),
            ("Контакт:", export_data.contact.get('full_name', '') if export_data.contact else ''),
            ("Тел:", export_data.contact.get('phone', '') if export_data.contact else ''),
            ("Email:", export_data.contact.get('email', '') if export_data.contact else ''),
        ]

        row = 2
        for label, value in customer_data:
            ws[f'H{row}'] = label
            ws[f'H{row}'].font = label_font
            ws[f'I{row}'] = value
            ws[f'I{row}'].font = value_font
            ws.merge_cells(f'I{row}:N{row}')
            for col in ['H', 'I', 'J', 'K', 'L', 'M', 'N']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Empty rows to match card height
        for _ in range(2):
            for col in ['H', 'I', 'J', 'K', 'L', 'M', 'N']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Card 3: Quote Info (O2:U11) - spans 7 columns, taller card
        # Calculate total
        total = Decimal('0')
        for item in export_data.items:
            calc_results = item.get('calculation_results', {})
            total += Decimal(str(calc_results.get('sales_price_total_with_vat', 0)))

        quote_data = [
            ("Дата КП:", export_data.quote.get('quote_date', '')),
            ("Срок поставки:", f"{export_data.variables.get('delivery_time_days', '')} дней"),
            ("Базис:", export_data.variables.get('offer_incoterms', '')),
            ("Оплата:", f"{export_data.variables.get('advance_from_client', '')}%"),
            ("Сумма с НДС:", f"{float(total):,.2f} ₽"),
        ]

        row = 2
        for label, value in quote_data:
            ws[f'O{row}'] = label
            ws[f'O{row}'].font = label_font
            ws[f'P{row}'] = value
            ws[f'P{row}'].font = value_font
            ws.merge_cells(f'P{row}:U{row}')
            for col in ['O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Delivery description (if present) - spans 2 rows
        delivery_desc = export_data.quote.get('delivery_terms', '')
        if delivery_desc:
            ws[f'O{row}'] = delivery_desc
            ws[f'O{row}'].font = Font(size=9, italic=True)
            ws[f'O{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'O{row}:U{row+1}')
            for r in [row, row+1]:
                for col in ['O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                    ws[f'{col}{r}'].fill = card_fill
                    ws[f'{col}{r}'].border = card_border
            row += 2
        else:
            # Empty row to match card height
            for col in ['O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                ws[f'{col}{row}'].fill = card_fill
                ws[f'{col}{row}'].border = card_border
            row += 1

        # Blank row before table
        row = 9
        
        # ==== DATA TABLE ====
        table_start_row = row
        
        # Column headers (21 columns)
        headers = [
            'Бренд', 'Артикул', 'Наименование', 'Кол-во', 'Валюта',
            'Цена без НДС', 'Сумма инвойса', 'Цена в валюте КП (₽)',
            'Логистика (₽)', 'Код ТН ВЭД', 'Пошлина (%)', 'Таможенный сбор (₽)',
            'Акциз (₽)', 'Утиль. сбор (₽)', 'Комиссия транзит (₽)',
            'Цена за ед. (₽)', 'Сумма (₽)', 'НДС (₽)',
            'Цена с НДС за ед. (₽)', 'Сумма с НДС (₽)'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row += 1
        
        # Data rows
        for item in export_data.items:
            calc_results = item.get('calculation_results', {})
            
            # Columns 1-4: Basic info
            ws.cell(row=row, column=1).value = item.get('brand', '')
            ws.cell(row=row, column=2).value = item.get('product_code', '')
            ws.cell(row=row, column=3).value = item.get('product_name', '')
            ws.cell(row=row, column=4).value = item.get('quantity', 0)

            # Column 5: Currency
            ws.cell(row=row, column=5).value = export_data.variables.get('currency_of_base_price', 'USD')
            
            # Column 6-7: Purchase prices
            purchase_price_no_vat = Decimal(str(calc_results.get('purchase_price_no_vat', 0)))
            purchase_price_total = Decimal(str(calc_results.get('purchase_price_total_quote_currency', 0)))

            ws.cell(row=row, column=6).value = float(purchase_price_no_vat)
            ws.cell(row=row, column=7).value = float(purchase_price_total)

            # Column 8-9: Converted price and logistics (per unit values)
            quantity = item.get('quantity', 1)
            purchase_price_per_unit = purchase_price_total / quantity if quantity > 0 else Decimal('0')
            logistics_per_unit = Decimal(str(calc_results.get('logistics_total', 0))) / quantity if quantity > 0 else Decimal('0')

            ws.cell(row=row, column=8).value = float(purchase_price_per_unit)
            ws.cell(row=row, column=9).value = float(logistics_per_unit)
            
            # Column 10-15: Customs and fees
            ws.cell(row=row, column=10).value = item.get('customs_code', '')
            # Import tariff comes from variables, not calc_results
            import_tariff = export_data.variables.get('import_tariff', 0)
            ws.cell(row=row, column=11).value = float(Decimal(str(import_tariff)))
            ws.cell(row=row, column=12).value = float(Decimal(str(calc_results.get('customs_fee', 0))))
            ws.cell(row=row, column=13).value = float(Decimal(str(calc_results.get('excise_tax_amount', 0))))
            ws.cell(row=row, column=14).value = float(Decimal(str(calc_results.get('recycling_fee', 0))))
            ws.cell(row=row, column=15).value = float(Decimal(str(calc_results.get('transit_commission', 0))))

            # Column 16-20: Selling prices
            ws.cell(row=row, column=16).value = float(Decimal(str(calc_results.get('sales_price_per_unit_no_vat', 0))))
            ws.cell(row=row, column=16).number_format = '#,##0.00 ₽'

            ws.cell(row=row, column=17).value = float(Decimal(str(calc_results.get('sales_price_total_no_vat', 0))))
            ws.cell(row=row, column=17).number_format = '#,##0.00 ₽'

            ws.cell(row=row, column=18).value = float(Decimal(str(calc_results.get('vat_net_payable', 0))))
            ws.cell(row=row, column=18).number_format = '#,##0.00 ₽'

            ws.cell(row=row, column=19).value = float(Decimal(str(calc_results.get('sales_price_per_unit_with_vat', 0))))
            ws.cell(row=row, column=19).number_format = '#,##0.00 ₽'

            ws.cell(row=row, column=20).value = float(Decimal(str(calc_results.get('sales_price_total_with_vat', 0))))
            ws.cell(row=row, column=20).number_format = '#,##0.00 ₽'
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['J'].width = 15
        
        for col_letter in ['F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
            ws.column_dimensions[col_letter].width = 12
        
        # Freeze panes at table start
        ws.freeze_panes = f'A{table_start_row + 1}'
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
