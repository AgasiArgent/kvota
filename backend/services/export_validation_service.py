"""
Export Validation Service - Generates Excel with API vs Excel comparison

Creates an Excel file with:
1. API_Inputs tab - All uploaded input values
2. расчет tab - Modified to reference API_Inputs (formulas recalculate)
3. API_Results tab - API calculation outputs
4. Conditional formatting highlighting differences > 0.01%
"""

import io
import logging
from copy import copy
from decimal import Decimal
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Template file path
TEMPLATE_PATH = "/home/novi/Downloads/test_raschet_new_template_vat22.xlsm"

# Highlight style for differences > 0.01%
DIFF_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Thin border style
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# VALUE MAPPINGS - Convert API values to Excel-expected values
# =============================================================================

# Offer sale type mapping (API -> Russian dropdown values)
SALE_TYPE_MAP = {
    # English/code values
    "openbook": "поставка",
    "supply": "поставка",
    "transit": "транзит",
    "fin_transit": "финтранзит",
    "export": "экспорт",
    # Russian values (passthrough)
    "поставка": "поставка",
    "транзит": "транзит",
    "финтранзит": "финтранзит",
    "экспорт": "экспорт",
}

# Supplier country mapping (code -> Russian dropdown values)
# Includes both 2-letter codes and Russian names (for passthrough)
COUNTRY_MAP = {
    # 2-letter country codes
    "TR": "Турция",
    "CN": "Китай",
    "RU": "Россия",
    "DE": "Германия",
    "IT": "Италия",
    "LT": "Литва",
    "LV": "Латвия",
    "BG": "Болгария",
    "PL": "Польша",
    "AE": "ОАЭ",
    # Russian names (passthrough)
    "Турция": "Турция",
    "Китай": "Китай",
    "Россия": "Россия",
    "Германия": "Германия",
    "Италия": "Италия",
    "Литва": "Литва",
    "Латвия": "Латвия",
    "Болгария": "Болгария",
    "Польша": "Польша",
    "ОАЭ": "ОАЭ",
    "Прочие": "Прочие",
    "Турция (транзитная зона)": "Турция (транзитная зона)",
    "ЕС (закупка между странами ЕС)": "ЕС (закупка между странами ЕС)",
}

# Fields that are percentages (need to be divided by 100)
PERCENTAGE_FIELDS = {
    "advance_to_supplier",
    "advance_from_client",
    "advance_on_loading",
    "advance_on_shipping",
    "advance_on_customs",
    "rate_forex_risk",
    "supplier_discount",
    "import_tariff",
    "markup",
}


# =============================================================================
# CELL MAPPINGS - Quote Level Inputs
# =============================================================================

QUOTE_INPUT_MAPPING = {
    # Cell in расчет -> (API field path, display name)
    "D5": ("seller_company", "Компания-продавец"),
    "D6": ("offer_sale_type", "Вид КП"),
    "D7": ("incoterms", "Базис поставки"),
    "D8": ("quote_currency", "Валюта КП"),
    "D9": ("delivery_days", "Срок поставки (дней)"),
    "D11": ("advance_to_supplier", "Аванс поставщику (%)"),
    # Payment milestones
    "J5": ("advance_from_client", "Аванс от клиента (%)"),
    "K5": ("time_to_advance", "Дней до аванса"),
    "J6": ("advance_on_loading", "Аванс при заборе (%)"),
    "K6": ("time_to_advance_loading", "Дней до аванса загрузки"),
    "J7": ("advance_on_shipping", "Аванс при отправке (%)"),
    "K7": ("time_to_advance_shipping", "Дней до аванса отправки"),
    "J8": ("advance_on_customs", "Аванс при таможне (%)"),
    "K8": ("time_to_advance_customs", "Дней до аванса таможни"),
    # NOTE: time_to_payment is now mapped to F7 (days after receiving)
    "F7": ("time_to_payment", "Дней до оплаты после получения"),
    # Logistics costs
    "W2": ("logistics_supplier_hub", "Логистика: Поставщик-Хаб"),
    "W3": ("logistics_hub_customs", "Логистика: Хаб-Таможня"),
    "W4": ("logistics_customs_client", "Логистика: Таможня-Клиент"),
    # Brokerage costs
    "W5": ("brokerage_hub", "Брокерские услуги: Хаб"),
    "W6": ("brokerage_customs", "Брокерские услуги: Таможня"),
    "W7": ("warehousing", "Расходы на СВХ"),
    "W8": ("documentation", "Разрешительные документы"),
    "W9": ("other_costs", "Прочее"),
    # DM Fee - type determines placement:
    # AG3: dm_fee_type ("Фикс" or "комиссия %")
    # AG4: dm_fee_value when type is "Фикс" (fixed amount)
    # AG6: dm_fee_value when type is "комиссия %" (percentage, divided by 100)
    "AG3": ("dm_fee_type", "Тип вознаграждения ЛПР"),
    # NOTE: dm_fee_value is handled separately in _modify_raschet_references()
    # Admin settings
    "AH11": ("rate_forex_risk", "Резерв на курсовую разницу (%)"),
}

# =============================================================================
# CELL MAPPINGS - Product Level Inputs (Row 16 = first product)
# =============================================================================

PRODUCT_INPUT_COLUMNS = {
    # Column letter -> (API field, display name)
    "B": ("brand", "Бренд"),
    "C": ("sku", "Артикул"),
    "D": ("name", "Название товара"),
    "E": ("quantity", "Количество"),
    "G": ("weight_in_kg", "Вес, кг"),
    "J": ("currency_of_base_price", "Валюта закупки"),
    "K": ("base_price_vat", "Цена закупки (с VAT)"),
    "L": ("supplier_country", "Страна закупки"),
    "O": ("supplier_discount", "Скидка поставщика (%)"),
    "Q": ("exchange_rate", "Курс к валюте КП"),
    "W": ("customs_code", "Код ТН ВЭД"),
    "X": ("import_tariff", "Пошлина (%)"),
    "AC": ("markup", "Наценка (%)"),
}

# =============================================================================
# CELL MAPPINGS - Product Level Outputs (Calculated)
# =============================================================================

PRODUCT_OUTPUT_COLUMNS = {
    # Column letter -> (API field from ProductCalculationResult, display name)
    "N": ("purchase_price_no_vat", "Цена без VAT"),
    "P": ("purchase_price_after_discount", "После скидки"),
    "R": ("purchase_price_per_unit_quote_currency", "Цена за ед. в валюте КП"),
    "S": ("purchase_price_total_quote_currency", "Стоимость закупки"),
    "T": ("logistics_first_leg", "Логистика первый этап"),
    "U": ("logistics_last_leg", "Логистика второй этап"),
    "V": ("logistics_total", "Логистика итого"),
    "Y": ("customs_fee", "Пошлина (сумма)"),
    "Z": ("excise_tax_amount", "Акциз (сумма)"),
    "AA": ("cogs_per_unit", "Себестоимость за ед."),
    "AB": ("cogs_per_product", "Себестоимость товара"),
    "AD": ("sale_price_per_unit_excl_financial", "Цена продажи за ед. (без фин.)"),
    "AE": ("sale_price_total_excl_financial", "Цена продажи итого (без фин.)"),
    "AF": ("profit", "Прибыль"),
    "AG": ("dm_fee", "Вознаграждение ЛПР"),
    "AH": ("forex_reserve", "Резерв на курсовую разницу"),
    "AI": ("financial_agent_fee", "Комиссия ФинАгента"),
    "AJ": ("sales_price_per_unit_no_vat", "Цена продажи за ед. (без НДС)"),
    "AK": ("sales_price_total_no_vat", "Цена продажи итого (без НДС)"),
    "AL": ("sales_price_total_with_vat", "Цена продажи итого (с НДС)"),
    "AM": ("sales_price_per_unit_with_vat", "Цена продажи за ед. (с НДС)"),
    "AN": ("vat_from_sales", "НДС с продажи"),
    "AO": ("vat_on_import", "НДС к вычету"),
    "AP": ("vat_net_payable", "НДС к уплате"),
    "AQ": ("transit_commission", "Транзитная комиссия"),
    "AX": ("internal_sale_price_per_unit", "Внутренняя цена за ед."),
    "AY": ("internal_sale_price_total", "Внутренняя стоимость"),
    "BA": ("financing_cost_initial", "Начальное финансирование"),
    "BB": ("financing_cost_credit", "Проценты по отсрочке"),
}

# =============================================================================
# CELL MAPPINGS - Quote Level Totals (Row 13)
# =============================================================================

QUOTE_TOTAL_CELLS = {
    # Cell -> (API field, display name)
    "S13": ("total_purchase_price", "Итого стоимость закупки"),
    "T13": ("total_logistics_first", "Логистика первый этап (итого)"),
    "U13": ("total_logistics_last", "Логистика второй этап (итого)"),
    "V13": ("total_logistics", "Логистика итого"),
    "AB13": ("total_cogs", "Себестоимость итого"),
    "AK13": ("total_revenue", "Выручка (без НДС)"),
    "AL13": ("total_revenue_with_vat", "Выручка (с НДС)"),
    "AF13": ("total_profit", "Прибыль итого"),
}

# =============================================================================
# CELL MAPPINGS - Financing Outputs
# =============================================================================

FINANCING_CELLS = {
    # Cell -> (API field, display name)
    "BH2": ("evaluated_revenue", "Оценочная выручка"),
    "BH3": ("client_advance", "Аванс клиента"),
    "BH4": ("total_before_forwarding", "Итого до экспедирования"),
    "BH6": ("supplier_payment", "Платеж поставщику"),
    "BJ7": ("supplier_financing_cost", "Стоимость фин-ия поставщика"),
    "BJ10": ("operational_financing_cost", "Стоимость операционного фин-ия"),
    "BJ11": ("total_financing_cost", "Итого стоимость финансирования"),
    "BL3": ("credit_sales_amount", "Сумма к оплате клиентом"),
    "BL4": ("credit_sales_fv", "FV с процентами"),
    "BL5": ("credit_sales_interest", "Проценты по отсрочке"),
}


class ExportValidationService:
    """Service to generate validation Excel with API vs Excel comparison."""

    def __init__(self, template_path: str = TEMPLATE_PATH):
        self.template_path = template_path

    def generate_validation_export(
        self,
        quote_inputs: Dict[str, Any],
        product_inputs: List[Dict[str, Any]],
        api_results: Dict[str, Any],
        product_results: List[Dict[str, Any]],
    ) -> bytes:
        """
        Generate validation Excel file.

        Args:
            quote_inputs: Quote-level input values
            product_inputs: List of product input dicts
            api_results: Quote-level API results (totals, financing)
            product_results: List of product result dicts from API

        Returns:
            Excel file as bytes
        """
        # Load template
        wb = openpyxl.load_workbook(self.template_path, keep_vba=True)

        # 1. Create API_Inputs sheet
        self._create_inputs_sheet(wb, quote_inputs, product_inputs)

        # 2. Modify расчет to reference API_Inputs
        self._modify_raschet_references(wb, len(product_inputs))

        # 3. Create API_Results sheet
        self._create_results_sheet(wb, api_results, product_results)

        # 4. Add conditional formatting for comparison
        self._add_comparison_formatting(wb, len(product_results))

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_inputs_sheet(
        self,
        wb: openpyxl.Workbook,
        quote_inputs: Dict[str, Any],
        product_inputs: List[Dict[str, Any]],
    ) -> None:
        """Create API_Inputs sheet with all uploaded values."""

        # Create or get sheet
        if "API_Inputs" in wb.sheetnames:
            ws = wb["API_Inputs"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("API_Inputs", 0)

        # Header
        ws["A1"] = "API Input Values"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:C1")

        # Section: Quote Settings
        row = 3
        ws[f"A{row}"] = "Quote Settings"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        ws[f"B{row}"].fill = HEADER_FILL
        ws[f"C{row}"].fill = HEADER_FILL
        row += 1

        ws[f"A{row}"] = "Cell"
        ws[f"B{row}"] = "Field"
        ws[f"C{row}"] = "Value"
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].border = THIN_BORDER
        row += 1

        # Quote input values
        quote_input_start_row = row
        for cell_addr, (field, display_name) in QUOTE_INPUT_MAPPING.items():
            ws[f"A{row}"] = cell_addr
            ws[f"B{row}"] = display_name
            value = quote_inputs.get(field)
            # Use field-aware formatting (percentages, enums)
            ws[f"C{row}"] = self._format_input_value(value, field)
            ws[f"C{row}"].alignment = Alignment(horizontal="right")
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].border = THIN_BORDER
            row += 1

        # Store advance_from_client for D10 calculation
        self._advance_from_client = quote_inputs.get("advance_from_client", 0)

        # Handle dm_fee_value separately - placement depends on type
        dm_fee_type = quote_inputs.get("dm_fee_type", "")
        dm_fee_value = quote_inputs.get("dm_fee_value", 0)

        # Determine target cell based on type:
        # "Фикс" or "фикс. сумма" -> AG4 (fixed amount)
        # "комиссия %" or "% от суммы" -> AG6 (percentage, divide by 100)
        is_percentage = any(pct in str(dm_fee_type).lower() for pct in ["%", "комиссия", "процент"])

        if is_percentage:
            dm_target_cell = "AG6"
            # Treat as percentage - divide by 100
            formatted_value = dm_fee_value / 100 if dm_fee_value else 0
        else:
            dm_target_cell = "AG4"
            formatted_value = dm_fee_value

        # Add dm_fee_value to API_Inputs
        ws[f"A{row}"] = dm_target_cell
        ws[f"B{row}"] = "Вознаграждение ЛПР"
        ws[f"C{row}"] = formatted_value
        ws[f"C{row}"].alignment = Alignment(horizontal="right")
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].border = THIN_BORDER

        # Store for _modify_raschet_references
        self._dm_fee_row = row
        self._dm_fee_target_cell = dm_target_cell
        row += 1

        # Section: Products
        row += 2
        product_header_row = row
        ws[f"A{row}"] = "Products"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        # Product headers
        product_start_row = row + 1
        ws[f"A{row}"] = "#"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = THIN_BORDER

        col_idx = 2
        product_col_map = {}  # column letter -> field name
        for col_letter, (field, display_name) in PRODUCT_INPUT_COLUMNS.items():
            cell = ws.cell(row=row, column=col_idx)
            cell.value = display_name
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            product_col_map[col_idx] = field
            col_idx += 1
        row += 1

        # Product data
        for i, product in enumerate(product_inputs, 1):
            ws[f"A{row}"] = i
            ws[f"A{row}"].border = THIN_BORDER

            col_idx = 2
            for _, (field, _) in PRODUCT_INPUT_COLUMNS.items():
                cell = ws.cell(row=row, column=col_idx)
                value = product.get(field)
                # Use field-aware formatting (percentages, country codes)
                cell.value = self._format_input_value(value, field)
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
                col_idx += 1
            row += 1

        # Store row references as instance attributes for later use
        self._quote_input_start = quote_input_start_row
        self._product_start_row = product_start_row

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 20
        for col in range(4, col_idx + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _modify_raschet_references(
        self, wb: openpyxl.Workbook, num_products: int
    ) -> None:
        """Modify расчет sheet to reference API_Inputs."""

        ws = wb["расчет"]
        inputs_sheet = "API_Inputs"

        # Quote-level inputs - create reference formulas
        row = 5  # Start row in API_Inputs for quote values
        for cell_addr, (field, _) in QUOTE_INPUT_MAPPING.items():
            # Create formula referencing API_Inputs
            # The value is in column C of API_Inputs
            ws[cell_addr] = f"='{inputs_sheet}'!C{row}"
            row += 1

        # D10 - Payment type based on advance_from_client
        # Keep as value, not formula reference (per user request)
        payment_type = self._get_payment_type_value(self._advance_from_client)
        ws["D10"] = payment_type

        # DM Fee value - placed in AG4 (Фикс) or AG6 (комиссия %)
        # The target cell and row were determined in _create_inputs_sheet
        if hasattr(self, '_dm_fee_target_cell') and hasattr(self, '_dm_fee_row'):
            ws[self._dm_fee_target_cell] = f"='{inputs_sheet}'!C{self._dm_fee_row}"

        # Product-level inputs
        # Use the stored row from _create_inputs_sheet (row after headers)
        # Products start at row 16 in расчет
        inputs_product_start = self._product_start_row  # Set by _create_inputs_sheet
        raschet_product_start = 16

        for prod_idx in range(num_products):
            inputs_row = inputs_product_start + prod_idx
            raschet_row = raschet_product_start + prod_idx

            # Map each product input column
            input_col = 2  # Start at column B in API_Inputs
            for col_letter, (field, _) in PRODUCT_INPUT_COLUMNS.items():
                raschet_cell = f"{col_letter}{raschet_row}"
                inputs_cell = f"'{inputs_sheet}'!{get_column_letter(input_col)}{inputs_row}"
                ws[raschet_cell] = f"={inputs_cell}"
                input_col += 1

    def _create_results_sheet(
        self,
        wb: openpyxl.Workbook,
        api_results: Dict[str, Any],
        product_results: List[Dict[str, Any]],
    ) -> None:
        """Create API_Results sheet with calculation outputs."""

        # Create or get sheet
        if "API_Results" in wb.sheetnames:
            ws = wb["API_Results"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("API_Results")

        # Header
        ws["A1"] = "API Calculation Results"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")

        # Section: Quote Totals
        row = 3
        ws[f"A{row}"] = "Quote Totals"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        ws[f"A{row}"] = "Cell"
        ws[f"B{row}"] = "Field"
        ws[f"C{row}"] = "API Value"
        ws[f"D{row}"] = "Excel Value"
        ws[f"E{row}"] = "Diff %"
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].border = THIN_BORDER
        row += 1

        # Quote totals
        for cell_addr, (field, display_name) in QUOTE_TOTAL_CELLS.items():
            ws[f"A{row}"] = cell_addr
            ws[f"B{row}"] = display_name
            api_value = api_results.get(field)
            ws[f"C{row}"] = self._format_value(api_value)
            ws[f"D{row}"] = f"=расчет!{cell_addr}"
            # Diff formula
            ws[f"E{row}"] = f'=IF(D{row}=0,"N/A",ABS(C{row}-D{row})/ABS(D{row})*100)'
            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}{row}"].border = THIN_BORDER
            row += 1

        # Financing cells
        row += 1
        ws[f"A{row}"] = "Financing"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        for cell_addr, (field, display_name) in FINANCING_CELLS.items():
            ws[f"A{row}"] = cell_addr
            ws[f"B{row}"] = display_name
            api_value = api_results.get(field)
            ws[f"C{row}"] = self._format_value(api_value)
            ws[f"D{row}"] = f"=расчет!{cell_addr}"
            ws[f"E{row}"] = f'=IF(D{row}=0,"N/A",ABS(C{row}-D{row})/ABS(D{row})*100)'
            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}{row}"].border = THIN_BORDER
            row += 1

        # Section: Product Results
        row += 2
        ws[f"A{row}"] = "Product Results"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        # Product headers
        ws[f"A{row}"] = "#"
        ws[f"A{row}"].font = Font(bold=True)

        col_idx = 2
        for col_letter, (field, display_name) in PRODUCT_OUTPUT_COLUMNS.items():
            cell = ws.cell(row=row, column=col_idx)
            cell.value = f"{col_letter}: {display_name[:15]}"
            cell.font = Font(bold=True, size=9)
            cell.border = THIN_BORDER
            col_idx += 1
        row += 1

        product_results_start = row

        # Product data
        for i, product in enumerate(product_results):
            raschet_row = 16 + i  # Row in расчет sheet

            ws[f"A{row}"] = i + 1
            ws[f"A{row}"].border = THIN_BORDER

            col_idx = 2
            for col_letter, (field, _) in PRODUCT_OUTPUT_COLUMNS.items():
                cell = ws.cell(row=row, column=col_idx)
                api_value = product.get(field)
                cell.value = self._format_value(api_value)
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
                col_idx += 1
            row += 1

        # Store for conditional formatting
        self._product_results_start = product_results_start
        self._num_products = len(product_results)

        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 10

    def _add_comparison_formatting(
        self, wb: openpyxl.Workbook, num_products: int
    ) -> None:
        """Add conditional formatting to highlight differences > 0.01%."""

        ws = wb["API_Results"]

        # Highlight E column (Diff %) where value > 0.01
        diff_rule = FormulaRule(
            formula=["E5>0.01"],
            fill=DIFF_FILL,
        )

        # Apply to diff column range
        ws.conditional_formatting.add("E5:E100", diff_rule)

        # Also add comparison sheet for detailed product analysis
        self._create_comparison_sheet(wb, num_products)

    def _create_comparison_sheet(
        self, wb: openpyxl.Workbook, num_products: int
    ) -> None:
        """Create detailed comparison sheet for products."""

        if "Comparison" in wb.sheetnames:
            ws = wb["Comparison"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("Comparison")

        ws["A1"] = "Detailed Comparison (Cells with >0.01% difference highlighted)"
        ws["A1"].font = Font(bold=True, size=12)
        ws.merge_cells("A1:F1")

        row = 3
        ws[f"A{row}"] = "Product"
        ws[f"B{row}"] = "Cell"
        ws[f"C{row}"] = "Field"
        ws[f"D{row}"] = "API Value"
        ws[f"E{row}"] = "Excel Value"
        ws[f"F{row}"] = "Diff %"
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].fill = HEADER_FILL
            ws[f"{col}{row}"].font = HEADER_FONT
            ws[f"{col}{row}"].border = THIN_BORDER
        row += 1

        # For each product and each output column
        for prod_idx in range(num_products):
            raschet_row = 16 + prod_idx
            api_results_row = self._product_results_start + prod_idx

            for col_letter, (field, display_name) in PRODUCT_OUTPUT_COLUMNS.items():
                ws[f"A{row}"] = prod_idx + 1
                ws[f"B{row}"] = f"{col_letter}{raschet_row}"
                ws[f"C{row}"] = display_name

                # API value from API_Results
                api_col = list(PRODUCT_OUTPUT_COLUMNS.keys()).index(col_letter) + 2
                ws[f"D{row}"] = f"=API_Results!{get_column_letter(api_col)}{api_results_row}"

                # Excel value from расчет
                ws[f"E{row}"] = f"=расчет!{col_letter}{raschet_row}"

                # Diff percentage
                ws[f"F{row}"] = f'=IF(E{row}=0,"N/A",ABS(D{row}-E{row})/ABS(E{row})*100)'

                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws[f"{col}{row}"].border = THIN_BORDER
                row += 1

        # Conditional formatting for diff > 0.01%
        diff_rule = FormulaRule(
            formula=["$F4>0.01"],
            fill=DIFF_FILL,
        )
        ws.conditional_formatting.add(f"A4:F{row}", diff_rule)

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 12

    def _format_value(self, value: Any) -> Any:
        """Format value for Excel cell."""
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, str):
            # Try to convert numeric strings
            try:
                return float(value)
            except ValueError:
                return value
        return value

    def _format_input_value(self, value: Any, field_name: str) -> Any:
        """
        Format input value for Excel, with field-aware transformations.

        - Percentages: 30 -> 0.3 (Excel expects decimal format)
        - Sale type: openbook -> поставка
        - Country: TR -> Турция
        """
        if value is None:
            return ""

        # Convert sale type to Russian
        if field_name == "offer_sale_type":
            str_val = str(value).lower()
            return SALE_TYPE_MAP.get(str_val, str(value))

        # Convert country code to Russian
        # Handle both 2-letter codes (TR) and Russian names (Турция)
        if field_name == "supplier_country":
            str_val = str(value)
            # First try exact match (handles Russian names)
            if str_val in COUNTRY_MAP:
                return COUNTRY_MAP[str_val]
            # Then try uppercase (handles 2-letter codes like "tr" -> "TR")
            upper_val = str_val.upper()
            if upper_val in COUNTRY_MAP:
                return COUNTRY_MAP[upper_val]
            # Return as-is if no match
            return str_val

        # Convert percentage fields (30 -> 0.3)
        if field_name in PERCENTAGE_FIELDS:
            try:
                num_val = float(value) if not isinstance(value, (int, float, Decimal)) else float(value)
                # If value > 1, assume it's in percent form (30 means 30%)
                if num_val > 1:
                    return num_val / 100
                return num_val
            except (ValueError, TypeError):
                return value

        # Default formatting
        return self._format_value(value)

    def _get_payment_type_value(self, advance_from_client: float) -> str:
        """
        Get payment type for D10 based on advance_from_client percentage.

        Returns:
            - "100% предоплата" if advance is 100% (or 1.0)
            - "100% постоплата" if advance is 0%
            - "частичная оплата" otherwise
        """
        # Normalize: if > 1, assume percent form
        if advance_from_client > 1:
            advance_from_client = advance_from_client / 100

        if advance_from_client >= 0.99:  # ~100%
            return "100% предоплата"
        elif advance_from_client <= 0.01:  # ~0%
            return "100% постоплата"
        else:
            return "частичная оплата"


# Convenience function
def generate_validation_export(
    quote_inputs: Dict[str, Any],
    product_inputs: List[Dict[str, Any]],
    api_results: Dict[str, Any],
    product_results: List[Dict[str, Any]],
) -> bytes:
    """Generate validation Excel file."""
    service = ExportValidationService()
    return service.generate_validation_export(
        quote_inputs, product_inputs, api_results, product_results
    )
