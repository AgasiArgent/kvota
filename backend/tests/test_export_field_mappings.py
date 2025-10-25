#!/usr/bin/env python3
"""
Export Field Mapping Validation Tests
Session 24 - Automated validation of field mappings across all export formats

Tests all 7 export formats (4 PDF + 3 Excel) to ensure:
- Field mappings are correct
- Calculation results properly mapped to export templates
- Critical columns contain valid data (not 0 or empty)
"""
import pytest
import sys
import os
from decimal import Decimal
from openpyxl import load_workbook
import asyncpg
from io import BytesIO

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + '/..')

from pdf_service import QuotePDFService
from services.excel_service import QuoteExcelService
from services.export_data_mapper import fetch_export_data

# Test configuration
TEST_QUOTE_ID = "3b2e851a-a579-4019-aec7-4ac940ea520b"  # КП25-0009
DATABASE_URL = os.getenv("DATABASE_URL")


@pytest.fixture
async def export_data():
    """Fetch test quote export data"""
    if not DATABASE_URL:
        pytest.skip("DATABASE_URL not configured")

    conn = await asyncpg.connect(DATABASE_URL)
    try:
        data = await fetch_export_data(conn, TEST_QUOTE_ID)
        return data
    finally:
        await conn.close()


class TestSupplyGridFieldMappings:
    """Validate supply grid Excel export field mappings"""

    @pytest.mark.asyncio
    async def test_supply_grid_columns_7_8_9(self, export_data):
        """Validate supply grid columns 7, 8, 9 (selling prices) have correct values"""
        service = QuoteExcelService()
        excel_bytes = await service.generate_supply_grid_excel(export_data)

        # Save and load workbook
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Find grid start row (after header)
        grid_start = None
        for row in range(1, 25):
            val = ws.cell(row=row, column=1).value
            if val and ("№" in str(val) or "Бренд" in str(val)):
                grid_start = row + 1
                break

        assert grid_start is not None, "Could not find grid start row"

        # Validate first data row has values
        col_7 = ws.cell(row=grid_start, column=7).value  # sales_price_per_unit_no_vat
        col_8 = ws.cell(row=grid_start, column=8).value  # sales_price_total_no_vat
        col_9 = ws.cell(row=grid_start, column=9).value  # sales_price_total_with_vat

        assert col_7 is not None and col_7 > 0, \
            f"Column 7 (Цена за ед. без НДС) is {col_7}, should be > 0"
        assert col_8 is not None and col_8 > 0, \
            f"Column 8 (Сумма без НДС) is {col_8}, should be > 0"
        assert col_9 is not None and col_9 > 0, \
            f"Column 9 (Сумма с НДС) is {col_9}, should be > 0"


class TestOpenbookGridFieldMappings:
    """Validate openbook grid Excel export field mappings"""

    @pytest.mark.asyncio
    async def test_openbook_grid_columns_7_8_9_10_11(self, export_data):
        """Validate openbook grid columns 7-11 (invoice, logistics, customs) have correct values"""
        service = QuoteExcelService()
        excel_bytes = await service.generate_openbook_grid_excel(export_data)

        # Save and load workbook
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Find grid start row
        grid_start = None
        for row in range(1, 25):
            val = ws.cell(row=row, column=1).value
            if val and "№" in str(val):
                grid_start = row + 1
                break

        assert grid_start is not None, "Could not find grid start row"

        # Validate critical columns
        col_7 = ws.cell(row=grid_start, column=7).value   # purchase_price_total_quote_currency
        col_8 = ws.cell(row=grid_start, column=8).value   # purchase_price_per_unit_quote_currency
        col_9 = ws.cell(row=grid_start, column=9).value   # logistics_per_unit
        col_10 = ws.cell(row=grid_start, column=10).value # customs_code
        col_11 = ws.cell(row=grid_start, column=11).value # import_tariff %

        assert col_7 is not None and col_7 > 0, \
            f"Column 7 (Сумма инвойса) is {col_7}, should be > 0"
        assert col_8 is not None and col_8 > 0, \
            f"Column 8 (Цена за ед. в валюте КП) is {col_8}, should be > 0"
        assert col_9 is not None and col_9 > 0, \
            f"Column 9 (Логистика за ед.) is {col_9}, should be > 0"
        assert col_10 is not None and len(str(col_10)) > 0, \
            f"Column 10 (КОД ТН ВЭД) is empty or None"
        assert col_11 is not None and col_11 >= 0, \
            f"Column 11 (Пошлина %) is {col_11}, should be >= 0"


class TestPDFExports:
    """Validate all 4 PDF exports generate successfully"""

    @pytest.mark.asyncio
    async def test_supply_quote_pdf_generation(self, export_data):
        """Validate supply quote PDF generates with valid size"""
        service = QuotePDFService()
        pdf_bytes = await service.generate_supply_quote_pdf(export_data)

        assert pdf_bytes is not None, "PDF generation returned None"
        assert len(pdf_bytes) > 1000, f"PDF too small ({len(pdf_bytes)} bytes), likely empty"

    @pytest.mark.asyncio
    async def test_supply_letter_pdf_generation(self, export_data):
        """Validate supply letter PDF generates with valid size"""
        service = QuotePDFService()
        pdf_bytes = await service.generate_supply_letter_pdf(export_data)

        assert pdf_bytes is not None, "PDF generation returned None"
        assert len(pdf_bytes) > 1000, f"PDF too small ({len(pdf_bytes)} bytes), likely empty"

    @pytest.mark.asyncio
    async def test_openbook_quote_pdf_generation(self, export_data):
        """Validate openbook quote PDF generates with valid size"""
        service = QuotePDFService()
        pdf_bytes = await service.generate_openbook_quote_pdf(export_data)

        assert pdf_bytes is not None, "PDF generation returned None"
        assert len(pdf_bytes) > 1000, f"PDF too small ({len(pdf_bytes)} bytes), likely empty"

    @pytest.mark.asyncio
    async def test_openbook_letter_pdf_generation(self, export_data):
        """Validate openbook letter PDF generates with valid size"""
        service = QuotePDFService()
        pdf_bytes = await service.generate_openbook_letter_pdf(export_data)

        assert pdf_bytes is not None, "PDF generation returned None"
        assert len(pdf_bytes) > 1000, f"PDF too small ({len(pdf_bytes)} bytes), likely empty"


class TestExcelExports:
    """Validate all 3 Excel exports generate successfully"""

    @pytest.mark.asyncio
    async def test_supply_grid_excel_generation(self, export_data):
        """Validate supply grid Excel generates with valid size"""
        service = QuoteExcelService()
        excel_bytes = await service.generate_supply_grid_excel(export_data)

        assert excel_bytes is not None, "Excel generation returned None"
        assert len(excel_bytes) > 1000, f"Excel too small ({len(excel_bytes)} bytes), likely empty"

    @pytest.mark.asyncio
    async def test_openbook_grid_excel_generation(self, export_data):
        """Validate openbook grid Excel generates with valid size"""
        service = QuoteExcelService()
        excel_bytes = await service.generate_openbook_grid_excel(export_data)

        assert excel_bytes is not None, "Excel generation returned None"
        assert len(excel_bytes) > 1000, f"Excel too small ({len(excel_bytes)} bytes), likely empty"

    @pytest.mark.asyncio
    async def test_openbook_detailed_excel_generation(self, export_data):
        """Validate openbook detailed Excel generates with valid size"""
        service = QuoteExcelService()
        excel_bytes = await service.generate_openbook_detailed_excel(export_data)

        assert excel_bytes is not None, "Excel generation returned None"
        assert len(excel_bytes) > 1000, f"Excel too small ({len(excel_bytes)} bytes), likely empty"


class TestDataConsistency:
    """Validate data consistency across export formats"""

    @pytest.mark.asyncio
    async def test_export_data_has_required_fields(self, export_data):
        """Validate export data contains all required fields"""
        assert export_data is not None, "Export data is None"
        assert "quote" in export_data, "Missing 'quote' field"
        assert "items" in export_data, "Missing 'items' field"
        assert "customer" in export_data, "Missing 'customer' field"
        assert "organization" in export_data, "Missing 'organization' field"
        assert "variables" in export_data, "Missing 'variables' field"
        assert "calculations" in export_data, "Missing 'calculations' field"

        # Validate items have calculation results
        assert len(export_data["items"]) > 0, "No items in export data"
        first_item = export_data["items"][0]
        assert "calculation_results" in first_item, "Items missing 'calculation_results'"


if __name__ == "__main__":
    # Run tests with verbose output
    pytest.main([__file__, "-v", "-s"])
