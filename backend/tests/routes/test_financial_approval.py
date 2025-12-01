"""
Unit tests for Financial Approval API endpoints

Tests:
- GET /api/quotes/{id}/financial-review - Excel download
- POST /api/quotes/{id}/approve - Approve quote
- POST /api/quotes/{id}/send-back - Send back with comments
"""
import pytest
from decimal import Decimal
from io import BytesIO
from openpyxl import load_workbook


def test_financial_review_excel_generation():
    """Test Excel generation with validations"""
    from services.financial_review_export import create_financial_review_excel

    # Test data with low markup (should trigger validation)
    quote_data = {
        'quote_number': 'KP-TEST-001',
        'customer_name': 'Test Customer LLC',
        'total_amount': Decimal('10000.00'),

        # Payment terms that trigger validation
        'advance_from_client': Decimal('50.0'),  # 50% advance
        'delivery_time': 60,  # 60 days → requires 15% markup

        # Low markup (should be red)
        'markup': Decimal('3.0'),  # TOO LOW!

        # DM fee > margin (should be red)
        'dm_fee_value': Decimal('600.00'),
        'total_margin': Decimal('500.00'),  # DM fee exceeds margin!

        # VAT removed
        'vat_removed': True,

        # Other required fields
        'seller_company': 'Master Bearing LLC',
        'offer_sale_type': 'FCA',
        'currency_of_quote': 'RUB',
        'advance_to_supplier': Decimal('30.0'),
        'time_to_advance': 10,
        'total_logistics': Decimal('1000.00'),
        'total_cogs': Decimal('8000.00'),
        'total_revenue_no_vat': Decimal('9200.00'),
        'total_revenue_with_vat': Decimal('10856.00'),

        # Products
        'products': [{
            'name': 'Test Product',
            'quantity': 10,
            'markup': Decimal('2.0'),  # Also too low
            'cogs': Decimal('800.00'),
            'price_no_vat': Decimal('816.00'),
            'price_with_vat': Decimal('962.88')
        }]
    }

    # Generate Excel
    workbook = create_financial_review_excel(quote_data)

    # Verify workbook created
    assert workbook is not None
    assert 'Review' in workbook.sheetnames

    ws = workbook['Review']

    # Verify header (Russian)
    assert ws['A1'].value == 'ФИНАНСОВЫЙ АНАЛИЗ КП'

    # Verify quote info
    assert ws['B3'].value == 'KP-TEST-001'
    assert ws['E3'].value == 'Test Customer LLC'  # Customer name is in E3, not B4

    # Find the totals row by searching for 'ИТОГО ПО КП' header
    totals_row = None
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if cell.value and 'ИТОГО ПО КП' in str(cell.value):
                totals_row = row_idx + 2  # Values are 2 rows below the header
                break
        if totals_row:
            break

    assert totals_row is not None, "Totals row not found in Excel"

    # Verify markup cell (column E) is red and has validation comment
    markup_cell = ws[f'E{totals_row}']
    assert markup_cell.value == 3.0, f"Expected markup 3.0, got {markup_cell.value}"
    # Should be highlighted red
    assert markup_cell.fill.start_color.rgb == '00FFCCCC', f"Markup cell not highlighted red: {markup_cell.fill.start_color.rgb}"
    # Should have comment explaining issue
    assert markup_cell.comment is not None, "Markup cell has no comment"
    assert 'требуемого порога' in markup_cell.comment.text or 'below' in markup_cell.comment.text.lower()

    # Verify DM fee cell (column M) is red
    dm_fee_cell = ws[f'M{totals_row}']
    assert dm_fee_cell.value == 600.0, f"Expected DM fee 600.0, got {dm_fee_cell.value}"
    # Should be red (DM fee > margin)
    assert dm_fee_cell.fill.start_color.rgb == '00FFCCCC', f"DM fee cell not highlighted red: {dm_fee_cell.fill.start_color.rgb}"
    assert dm_fee_cell.comment is not None, "DM fee cell has no comment"
    assert 'превышает' in dm_fee_cell.comment.text or 'exceeds' in dm_fee_cell.comment.text.lower()

    print("✅ All validations working in Excel")


def test_markup_calculation_logic():
    """Test required markup calculation"""
    from services.financial_review_export import calculate_required_markup

    # Test case 1: advance=50%, delivery=60 days → 15%
    result = calculate_required_markup(Decimal('50.0'), 60)
    assert result == Decimal('15.0'), f"Expected 15%, got {result}%"

    # Test case 2: advance=30%, delivery=90 days → 20%
    result = calculate_required_markup(Decimal('30.0'), 90)
    assert result == Decimal('20.0'), f"Expected 20%, got {result}%"

    # Test case 3: advance=70%, delivery=60 days → 5% (no penalty)
    result = calculate_required_markup(Decimal('70.0'), 60)
    assert result == Decimal('5.0'), f"Expected 5%, got {result}%"

    print("✅ Markup calculation logic correct")


def test_dm_fee_validation():
    """Test DM fee vs margin validation"""
    from services.financial_review_export import validate_dm_fee

    # Test case 1: DM fee > margin (FAIL)
    is_valid, msg = validate_dm_fee(Decimal('600.00'), Decimal('500.00'))
    assert not is_valid
    assert 'exceeds deal margin' in msg

    # Test case 2: DM fee < margin (PASS)
    is_valid, msg = validate_dm_fee(Decimal('400.00'), Decimal('500.00'))
    assert is_valid
    assert msg == ''

    # Test case 3: DM fee = margin (PASS, boundary)
    is_valid, msg = validate_dm_fee(Decimal('500.00'), Decimal('500.00'))
    assert is_valid

    print("✅ DM fee validation working")


def test_markup_validation():
    """Test markup validation logic"""
    from services.financial_review_export import validate_markup

    # Test case 1: 3% markup with 50% advance, 60 days (FAIL - need 15%)
    is_valid, msg = validate_markup(
        markup=Decimal('3.0'),
        advance_percent=Decimal('50.0'),
        delivery_time=60,
        level='quote'
    )
    assert not is_valid
    assert '15.00%' in msg  # Required threshold

    # Test case 2: 16% markup with same conditions (PASS)
    is_valid, msg = validate_markup(
        markup=Decimal('16.0'),
        advance_percent=Decimal('50.0'),
        delivery_time=60,
        level='quote'
    )
    assert is_valid
    assert msg == ''

    # Test case 3: 6% markup with 70% advance (PASS - only need 5%)
    is_valid, msg = validate_markup(
        markup=Decimal('6.0'),
        advance_percent=Decimal('70.0'),
        delivery_time=60,
        level='product'
    )
    assert is_valid

    print("✅ Markup validation working")


if __name__ == '__main__':
    """Run tests directly with python3"""
    print("Running Financial Approval Tests...\n")

    test_markup_calculation_logic()
    test_dm_fee_validation()
    test_markup_validation()
    test_financial_review_excel_generation()

    print("\n✅ All tests passed!")
