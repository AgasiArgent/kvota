"""
Tests for Financial Review Excel Export Service
Following TDD approach: Write failing test → Implement → Test passes
"""
import pytest
from decimal import Decimal
from services.financial_review_export import create_financial_review_excel


def test_creates_excel_with_quote_data():
    """Test basic Excel creation with quote data"""
    quote_data = {
        'quote_number': 'KP-2025-001',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'markup': Decimal('15.0'),
        'rate_forex_risk': Decimal('3.0'),
        'dm_fee_value': Decimal('1000.00'),
        'products': []
    }

    workbook = create_financial_review_excel(quote_data)

    assert workbook is not None
    assert 'Review' in workbook.sheetnames
    sheet = workbook['Review']
    assert sheet['A1'].value == 'Financial Review'


def test_vat_summary_all_products_removed():
    """Test VAT summary when all products have VAT removed (K16 ≠ N16)"""
    quote_data = {
        'quote_number': 'КП25-TEST',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'products': [
            {
                'name': 'Product 1',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('1000.00'),  # K16
                'calc_n16_price_without_vat': Decimal('847.46'),  # N16 (VAT removed)
                'quantity': 10,
            },
            {
                'name': 'Product 2',
                'supplier_country': 'EU',
                'base_price_vat': Decimal('500.00'),  # K16
                'calc_n16_price_without_vat': Decimal('423.73'),  # N16 (VAT removed)
                'quantity': 5,
            },
        ]
    }

    wb = create_financial_review_excel(quote_data)
    ws = wb.active

    # Find VAT summary row (search for "НДС очищен на:")
    vat_summary_found = False
    for row in ws.iter_rows(min_row=1, max_row=30):
        if row[0].value == 'НДС очищен на:':
            assert row[1].value == '2 из 2 продуктов'
            vat_summary_found = True
            break

    assert vat_summary_found, "VAT summary row not found"


def test_vat_summary_partial_removal():
    """Test VAT summary when only some products have VAT removed"""
    quote_data = {
        'quote_number': 'КП25-TEST',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'products': [
            {
                'name': 'Product 1 Turkey',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('1000.00'),  # K16
                'calc_n16_price_without_vat': Decimal('847.46'),  # N16 (VAT removed)
                'quantity': 10,
            },
            {
                'name': 'Product 2 China',
                'supplier_country': 'CN',
                'base_price_vat': Decimal('500.00'),  # K16
                'calc_n16_price_without_vat': Decimal('500.00'),  # N16 (same - no VAT)
                'quantity': 5,
            },
            {
                'name': 'Product 3 Turkey',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('800.00'),  # K16
                'calc_n16_price_without_vat': Decimal('677.97'),  # N16 (VAT removed)
                'quantity': 8,
            },
        ]
    }

    wb = create_financial_review_excel(quote_data)
    ws = wb.active

    # Find VAT summary row
    vat_summary_found = False
    for row in ws.iter_rows(min_row=1, max_row=30):
        if row[0].value == 'НДС очищен на:':
            assert row[1].value == '2 из 3 продуктов'
            vat_summary_found = True
            break

    assert vat_summary_found, "VAT summary row not found"


def test_vat_summary_no_removal():
    """Test VAT summary when no products have VAT removed (all China)"""
    quote_data = {
        'quote_number': 'КП25-TEST',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'products': [
            {
                'name': 'Product 1 China',
                'supplier_country': 'CN',
                'base_price_vat': Decimal('1000.00'),  # K16
                'calc_n16_price_without_vat': Decimal('1000.00'),  # N16 (same)
                'quantity': 10,
            },
            {
                'name': 'Product 2 China',
                'supplier_country': 'CN',
                'base_price_vat': Decimal('500.00'),  # K16
                'calc_n16_price_without_vat': Decimal('500.00'),  # N16 (same)
                'quantity': 5,
            },
        ]
    }

    wb = create_financial_review_excel(quote_data)
    ws = wb.active

    # Find VAT summary row
    vat_summary_found = False
    for row in ws.iter_rows(min_row=1, max_row=30):
        if row[0].value == 'НДС очищен на:':
            assert row[1].value == '0 из 2 продуктов'
            vat_summary_found = True
            break

    assert vat_summary_found, "VAT summary row not found"


def test_vat_summary_calculation_logic():
    """Verify VAT removal detection uses correct formula: K16 != N16"""
    quote_data = {
        'quote_number': 'КП25-TEST',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'products': [
            {
                'name': 'Product 1 - Tiny Difference',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('1000.00'),      # K16
                'calc_n16_price_without_vat': Decimal('999.99'),  # N16 (diff = 0.01 - should count)
                'quantity': 10,
            },
            {
                'name': 'Product 2 - No Difference',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('500.00'),       # K16
                'calc_n16_price_without_vat': Decimal('500.00'),  # N16 (diff = 0 - should NOT count)
                'quantity': 5,
            },
        ]
    }

    wb = create_financial_review_excel(quote_data)
    ws = wb.active

    vat_summary_found = False
    for row in ws.iter_rows(min_row=1, max_row=30):
        if row[0].value == 'НДС очищен на:':
            assert row[1].value == '1 из 2 продуктов'  # Only Product 1 counted
            vat_summary_found = True
            break

    assert vat_summary_found, "VAT summary row not found"


def test_product_table_has_vat_columns():
    """Test that product table includes supplier country and VAT comparison columns"""
    quote_data = {
        'quote_number': 'КП25-TEST',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'products': [
            {
                'name': 'Product 1',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('1000.00'),
                'calc_n16_price_without_vat': Decimal('847.46'),
                'quantity': 10,
            }
        ]
    }

    wb = create_financial_review_excel(quote_data)
    ws = wb.active

    # Find product table header row (search for "ТОВАРЫ")
    header_row_num = None
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=30), start=1):
        if row[0].value == 'ТОВАРЫ':
            header_row_num = row_num + 1  # Headers are next row
            break

    assert header_row_num is not None, "Product table header not found"

    # Check column headers (note: B is merged with C for "Наименование")
    assert ws[f'D{header_row_num}'].value == 'Страна'
    assert ws[f'E{header_row_num}'].value == 'Цена с НДС\n(K16)'
    assert ws[f'F{header_row_num}'].value == 'Цена без НДС\n(N16)'
    assert ws[f'G{header_row_num}'].value == 'Кол-во'  # Shifted from D


def test_product_highlighting_when_vat_removed():
    """Test yellow highlighting on products where K16 ≠ N16"""
    quote_data = {
        'quote_number': 'КП25-TEST',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'products': [
            {
                'name': 'Product 1 Turkey (VAT removed)',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('1000.00'),  # K16
                'calc_n16_price_without_vat': Decimal('847.46'),  # N16 - different!
                'quantity': 10,
            },
            {
                'name': 'Product 2 China (no VAT)',
                'supplier_country': 'CN',
                'base_price_vat': Decimal('500.00'),  # K16
                'calc_n16_price_without_vat': Decimal('500.00'),  # N16 - same
                'quantity': 5,
            }
        ]
    }

    wb = create_financial_review_excel(quote_data)
    ws = wb.active

    # Find first product data row
    product_row_start = None
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=35), start=1):
        if row[0].value == 1:  # Row number column
            product_row_start = row_num
            break

    assert product_row_start is not None, "Product data rows not found"

    # Product 1 (Turkey): Should have yellow highlighting on E and F
    from openpyxl.styles import PatternFill
    YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    row1 = product_row_start
    assert ws[f'E{row1}'].fill == YELLOW_FILL, "Column E (K16) should be highlighted for Product 1"
    assert ws[f'F{row1}'].fill == YELLOW_FILL, "Column F (N16) should be highlighted for Product 1"

    # Product 2 (China): Should NOT have highlighting
    row2 = product_row_start + 1
    assert ws[f'E{row2}'].fill.start_color.rgb != 'FFFFCC', "Column E should NOT be highlighted for Product 2"
    assert ws[f'F{row2}'].fill.start_color.rgb != 'FFFFCC', "Column F should NOT be highlighted for Product 2"


def test_supplier_country_displayed():
    """Test that supplier country is shown in column D"""
    quote_data = {
        'quote_number': 'КП25-TEST',
        'customer_name': 'Test Customer',
        'total_amount': Decimal('10000.00'),
        'products': [
            {
                'name': 'Product Turkey',
                'supplier_country': 'TR',
                'base_price_vat': Decimal('1000.00'),
                'calc_n16_price_without_vat': Decimal('847.46'),
                'quantity': 10,
            },
            {
                'name': 'Product China',
                'supplier_country': 'CN',
                'base_price_vat': Decimal('500.00'),
                'calc_n16_price_without_vat': Decimal('500.00'),
                'quantity': 5,
            },
            {
                'name': 'Product EU',
                'supplier_country': 'EU',
                'base_price_vat': Decimal('800.00'),
                'calc_n16_price_without_vat': Decimal('677.97'),
                'quantity': 8,
            }
        ]
    }

    wb = create_financial_review_excel(quote_data)
    ws = wb.active

    # Find first product data row
    product_row_start = None
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=35), start=1):
        if row[0].value == 1:
            product_row_start = row_num
            break

    assert product_row_start is not None

    # Check supplier countries
    assert ws[f'D{product_row_start}'].value == 'TR'
    assert ws[f'D{product_row_start + 1}'].value == 'CN'
    assert ws[f'D{product_row_start + 2}'].value == 'EU'
