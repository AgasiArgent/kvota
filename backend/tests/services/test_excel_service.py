"""
Tests for Excel Export Service

Tests cover:
- Validation export generation
- Grid export generation (2 sheets)
- Russian number formatting
- Currency symbols and formats
- Excel file integrity
- Column formulas (Sheet 2 Column 7)
"""
import pytest
from decimal import Decimal
from openpyxl import load_workbook
from io import BytesIO
from services.excel_service import QuoteExcelService
from services.export_data_mapper import ExportData


# ============================================================================
# FIXTURES
# ============================================================================

@pytest.fixture
def sample_export_data():
    """Sample export data for testing"""
    return ExportData(
        quote={
            'id': 'quote-123',
            'quote_number': 'КП25-0001',
            'title': 'Test Quote',
            'status': 'draft',
            'created_at': '2025-10-24T10:00:00Z'
        },
        items=[
            {
                'id': 'item-001',
                'brand': 'SKF',
                'sku': '6205',
                'product_name': 'Bearing SKF 6205',
                'quantity': 10,
                'base_price_vat': 1200.00,
                'weight_in_kg': 0.5,
                'customs_code': '8482100000',
                'import_tariff': 5.0,
                'excise_tax': 0,
                'calculation_results': {
                    'purchase_price_no_vat': 1000.00,
                    'purchase_price_total_quote_currency': 95000.00,
                    'logistics_total': 5000.00,
                    'customs_fee': 4750.00,
                    'transit_commission': 0,
                    'util_fee': 0,
                    'sales_price_per_unit': 120.85,
                    'sales_price_total_no_vat': 1208.50,
                    'vat_amount': 241.70,
                    'sales_price_per_unit_with_vat': 145.02,
                    'sales_price_total_with_vat': 1450.20,
                },
            },
            {
                'id': 'item-002',
                'brand': 'FAG',
                'sku': '6206',
                'product_name': 'Bearing FAG 6206',
                'quantity': 5,
                'base_price_vat': 1500.00,
                'weight_in_kg': 0.8,
                'customs_code': '8482100000',
                'import_tariff': 5.0,
                'excise_tax': 0,
                'calculation_results': {
                    'purchase_price_no_vat': 1250.00,
                    'purchase_price_total_quote_currency': 118750.00,
                    'logistics_total': 6250.00,
                    'customs_fee': 5937.50,
                    'transit_commission': 0,
                    'util_fee': 0,
                    'sales_price_per_unit': 151.06,
                    'sales_price_total_no_vat': 755.30,
                    'vat_amount': 151.06,
                    'sales_price_per_unit_with_vat': 181.27,
                    'sales_price_total_with_vat': 906.36,
                },
            }
        ],
        customer={'name': 'ООО "Тестовая компания"', 'inn': '7701234567'},
        contact=None,
        manager=None,
        organization={'name': 'МАСТЕР БЭРИНГ ООО'},
        variables={
            'currency_of_quote': 'RUB',
            'currency_of_base_price': 'USD',
            'exchange_rate_base_price_to_quote': 95.0,
            'seller_company': 'МАСТЕР БЭРИНГ ООО',
            'offer_incoterms': 'DDP',
            'offer_sale_type': 'поставка',
            'delivery_time': 60,
            'advance_from_client': 100,
            'markup': 15.0,
        },
        calculations={
            'total_subtotal': 1963.80,
            'total_with_vat': 2356.56,
            'currency': 'RUB',
        }
    )


# ============================================================================
# TESTS: Russian Number Formatting
# ============================================================================

def test_format_russian_number_with_decimal():
    """Test Russian number formatting with Decimal"""
    result = QuoteExcelService.format_russian_number(Decimal('1234.56'))
    assert result == '1 234,56'


def test_format_russian_number_with_float():
    """Test Russian number formatting with float"""
    result = QuoteExcelService.format_russian_number(1234.56)
    assert result == '1 234,56'


def test_format_russian_number_with_int():
    """Test Russian number formatting with integer"""
    result = QuoteExcelService.format_russian_number(1234)
    assert result == '1 234,00'


def test_format_russian_number_with_none():
    """Test Russian number formatting with None"""
    result = QuoteExcelService.format_russian_number(None)
    assert result == '0,00'


def test_format_russian_number_large():
    """Test Russian number formatting with large numbers"""
    result = QuoteExcelService.format_russian_number(1234567.89)
    assert result == '1 234 567,89'


def test_format_russian_number_zero():
    """Test Russian number formatting with zero"""
    result = QuoteExcelService.format_russian_number(0)
    assert result == '0,00'


# ============================================================================
# TESTS: Validation Export
# ============================================================================

def test_generate_validation_export_basic(sample_export_data):
    """Test basic validation export generation"""
    excel_bytes = QuoteExcelService.generate_validation_export(sample_export_data)

    assert excel_bytes is not None
    assert len(excel_bytes) > 0
    assert excel_bytes[:2] == b'PK'  # ZIP header (Excel files are ZIP archives)


def test_generate_validation_export_structure(sample_export_data):
    """Test validation export has correct structure"""
    excel_bytes = QuoteExcelService.generate_validation_export(sample_export_data)

    # Load workbook
    wb = load_workbook(BytesIO(excel_bytes))

    # Check sheet exists
    assert 'Проверка расчетов' in wb.sheetnames

    ws = wb['Проверка расчетов']

    # Check headers
    assert ws['A1'].value == 'Excel Cell'
    assert ws['B1'].value == 'Название'
    assert ws['C1'].value == 'Значение'

    # Check header styling (openpyxl may add '00' or 'FF' prefix)
    header_color = ws['A1'].fill.start_color.rgb
    assert header_color in ['FF2C5AA0', '002C5AA0']  # Blue background
    assert ws['A1'].font.bold is True
    text_color = ws['A1'].font.color.rgb
    assert text_color in ['FFFFFFFF', '00FFFFFF']  # White text


def test_generate_validation_export_has_sections(sample_export_data):
    """Test validation export has input and output sections"""
    excel_bytes = QuoteExcelService.generate_validation_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb['Проверка расчетов']

    # Find section headers
    section_headers = []
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value and isinstance(cell_value, str):
            if 'ВХОДНЫЕ ДАННЫЕ' in cell_value or 'РАССЧИТАННЫЕ ЗНАЧЕНИЯ' in cell_value:
                section_headers.append(cell_value)

    assert 'ВХОДНЫЕ ДАННЫЕ' in section_headers
    assert 'РАССЧИТАННЫЕ ЗНАЧЕНИЯ' in section_headers


def test_generate_validation_export_has_item_data(sample_export_data):
    """Test validation export includes item-specific data"""
    excel_bytes = QuoteExcelService.generate_validation_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb['Проверка расчетов']

    # Check for SKF item
    found_skf = False
    for row in ws.iter_rows(min_row=1, max_row=100):
        if row[0].value and 'SKF' in str(row[0].value):
            found_skf = True
            break

    assert found_skf, "Should find SKF item in validation export"


def test_generate_validation_export_column_widths(sample_export_data):
    """Test validation export has proper column widths"""
    excel_bytes = QuoteExcelService.generate_validation_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb['Проверка расчетов']

    assert ws.column_dimensions['A'].width == 15
    assert ws.column_dimensions['B'].width == 40
    assert ws.column_dimensions['C'].width == 20


# ============================================================================
# TESTS: Grid Export
# ============================================================================

def test_generate_grid_export_basic(sample_export_data):
    """Test basic grid export generation"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    assert excel_bytes is not None
    assert len(excel_bytes) > 0
    assert excel_bytes[:2] == b'PK'  # ZIP header


def test_generate_grid_export_has_two_sheets(sample_export_data):
    """Test grid export has exactly 2 sheets"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))

    assert len(wb.sheetnames) == 2
    assert 'КП поставка' in wb.sheetnames
    assert 'КП open book' in wb.sheetnames


def test_generate_grid_export_sheet1_structure(sample_export_data):
    """Test Sheet 1 (КП поставка) has correct structure"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws1 = wb['КП поставка']

    # Check headers (9 columns)
    expected_headers = [
        'Бренд', 'Артикул', 'Наименование', 'Кол-во',
        'Цена за ед. (₽)', 'Сумма (₽)', 'НДС (₽)',
        'Цена с НДС за ед. (₽)', 'Сумма с НДС (₽)'
    ]

    for col_idx, expected_header in enumerate(expected_headers, 1):
        actual_header = ws1.cell(row=1, column=col_idx).value
        assert actual_header == expected_header, f"Column {col_idx} header mismatch"

    # Check header styling (openpyxl may add '00' or 'FF' prefix)
    header_color = ws1.cell(row=1, column=1).fill.start_color.rgb
    assert header_color in ['FF2C5AA0', '002C5AA0']  # Blue
    assert ws1.cell(row=1, column=1).font.bold is True
    text_color = ws1.cell(row=1, column=1).font.color.rgb
    assert text_color in ['FFFFFFFF', '00FFFFFF']  # White


def test_generate_grid_export_sheet1_data(sample_export_data):
    """Test Sheet 1 has correct data"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws1 = wb['КП поставка']

    # Check first item (row 2)
    assert ws1.cell(row=2, column=1).value == 'SKF'  # Brand
    assert ws1.cell(row=2, column=2).value == '6205'  # SKU
    assert ws1.cell(row=2, column=3).value == 'Bearing SKF 6205'  # Product name
    assert ws1.cell(row=2, column=4).value == 10  # Quantity
    assert ws1.cell(row=2, column=5).value == 120.85  # Price per unit
    assert ws1.cell(row=2, column=6).value == 1208.50  # Total price
    assert ws1.cell(row=2, column=7).value == 241.70  # VAT

    # Check second item (row 3)
    assert ws1.cell(row=3, column=1).value == 'FAG'


def test_generate_grid_export_sheet1_totals(sample_export_data):
    """Test Sheet 1 has totals row"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws1 = wb['КП поставка']

    # Find ИТОГО row (should be row 4 with 2 items)
    totals_row = 4
    assert ws1.cell(row=totals_row, column=1).value == 'ИТОГО'
    assert ws1.cell(row=totals_row, column=1).font.bold is True

    # Check quantity total
    assert ws1.cell(row=totals_row, column=4).value == 15  # 10 + 5


def test_generate_grid_export_sheet1_number_formats(sample_export_data):
    """Test Sheet 1 has correct number formats"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws1 = wb['КП поставка']

    # Check currency format on monetary columns (5-9)
    for col in range(5, 10):
        cell_format = ws1.cell(row=2, column=col).number_format
        assert '₽' in cell_format or '#,##0.00' in cell_format


def test_generate_grid_export_sheet1_freeze_panes(sample_export_data):
    """Test Sheet 1 has frozen header row"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws1 = wb['КП поставка']

    assert ws1.freeze_panes == 'A2'


def test_generate_grid_export_sheet2_structure(sample_export_data):
    """Test Sheet 2 (КП open book) has correct structure"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws2 = wb['КП open book']

    # Check headers (20 columns)
    expected_headers = [
        'Бренд', 'Артикул', 'Наименование', 'Кол-во', 'Валюта',
        'Цена без НДС', 'Сумма инвойса', 'Цена в валюте КП (₽)',
        'Логистика (₽)', 'Код ТН ВЭД', 'Пошлина (%)', 'Таможенный сбор (₽)',
        'Акциз (₽)', 'Утиль. сбор (₽)', 'Комиссия транзит (₽)',
        'Цена за ед. (₽)', 'Сумма (₽)', 'НДС (₽)',
        'Цена с НДС за ед. (₽)', 'Сумма с НДС (₽)'
    ]

    for col_idx, expected_header in enumerate(expected_headers, 1):
        actual_header = ws2.cell(row=1, column=col_idx).value
        assert actual_header == expected_header, f"Sheet 2 Column {col_idx} header mismatch"


def test_generate_grid_export_sheet2_data(sample_export_data):
    """Test Sheet 2 has correct detailed data"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws2 = wb['КП open book']

    # Check first item (row 2)
    assert ws2.cell(row=2, column=1).value == 'SKF'  # Brand
    assert ws2.cell(row=2, column=2).value == '6205'  # SKU
    assert ws2.cell(row=2, column=5).value == 'USD'  # Currency
    assert ws2.cell(row=2, column=6).value == 1000.00  # Price without VAT
    assert ws2.cell(row=2, column=7).value == 10000.00  # Invoice amount (1000 * 10)
    assert ws2.cell(row=2, column=10).value == '8482100000'  # Customs code
    assert ws2.cell(row=2, column=11).value == 5.0  # Import tariff


def test_generate_grid_export_sheet2_invoice_amount_formula(sample_export_data):
    """Test Sheet 2 Column 7 (Invoice Amount) is calculated correctly"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws2 = wb['КП open book']

    # Row 2: Price 1000 × Quantity 10 = 10,000
    invoice_amount_row2 = ws2.cell(row=2, column=7).value
    assert invoice_amount_row2 == 10000.00

    # Row 3: Price 1250 × Quantity 5 = 6,250
    invoice_amount_row3 = ws2.cell(row=3, column=7).value
    assert invoice_amount_row3 == 6250.00


def test_generate_grid_export_sheet2_number_formats(sample_export_data):
    """Test Sheet 2 has correct number formats for different column types"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws2 = wb['КП open book']

    # Check percentage format on column 11 (Import tariff)
    tariff_format = ws2.cell(row=2, column=11).number_format
    assert '%' in tariff_format or '0.00%' in tariff_format

    # Check currency format on monetary columns
    for col in [6, 7, 8, 9, 12, 14, 15, 16, 17, 18, 19, 20]:
        cell_format = ws2.cell(row=2, column=col).number_format
        assert '₽' in cell_format or '#,##0.00' in cell_format


def test_generate_grid_export_sheet2_freeze_panes(sample_export_data):
    """Test Sheet 2 has frozen header row"""
    excel_bytes = QuoteExcelService.generate_grid_export(sample_export_data)

    wb = load_workbook(BytesIO(excel_bytes))
    ws2 = wb['КП open book']

    assert ws2.freeze_panes == 'A2'


# ============================================================================
# TESTS: Edge Cases
# ============================================================================

def test_generate_validation_export_empty_items():
    """Test validation export with no items"""
    export_data = ExportData(
        quote={'id': 'quote-123', 'quote_number': 'КП25-0001'},
        items=[],
        customer=None,
        organization={'name': 'Test Org'},
        variables={'currency_of_quote': 'RUB'},
        calculations={'total_subtotal': 0, 'total_with_vat': 0}
    )

    excel_bytes = QuoteExcelService.generate_validation_export(export_data)

    assert excel_bytes is not None
    assert len(excel_bytes) > 0

    wb = load_workbook(BytesIO(excel_bytes))
    assert 'Проверка расчетов' in wb.sheetnames


def test_generate_grid_export_empty_items():
    """Test grid export with no items"""
    export_data = ExportData(
        quote={'id': 'quote-123', 'quote_number': 'КП25-0001'},
        items=[],
        customer=None,
        organization={'name': 'Test Org'},
        variables={'currency_of_quote': 'RUB', 'currency_of_base_price': 'USD'},
        calculations={'total_subtotal': 0, 'total_with_vat': 0}
    )

    excel_bytes = QuoteExcelService.generate_grid_export(export_data)

    assert excel_bytes is not None
    wb = load_workbook(BytesIO(excel_bytes))

    # Should still have 2 sheets with headers
    assert len(wb.sheetnames) == 2
    ws1 = wb['КП поставка']
    assert ws1.cell(row=1, column=1).value == 'Бренд'


def test_generate_validation_export_missing_calculation_results():
    """Test validation export when item has no calculation results"""
    export_data = ExportData(
        quote={'id': 'quote-123', 'quote_number': 'КП25-0001'},
        items=[
            {
                'id': 'item-001',
                'brand': 'SKF',
                'sku': '6205',
                'product_name': 'Test Product',
                'quantity': 10,
                'calculation_results': None  # No calculations
            }
        ],
        customer=None,
        organization={'name': 'Test Org'},
        variables={},
        calculations={'total_subtotal': 0, 'total_with_vat': 0}
    )

    # Should not crash
    excel_bytes = QuoteExcelService.generate_validation_export(export_data)
    assert excel_bytes is not None


def test_generate_grid_export_missing_optional_fields():
    """Test grid export when items have missing optional fields"""
    export_data = ExportData(
        quote={'id': 'quote-123', 'quote_number': 'КП25-0001'},
        items=[
            {
                'id': 'item-001',
                'brand': '',  # Missing brand
                'sku': '',  # Missing SKU
                'product_name': 'Test Product',
                'quantity': 1,
                'calculation_results': {
                    'sales_price_per_unit': 100.0,
                    'sales_price_total_no_vat': 100.0,
                    'vat_amount': 20.0,
                    'sales_price_per_unit_with_vat': 120.0,
                    'sales_price_total_with_vat': 120.0,
                }
            }
        ],
        customer={'name': 'Test Customer'},
        organization={'name': 'Test Org'},
        variables={'currency_of_quote': 'RUB', 'currency_of_base_price': 'USD'},
        calculations={'total_subtotal': 100, 'total_with_vat': 120}
    )

    # Should not crash
    excel_bytes = QuoteExcelService.generate_grid_export(export_data)
    assert excel_bytes is not None

    wb = load_workbook(BytesIO(excel_bytes))
    ws1 = wb['КП поставка']

    # Empty fields should be empty strings (openpyxl may return None or '')
    brand_value = ws1.cell(row=2, column=1).value
    sku_value = ws1.cell(row=2, column=2).value
    assert brand_value in ['', None]  # Brand
    assert sku_value in ['', None]  # SKU
